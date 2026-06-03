from datetime import datetime
from decimal import Decimal
import csv
from io import TextIOWrapper

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from openpyxl import load_workbook

from main.models import (
    CurrentResidentRosterEntry,
    FinancialUpload,
    HousingApplication,
    Payment,
    Property,
    PropertyRoomRent,
)
from main.views import canonical_room_label, money, normalized_header, normalized_room_label, unique_headers


SKIP_TENANT_LABELS = {"", "office", "shop", "vacant", "total", "totals"}


def clean_match_value(value):
    return "".join(character.lower() for character in str(value or "") if character.isalnum())


def name_tokens(value):
    return {
        token.lower()
        for token in str(value or "").replace("/", " ").split()
        if token.strip()
    }


def names_are_compatible(left, right):
    left_tokens = name_tokens(left)
    right_tokens = name_tokens(right)

    if not left_tokens or not right_tokens:
        return False

    if left_tokens.issubset(right_tokens) or right_tokens.issubset(left_tokens):
        return True

    left_parts = str(left or "").strip().split()
    right_parts = str(right or "").strip().split()
    if len(left_parts) >= 2 and len(right_parts) >= 2:
        left_first, left_last = left_parts[0].lower(), left_parts[-1].lower()
        right_first, right_last = right_parts[0].lower(), right_parts[-1].lower()
        return left_last == right_last and (
            left_first.startswith(right_first[:3]) or right_first.startswith(left_first[:3])
        )

    return False


def split_name(full_name):
    parts = str(full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def parse_month(value):
    try:
        return datetime.strptime(value, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise CommandError("Use --month in YYYY-MM format, for example --month 2026-01.") from exc


def parse_date(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    for date_format in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue
    return None


def column_lookup(headers):
    aliases = {
        "room": ["room", "room #", "room number", "unit", "unit #", "space"],
        "tenant": ["tenant", "tenant name", "resident", "resident name", "name"],
        "lease_start": ["lease start", "lease start date", "move in", "move-in", "move in date"],
        "monthly_rent": ["monthly rent", "rent"],
        "rent_paid": ["rent paid", "paid rent"],
        "deposit": ["deposit", "deposit held"],
        "new_deposit": ["new deposit", "deposit paid"],
        "utilities": ["shared utilities", "utilities", "utility", "utility paid"],
    }
    normalized_to_header = {normalized_header(header): header for header in headers}
    found = {}
    for key, choices in aliases.items():
        for choice in choices:
            header = normalized_to_header.get(normalized_header(choice))
            if header:
                found[key] = header
                break
    return found


def read_raw_upload_rows(upload, selected_sheet_name=None):
    file_name = upload.file.name.lower()
    upload.file.open("rb")
    try:
        if file_name.endswith(".xlsx"):
            workbook = load_workbook(upload.file, read_only=True, data_only=True)
            sheet_name = selected_sheet_name or workbook.sheetnames[0]
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Sheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
            raw_rows = [
                [cell for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            return sheet_name, raw_rows

        wrapper = TextIOWrapper(upload.file, encoding="utf-8-sig", newline="")
        try:
            return "CSV", list(csv.reader(wrapper))
        except UnicodeDecodeError:
            upload.file.seek(0)
            wrapper = TextIOWrapper(upload.file, encoding="latin-1", newline="")
            return "CSV", list(csv.reader(wrapper))
    finally:
        upload.file.close()


def read_rent_roll_rows(upload, selected_sheet_name=None):
    sheet_name, raw_rows = read_raw_upload_rows(upload, selected_sheet_name)
    non_empty_rows = [
        list(row)
        for row in raw_rows
        if any(str(cell or "").strip() for cell in row)
    ]
    required_columns = ["room", "tenant", "monthly_rent", "rent_paid"]
    for header_index, raw_header_row in enumerate(non_empty_rows):
        headers = unique_headers(raw_header_row)
        columns = column_lookup(headers)
        if all(column in columns for column in required_columns):
            rows = []
            for row_number, raw_row in enumerate(non_empty_rows[header_index + 1:], start=header_index + 2):
                row_data = {}
                for index, header in enumerate(headers):
                    row_data[header] = raw_row[index] if index < len(raw_row) else ""
                rows.append({"row_number": row_number, "data": row_data})
            return sheet_name, headers, rows, columns, header_index + 1

    preview_rows = [" | ".join(str(cell or "").strip() for cell in row[:10]) for row in non_empty_rows[:8]]
    raise CommandError(
        "Missing required column(s): room, tenant, monthly_rent, rent_paid. "
        "Could not find a header row matching Room #, Tenant Name, Monthly Rent, Rent Paid. "
        f"First rows seen: {' / '.join(preview_rows)}"
    )


class Command(BaseCommand):
    help = "Import one monthly rent roll sheet into room settings, roster records, and completed historical payments."

    def add_arguments(self, parser):
        parser.add_argument("--property-name", required=True, help="Exact property name.")
        parser.add_argument("--upload-id", type=int, required=True, help="FinancialUpload ID containing the rent roll.")
        parser.add_argument("--month", required=True, help="Rent roll month in YYYY-MM format.")
        parser.add_argument("--sheet-name", help="Worksheet/tab name. Defaults to first sheet.")
        parser.add_argument("--default-deposit-required", default="450.00")
        parser.add_argument("--payment-method", default="cash", choices=[choice[0] for choice in Payment.PAYMENT_METHOD_CHOICES])
        parser.add_argument("--confirm", action="store_true", help="Write records. Without this flag, only previews.")

    def handle(self, *args, **options):
        property_obj = Property.objects.filter(name=options["property_name"].strip()).first()
        if not property_obj:
            raise CommandError(f"Property not found: {options['property_name']}")

        upload = FinancialUpload.objects.filter(id=options["upload_id"], property=property_obj).first()
        if not upload:
            raise CommandError(f"Upload not found for {property_obj.name}: {options['upload_id']}")

        service_month = parse_month(options["month"])
        default_deposit_required = money(options["default_deposit_required"])
        sheet_name, headers, rows, columns, header_row_number = read_rent_roll_rows(upload, selected_sheet_name=options.get("sheet_name"))

        planned = []
        skipped = []
        for row in rows:
            data = row["data"]
            room = canonical_room_label(data.get(columns["room"]))
            tenant_name = str(data.get(columns["tenant"], "") or "").strip()
            clean_tenant = normalized_header(tenant_name)
            if clean_tenant in SKIP_TENANT_LABELS or not room:
                skipped.append((room or "-", tenant_name or "-", "non-resident row"))
                continue

            monthly_rent = money(data.get(columns["monthly_rent"]))
            rent_paid = money(data.get(columns["rent_paid"]))
            deposit_paid = money(data.get(columns.get("deposit"))) if columns.get("deposit") else Decimal("0.00")
            new_deposit_paid = money(data.get(columns.get("new_deposit"))) if columns.get("new_deposit") else Decimal("0.00")
            utilities_paid = money(data.get(columns.get("utilities"))) if columns.get("utilities") else Decimal("0.00")
            lease_start = parse_date(data.get(columns.get("lease_start"))) if columns.get("lease_start") else None

            planned.append({
                "room": room,
                "tenant_name": tenant_name,
                "lease_start": lease_start,
                "monthly_rent": monthly_rent,
                "rent_paid": rent_paid,
                "deposit_paid": deposit_paid,
                "new_deposit_paid": new_deposit_paid,
                "utilities_paid": utilities_paid,
            })

        self.stdout.write("Monthly rent roll import preview")
        self.stdout.write("===============================")
        self.stdout.write(f"Property: {property_obj.name}")
        self.stdout.write(f"Upload: {upload.id} | {upload.name}")
        self.stdout.write(f"Sheet: {sheet_name}")
        self.stdout.write(f"Header row: {header_row_number}")
        self.stdout.write(f"Month: {service_month.strftime('%B %Y')}")
        self.stdout.write(f"Financial entries attached to upload: {upload.entries.count()}")
        self.stdout.write(f"Resident rows selected: {len(planned)}")

        for room, tenant_name, reason in skipped:
            self.stdout.write(f"SKIP | Room {room} | {tenant_name} | {reason}")

        rent_total = Decimal("0.00")
        utility_total = Decimal("0.00")
        new_deposit_total = Decimal("0.00")
        for item in planned:
            rent_total += item["rent_paid"]
            utility_total += item["utilities_paid"]
            new_deposit_total += item["new_deposit_paid"]
            self.stdout.write(
                f"IMPORT | Room {item['room']} | {item['tenant_name']} | "
                f"rent ${item['monthly_rent']} | rent paid ${item['rent_paid']} | "
                f"utilities paid ${item['utilities_paid']} | deposit held ${item['deposit_paid']}"
            )

        self.stdout.write("")
        self.stdout.write(f"Rent paid total: ${rent_total}")
        self.stdout.write(f"Utilities paid total: ${utility_total}")
        self.stdout.write(f"New deposit paid total: ${new_deposit_total}")

        if not options["confirm"]:
            self.stdout.write(self.style.WARNING("Dry run only. No records were changed."))
            self.stdout.write("Run again with --confirm to import this monthly rent roll.")
            return

        upload.entries.all().delete()
        payments_created = 0
        settings_updated = 0
        roster_updated = 0
        for item in planned:
            room = item["room"]
            deposit_required = max(default_deposit_required, item["deposit_paid"], item["new_deposit_paid"])
            setting, _created = PropertyRoomRent.objects.update_or_create(
                property=property_obj,
                room_unit_label=room,
                defaults={
                    "monthly_rent": item["monthly_rent"],
                    "utility_monthly": item["utilities_paid"],
                    "deposit_required": deposit_required,
                    "deposit_paid": item["deposit_paid"],
                    "is_active": True,
                },
            )
            settings_updated += 1

            first_name, last_name = split_name(item["tenant_name"])
            roster, roster_created = CurrentResidentRosterEntry.objects.get_or_create(
                property=property_obj,
                first_name=first_name,
                last_name=last_name,
                email="",
                room_unit_label=room,
                defaults={"is_active": True},
            )
            if not roster.is_active:
                roster.is_active = True
                roster.save(update_fields=["is_active"])
            if roster_created:
                roster_updated += 1

            application = self.find_or_create_application(property_obj, item, setting, deposit_required)
            payments_created += self.create_payment_if_missing(
                application,
                service_month,
                "rent",
                item["rent_paid"],
                options["payment_method"],
                upload.id,
            )
            payments_created += self.create_payment_if_missing(
                application,
                service_month,
                "utility",
                item["utilities_paid"],
                options["payment_method"],
                upload.id,
            )
            payments_created += self.create_payment_if_missing(
                application,
                service_month,
                "deposit",
                item["new_deposit_paid"],
                options["payment_method"],
                upload.id,
            )

        upload.parsed_at = timezone.now()
        upload.save(update_fields=["parsed_at"])
        self.stdout.write(self.style.SUCCESS("Monthly rent roll import complete."))
        self.stdout.write(f"Room settings updated: {settings_updated}")
        self.stdout.write(f"Roster records created: {roster_updated}")
        self.stdout.write(f"Payments created: {payments_created}")
        self.stdout.write("Financial entries removed from this upload so rent roll detail does not double count T-12 income.")

    def find_or_create_application(self, property_obj, item, setting, deposit_required):
        target_name = clean_match_value(item["tenant_name"])
        target_room = normalized_room_label(item["room"])
        applications = HousingApplication.objects.filter(property=property_obj)
        for application in applications:
            if clean_match_value(application.full_name) == target_name and normalized_room_label(application.space_label) == target_room:
                return application

        room_matches = [
            application
            for application in applications
            if normalized_room_label(application.space_label) == target_room
            and names_are_compatible(application.full_name, item["tenant_name"])
        ]
        if room_matches:
            return sorted(
                room_matches,
                key=lambda application: (
                    application.user_id is None,
                    -application.payments.count(),
                    application.id,
                ),
            )[0]

        return HousingApplication.objects.create(
            property=property_obj,
            full_name=item["tenant_name"],
            phone="",
            email="",
            age=0,
            space_type="Room",
            space_label=item["room"],
            monthly_rent=item["monthly_rent"],
            balance=Decimal("0.00"),
            rent_due_day=1,
            lease_start_date=item["lease_start"],
            deposit_required=deposit_required,
            deposit_paid=item["deposit_paid"],
            utility_monthly=setting.utility_monthly,
            utility_balance=Decimal("0.00"),
            income_source="Historical rent roll import",
            monthly_income=Decimal("0.00"),
            housing_need="Historical rent roll resident record.",
            additional_notes="Historical rent roll import record.",
        )

    def create_payment_if_missing(self, application, service_month, payment_type, amount, payment_method, upload_id):
        amount = money(amount)
        if amount <= 0:
            return 0

        existing_total = Decimal("0.00")
        for payment in application.payments.filter(payment_type=payment_type, status="completed", service_month=service_month):
            existing_total += payment.amount

        missing_amount = max(amount - existing_total, Decimal("0.00"))
        if missing_amount <= 0:
            return 0

        Payment.objects.create(
            application=application,
            payment_type=payment_type,
            payment_method=payment_method,
            description=f"Historical {service_month.strftime('%B %Y')} {payment_type} payment",
            notes=f"Historical rent roll imported from upload {upload_id}.",
            received_at=timezone.make_aware(datetime.combine(service_month, datetime.min.time())),
            service_month=service_month,
            months_covered=1,
            amount=missing_amount,
            status="completed",
        )
        return 1
