from collections import OrderedDict
import calendar
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
import csv
import html
from html.parser import HTMLParser
from io import TextIOWrapper
import base64
import json
import re
import secrets
import unicodedata
from urllib.parse import urlencode
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import stripe
from openpyxl import load_workbook

from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.management import call_command
from django.core.mail import EmailMessage, send_mail
from django.db.models import Count, Max, Q, Sum
from django.http import Http404, JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.html import strip_tags
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from .forms import (
    InviteCodeForm,
    BlogCommentForm,
    HousingApplicationForm,
    FinancialUploadForm,
    AccountingReceiptForm,
    SignUpForm,
    ManualPaymentForm,
    ResidentBalanceCorrectionForm,
    CustomReportForm,
    ResidentProfilePhotoForm,
    ReplacementInviteCodeForm,
    PropertyOwnerIntakeForm,
    PropertyOwnerLeadPipelineForm,
    LandlordSignUpForm,
    ExistingResidentIntakeForm,
    CurrentResidentRosterUploadForm,
    ResidentRoomTransferForm,
    GroupResidentMessageForm,
    CompanyEmailComposeForm,
    CompanyEmailReplyForm,
    ScreeningReviewForm,
    AdverseActionNoticeForm,
    RentalListingForm,
    RentalListingChannelForm,
)

from .models import (
    User,
    Property,
    BlogPost,
    HousingApplication,
    RentHistory,
    Payment,
    FinancialUpload,
    FinancialEntry,
    AccountingReceipt,
    ExpenseCategory,
    ResidentMessage,
    ResidentMessageReply,
    SmsMessageLog,
    ApplicantDocument,
    SignedDocument,
    PropertyOwnerIntake,
    ExistingResidentIntake,
    CurrentResidentRosterEntry,
    PropertyRoomRent,
    PropertyUtilityVendor,
    ResidentUtilitySetup,
    LandlordIntake,
    CompanyMailboxConnection,
    AdverseActionNotice,
    RentalListing,
    RentalListingPhoto,
    RentalListingChannel,
)
from .invite_utils import create_pending_portal_user, send_portal_access_invite_email
from .receipt_ocr import process_receipt_ocr
from .templatetags.formatting import phone_format

stripe.api_key = settings.STRIPE_SECRET_KEY

LATE_FEE_AMOUNT = Decimal("25.00")
T12_INCOME_PAYMENT_TYPES = ["rent", "utility", "late_fee", "application_fee", "background_check_fee", "other"]


def notify_resident_of_portal_reply(request, resident_message):
    application = resident_message.application

    if not application.email:
        return False

    dashboard_url = request.build_absolute_uri(reverse("resident_requests"))
    send_mail(
        f"New secure portal reply: {resident_message.subject}",
        f"""Hello {application.full_name},

You have a new secure reply in your Bowling Legacy resident portal.

Use this direct link to open My Requests. If you are not already signed in, the site will ask for your login first and then take you to the message area:
{dashboard_url}

For privacy, the reply content is stored inside your portal rather than in this email.

Thank you,
Bowling Legacy Housing
""",
        getattr(settings, "DEFAULT_FROM_EMAIL", None),
        [application.email],
        fail_silently=False,
    )
    return True


def notify_resident_of_portal_reply_sms(request, resident_message):
    portal_url = request.build_absolute_uri(reverse("resident_requests"))
    body = (
        "Bowling Legacy: You have a new secure portal reply. "
        f"Log in to view it: {portal_url} Reply STOP to opt out."
    )
    return send_sms_message(
        resident_message.application,
        body[:1500],
        request.user,
        resident_message=resident_message,
    )


def send_resident_portal_notification_email(request, application, subject, message, target_view_name):
    if not application.email:
        return False

    target_url = request.build_absolute_uri(reverse(target_view_name))
    send_mail(
        subject,
        f"""Hello {application.full_name},

{message}

Use this direct link. If you are not already signed in, the site will ask for your login first:
{target_url}

Thank you,
Bowling Legacy Housing
""",
        getattr(settings, "DEFAULT_FROM_EMAIL", None),
        [application.email],
        fail_silently=False,
    )
    return True


def money(value):
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    try:
        cleaned = str(value).replace("$", "").replace(",", "").strip()
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def csv_money(value):
    return f"{money(value):.2f}"


FINANCIAL_COLUMN_ALIASES = {
    "entry_date": ["date", "transaction date", "posted date", "posting date", "paid date", "receipt date"],
    "description": ["description", "memo", "payee", "vendor", "name", "transaction", "details"],
    "amount": ["amount", "debit", "credit", "paid", "payment", "total", "net amount"],
    "category": ["category", "account", "expense category", "income category", "gl account", "class", "overhead title", "overhead", "account title"],
    "entry_type": ["type", "entry type", "transaction type", "account type"],
    "property_name": ["property", "property name", "building", "location"],
}

FINANCIAL_TYPE_KEYWORDS = {
    "income": ["income", "rent", "rent income", "deposit", "revenue", "payment received"],
    "debt_service": ["mortgage", "loan", "debt", "principal", "interest"],
    "capital_expense": ["capital", "capex", "improvement", "renovation", "appliance", "roof"],
    "operating_expense": ["expense", "repair", "maintenance", "utility", "insurance", "tax", "supplies", "cleaning"],
}


def normalized_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def unique_headers(raw_headers):
    headers = []
    used = {}
    for index, header in enumerate(raw_headers, start=1):
        base = str(header or f"Column {index}").strip() or f"Column {index}"
        count = used.get(base, 0) + 1
        used[base] = count
        headers.append(base if count == 1 else f"{base} {count}")
    return headers


def guess_financial_columns(headers):
    normalized = {header: normalized_header(header) for header in headers}
    guesses = {}
    for field_name, aliases in FINANCIAL_COLUMN_ALIASES.items():
        for header, clean_header in normalized.items():
            if clean_header in aliases or any(alias in clean_header for alias in aliases):
                guesses[field_name] = header
                break
    return guesses


def financial_upload_sheet_names(upload):
    file_name = upload.file.name.lower()
    if not file_name.endswith(".xlsx"):
        return ["CSV"]

    upload.file.open("rb")
    try:
        workbook = load_workbook(upload.file, read_only=True, data_only=True)
        return workbook.sheetnames
    finally:
        upload.file.close()


def read_financial_upload_rows(upload, limit=None, selected_sheet_name=None):
    file_name = upload.file.name.lower()
    upload.file.open("rb")
    try:
        if file_name.endswith(".xlsx"):
            workbook = load_workbook(upload.file, read_only=True, data_only=True)
            worksheet = workbook[selected_sheet_name] if selected_sheet_name in workbook.sheetnames else workbook.active
            raw_rows = list(worksheet.iter_rows(values_only=True))
            sheet_name = worksheet.title
        else:
            wrapper = TextIOWrapper(upload.file, encoding="utf-8-sig", newline="")
            try:
                raw_rows = list(csv.reader(wrapper))
            except UnicodeDecodeError:
                upload.file.seek(0)
                wrapper = TextIOWrapper(upload.file, encoding="latin-1", newline="")
                raw_rows = list(csv.reader(wrapper))
            sheet_name = "CSV"
    finally:
        upload.file.close()

    non_empty_rows = [
        list(row)
        for row in raw_rows
        if any(str(cell or "").strip() for cell in row)
    ]
    if not non_empty_rows:
        return sheet_name, [], []

    headers = unique_headers(non_empty_rows[0])
    data_rows = non_empty_rows[1:]
    if limit:
        data_rows = data_rows[:limit]

    rows = []
    for row_number, row in enumerate(data_rows, start=2):
        row_data = {}
        for index, header in enumerate(headers):
            row_data[header] = row[index] if index < len(row) else ""
        rows.append({"row_number": row_number, "data": row_data})

    return sheet_name, headers, rows


def create_financial_entry_from_import(upload, property_obj, sheet_name, row, entry_date, entry_type, category, description, amount):
    if amount == Decimal("0.00"):
        return None

    if entry_type != "income" and category:
        ExpenseCategory.objects.get_or_create(
            name=category,
            defaults={
                "entry_type": entry_type if entry_type in dict(ExpenseCategory.ENTRY_TYPE_CHOICES) else "other",
                "created_by": None,
            },
        )

    property_name = property_obj.name if property_obj else ""
    normalized_amount = normalized_import_amount(amount, entry_type)
    duplicate = FinancialEntry.objects.filter(
        ledger_scope=upload.ledger_scope,
        property_name=property_name,
        entry_date=entry_date,
        month=entry_date.month if entry_date else None,
        year=entry_date.year if entry_date else None,
        entry_type=entry_type,
        category=category,
        description=description,
        amount=normalized_amount,
    ).exclude(upload=upload).exists()
    if duplicate:
        return None

    return FinancialEntry.objects.create(
        upload=upload,
        ledger_scope=upload.ledger_scope,
        property_name=property_name,
        sheet_name=sheet_name,
        row_number=row["row_number"],
        entry_date=entry_date,
        month=entry_date.month if entry_date else None,
        year=entry_date.year if entry_date else None,
        entry_type=entry_type,
        category=category,
        description=description,
        amount=normalized_amount,
    )


def duplicate_receipt_financial_entry_exists(receipt):
    if not receipt.category:
        return False

    return FinancialEntry.objects.filter(
        ledger_scope="property",
        property_name=receipt.property.name,
        entry_date=receipt.receipt_date,
        month=receipt.receipt_date.month if receipt.receipt_date else None,
        year=receipt.receipt_date.year if receipt.receipt_date else None,
        entry_type=receipt.entry_type,
        category=receipt.category.name,
        description=receipt.description or receipt.vendor,
        amount=receipt.amount,
    ).exists()


MONTH_NAME_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


SUMMARY_TOTAL_KEYWORDS = {
    "total",
    "subtotal",
    "ytd",
    "year to date",
    "quarter",
    "q1",
    "q2",
    "q3",
    "q4",
}
SUMMARY_CALCULATED_ROW_KEYWORDS = {
    "noi",
    "net operating income",
    "net income",
    "cash flow",
    "cashflow",
    "after debt service",
    "net after debt service",
    "total net after debt service",
    "profit",
    "loss",
}
UTILITY_PARENT_CATEGORIES = {"utility", "utilities"}
UTILITY_DETAIL_KEYWORDS = {"power", "electric", "electricity", "gas", "water", "sewer", "trash", "garbage", "internet", "utilities"}
SUMMARY_OPERATING_EXPENSE_LABELS = {
    "expense",
    "expenses",
    "operating expense",
    "operating expenses",
    "total expense",
    "total expenses",
    "total operating expense",
    "total operating expenses",
}
SUMMARY_DEBT_SERVICE_LABELS = {
    "debt service",
    "total debt service",
    "mortgage",
    "mortgage payment",
    "loan payment",
}


def is_summary_total_label(value):
    clean_value = normalized_header(value)
    return any(keyword in clean_value for keyword in SUMMARY_TOTAL_KEYWORDS)


def summary_category_entry_type(category, default_entry_type):
    clean_category = normalized_header(category)
    if not clean_category:
        return None

    if clean_category in SUMMARY_OPERATING_EXPENSE_LABELS:
        return "operating_expense"

    if clean_category in SUMMARY_DEBT_SERVICE_LABELS or "debt service" in clean_category:
        return "debt_service"

    if any(keyword in clean_category for keyword in SUMMARY_CALCULATED_ROW_KEYWORDS):
        return None

    return normalize_entry_type("", category, f"{category} summary", Decimal("1.00"), default_entry_type)


def summary_has_detail_rows(category, all_categories):
    clean_category = normalized_header(category)
    if clean_category not in SUMMARY_OPERATING_EXPENSE_LABELS:
        return False

    for other_category in all_categories:
        clean_other = normalized_header(other_category)
        if (
            clean_other
            and clean_other != clean_category
            and clean_other not in SUMMARY_OPERATING_EXPENSE_LABELS
            and clean_other not in SUMMARY_DEBT_SERVICE_LABELS
            and not any(keyword in clean_other for keyword in SUMMARY_CALCULATED_ROW_KEYWORDS)
            and summary_category_entry_type(clean_other, "operating_expense") == "operating_expense"
        ):
            return True

    return False


def should_skip_summary_category(category, all_categories):
    clean_category = normalized_header(category)
    if not clean_category:
        return True

    if any(keyword in clean_category for keyword in SUMMARY_CALCULATED_ROW_KEYWORDS):
        return True

    if is_summary_total_label(clean_category):
        entry_type = summary_category_entry_type(clean_category, "operating_expense")
        if not entry_type:
            return True
        if summary_has_detail_rows(clean_category, all_categories):
            return True
        return False

    if clean_category in UTILITY_PARENT_CATEGORIES:
        detail_categories = [
            normalized_header(other_category)
            for other_category in all_categories
            if normalized_header(other_category) != clean_category
        ]
        has_utility_details = any(
            any(keyword in detail_category for keyword in UTILITY_DETAIL_KEYWORDS)
            for detail_category in detail_categories
        )
        if has_utility_details:
            return True

    return False


def parse_month_header(value):
    clean_value = normalized_header(value)
    if not clean_value or is_summary_total_label(clean_value):
        return None

    for token in clean_value.replace("-", " ").replace("/", " ").split():
        if token in MONTH_NAME_MAP:
            return MONTH_NAME_MAP[token]

    return MONTH_NAME_MAP.get(clean_value[:3])


def summary_month_header_options(headers):
    return [
        {
            "name": header,
            "is_month": parse_month_header(header) is not None,
        }
        for header in headers
    ]


def parse_import_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, date):
        return value

    value = str(value).strip()
    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return timezone.datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return None


def normalize_entry_type(value, category, description, amount, default_entry_type):
    clean_value = normalized_header(value)
    for entry_type, keywords in FINANCIAL_TYPE_KEYWORDS.items():
        if clean_value == entry_type or clean_value in keywords:
            return entry_type

    combined = normalized_header(f"{category} {description}")
    if "rent" in combined and not any(expense_word in combined for expense_word in ["equipment rental", "rental equipment", "rent expense"]):
        return "income"

    for entry_type, keywords in FINANCIAL_TYPE_KEYWORDS.items():
        if any(keyword in combined for keyword in keywords):
            return entry_type

    if amount < 0:
        return "operating_expense"

    return default_entry_type or "operating_expense"


def normalized_import_amount(amount, entry_type):
    amount = money(amount)
    if entry_type == "income":
        return abs(amount)
    return abs(amount)


def sms_provider_name():
    return getattr(settings, "SMS_PROVIDER", "twilio").lower()


def sms_provider_configured():
    if sms_provider_name() == "telnyx":
        return bool(settings.TELNYX_API_KEY and settings.TELNYX_FROM_NUMBER)
    return bool(settings.TWILIO_ACCOUNT_SID and settings.TWILIO_AUTH_TOKEN and settings.TWILIO_FROM_NUMBER)


def resident_can_receive_sms(application):
    return bool(
        application.sms_opted_in
        and not application.sms_opted_out_at
        and normalize_phone_digits(application.phone)
    )


def sms_body(subject, message):
    body = f"{subject}\n\n{message}\n\nReply STOP to opt out."
    return body[:1500]


def sms_e164_phone(value):
    digits = normalize_phone_digits(value)
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    if str(value or "").strip().startswith("+"):
        return str(value).strip()
    return value


def send_twilio_sms(application, body):
    endpoint = f"https://api.twilio.com/2010-04-01/Accounts/{settings.TWILIO_ACCOUNT_SID}/Messages.json"
    payload = urlencode({
        "To": sms_e164_phone(application.phone),
        "From": settings.TWILIO_FROM_NUMBER,
        "Body": body,
    }).encode("utf-8")
    credentials = f"{settings.TWILIO_ACCOUNT_SID}:{settings.TWILIO_AUTH_TOKEN}".encode("utf-8")
    request = Request(endpoint, data=payload)
    request.add_header("Authorization", f"Basic {base64.b64encode(credentials).decode('ascii')}")
    response = urlopen(request, timeout=10)
    return response.read().decode("utf-8")


def send_telnyx_sms(application, body):
    request = Request(
        "https://api.telnyx.com/v2/messages",
        data=json.dumps({
            "to": sms_e164_phone(application.phone),
            "from": settings.TELNYX_FROM_NUMBER,
            "text": body,
            "type": "SMS",
        }).encode("utf-8"),
    )
    request.add_header("Authorization", f"Bearer {settings.TELNYX_API_KEY}")
    request.add_header("Content-Type", "application/json")
    request.add_header("Accept", "application/json")
    response = urlopen(request, timeout=10)
    response_body = response.read().decode("utf-8")
    try:
        payload = json.loads(response_body)
        return payload.get("data", {}).get("id") or response_body[:255]
    except json.JSONDecodeError:
        return response_body[:255]


def send_sms_message(application, body, sender, resident_message=None):
    log = SmsMessageLog.objects.create(
        application=application,
        resident_message=resident_message,
        to_phone=application.phone,
        body=body,
        sent_by=sender,
    )

    if not resident_can_receive_sms(application):
        log.status = "skipped_no_consent"
        log.error_message = "Resident has not opted in to SMS, has opted out, or has no phone number."
        log.save(update_fields=["status", "error_message"])
        return log

    if not sms_provider_configured():
        log.status = "not_configured"
        log.error_message = f"{sms_provider_name().title()} environment variables are not configured."
        log.save(update_fields=["status", "error_message"])
        return log

    try:
        if sms_provider_name() == "telnyx":
            provider_message_id = send_telnyx_sms(application, body)
        else:
            provider_message_id = send_twilio_sms(application, body)
    except Exception as exc:
        log.status = "failed"
        log.error_message = str(exc)
        log.save(update_fields=["status", "error_message"])
        return log

    log.status = "sent"
    log.provider_message_id = provider_message_id[:255]
    log.sent_at = timezone.now()
    log.save(update_fields=["status", "provider_message_id", "sent_at"])
    return log


def staff_required(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def reporting_required(user):
    return user.is_authenticated and (
        user.is_staff
        or user.is_superuser
        or getattr(user, "role", "") in ["property_owner", "landlord", "assistant", "admin"]
    )


def apply_completed_payment_to_balance(payment):
    application = payment.application

    if payment.payment_type == "rent":
        application.balance = max(Decimal("0.00"), application.balance - payment.amount)

    elif payment.payment_type == "deposit":
        application.deposit_paid = min(application.deposit_required, application.deposit_paid + payment.amount)

    elif payment.payment_type == "utility":
        application.utility_balance = max(Decimal("0.00"), application.utility_balance - payment.amount)

    elif payment.payment_type == "application_fee":
        application.application_fee_paid = min(application.application_fee_amount, application.application_fee_paid + payment.amount)

    elif payment.payment_type == "background_check_fee":
        application.background_check_fee_paid = min(application.background_check_fee_amount, application.background_check_fee_paid + payment.amount)
        if application.background_check_required and application.background_check_status == "pending":
            application.background_check_status = "ordered"

    elif payment.payment_type == "other" and "combined" in payment.description.lower():
        remaining = payment.amount

        rent_due = application.balance if application.balance > 0 else Decimal("0.00")
        rent_paid = min(rent_due, remaining)
        application.balance = max(Decimal("0.00"), application.balance - rent_paid)
        remaining -= rent_paid

        deposit_due = max(application.deposit_required - application.deposit_paid, Decimal("0.00"))
        deposit_paid = min(deposit_due, remaining)
        application.deposit_paid = min(application.deposit_required, application.deposit_paid + deposit_paid)
        remaining -= deposit_paid

        utility_due = application.utility_balance if application.utility_balance > 0 else Decimal("0.00")
        utility_paid = min(utility_due, remaining)
        application.utility_balance = max(Decimal("0.00"), application.utility_balance - utility_paid)

    application.save()


def recalculated_rent_due(application):
    return application.move_in_rent_charge if application.move_in_rent_charge > 0 else application.monthly_rent


def recalculated_utility_due(application):
    return application.move_in_utility_charge if application.move_in_utility_charge > 0 else application.utility_monthly


def expected_rent_for_month(application, month_start):
    if (
        application.lease_start_date
        and application.move_in_rent_charge > 0
        and application.lease_start_date.year == month_start.year
        and application.lease_start_date.month == month_start.month
    ):
        return application.move_in_rent_charge
    return application.monthly_rent or Decimal("0.00")


def expected_utility_for_month(application, month_start):
    if (
        application.lease_start_date
        and application.move_in_utility_charge > 0
        and application.lease_start_date.year == month_start.year
        and application.lease_start_date.month == month_start.month
    ):
        return application.move_in_utility_charge
    return application.utility_monthly or Decimal("0.00")


def first_day_of_month(value):
    return date(value.year, value.month, 1)


def add_months(value, months):
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def payment_service_month(payment):
    if payment.service_month:
        return first_day_of_month(payment.service_month)

    if payment.received_at:
        return first_day_of_month(timezone.localtime(payment.received_at).date())

    return first_day_of_month(timezone.localtime(payment.created_at).date())


def payment_month_allocations(payment):
    month_count = min(max(payment.months_covered or 1, 1), 24)
    base_amount = (payment.amount / Decimal(month_count)).quantize(Decimal("0.01"))
    allocated_total = Decimal("0.00")
    start_month = payment_service_month(payment)

    for index in range(month_count):
        amount = base_amount
        if index == month_count - 1:
            amount = payment.amount - allocated_total
        allocated_total += amount
        yield add_months(start_month, index), amount


def payment_amount_for_month(payments, year, month, payment_types=None):
    total = Decimal("0.00")
    payment_types = set(payment_types or [])

    for payment in payments:
        if payment_types and payment.payment_type not in payment_types:
            continue
        for accounting_month, amount in payment_month_allocations(payment):
            if accounting_month.year == year and accounting_month.month == month:
                total += amount

    return total


def selected_report_month(request):
    raw_month = (request.GET.get("month") or "").strip()
    if raw_month:
        try:
            return timezone.datetime.strptime(raw_month, "%Y-%m").date().replace(day=1)
        except ValueError:
            pass
    return timezone.localdate().replace(day=1)


def rent_roll_room_sort_key(room_label):
    clean_label = canonical_room_label(room_label)
    normalized_label = normalized_room_label(clean_label)
    if len(normalized_label) == 1 and normalized_label.isalpha():
        return (0, normalized_label.upper())
    return (1, normalized_label)


def resident_sort_key(resident):
    property_name = resident.property.name if getattr(resident, "property", None) else "No Property"
    return (
        property_name.lower(),
        rent_roll_room_sort_key(resident.space_label or ""),
        (resident.full_name or "").lower(),
        resident.id or 0,
    )


def sorted_resident_list(residents):
    return sorted(list(residents), key=resident_sort_key)


def rent_roll_base_row(property_obj, room_label, resident_name="No profile yet"):
    clean_room_label = canonical_room_label(room_label)
    return {
        "property": property_obj.name if property_obj else "No Property",
        "property_id": property_obj.id if property_obj else None,
        "room": clean_room_label or "-",
        "room_sort": rent_roll_room_sort_key(clean_room_label),
        "resident": resident_name or "No profile yet",
        "monthly_rent": Decimal("0.00"),
        "rent_paid": Decimal("0.00"),
        "rent_due_for_month": Decimal("0.00"),
        "current_rent_balance": Decimal("0.00"),
        "utility_monthly": Decimal("0.00"),
        "utility_paid": Decimal("0.00"),
        "utility_due_for_month": Decimal("0.00"),
        "current_utility_balance": Decimal("0.00"),
        "deposit_required": Decimal("0.00"),
        "deposit_paid": Decimal("0.00"),
        "deposit_due": Decimal("0.00"),
        "has_profile": False,
    }


def rent_roll_room_key(property_obj, room_label):
    return (property_obj.id if property_obj else None, normalized_room_label(room_label))


def apply_room_setting_to_rent_roll_row(row, setting):
    row["monthly_rent"] = setting.monthly_rent
    row["rent_due_for_month"] = setting.monthly_rent
    row["utility_monthly"] = setting.utility_monthly
    row["utility_due_for_month"] = setting.utility_monthly
    row["deposit_required"] = setting.deposit_required
    row["deposit_paid"] = setting.deposit_paid
    row["deposit_due"] = max(setting.deposit_required - setting.deposit_paid, Decimal("0.00"))


def apply_resident_to_rent_roll_row(row, resident, selected_month):
    completed_payments = list(resident.payments.filter(status="completed"))
    rent_paid = payment_amount_for_month(completed_payments, selected_month.year, selected_month.month, ["rent"])
    utility_paid = payment_amount_for_month(completed_payments, selected_month.year, selected_month.month, ["utility"])
    rent_expected = expected_rent_for_month(resident, selected_month)
    utility_expected = expected_utility_for_month(resident, selected_month)

    row["resident"] = resident.full_name
    row["monthly_rent"] = resident.monthly_rent
    row["rent_paid"] = rent_paid
    row["rent_due_for_month"] = max(rent_expected - rent_paid, Decimal("0.00"))
    row["current_rent_balance"] = resident.balance
    row["utility_monthly"] = resident.utility_monthly
    row["utility_paid"] = utility_paid
    row["utility_due_for_month"] = max(utility_expected - utility_paid, Decimal("0.00"))
    row["current_utility_balance"] = resident.utility_balance
    row["deposit_required"] = resident.deposit_required
    row["deposit_paid"] = resident.deposit_paid
    row["deposit_due"] = max(resident.deposit_required - resident.deposit_paid, Decimal("0.00"))
    row["has_profile"] = True


def selected_property_scope(request):
    properties = staff_managed_properties(request.user).order_by("name")
    selected_property = None
    property_id = (request.GET.get("property_id") or "").strip()

    if property_id:
        selected_property = get_object_or_404(properties, id=property_id)
        return properties.filter(id=selected_property.id), selected_property, False

    if properties.count() == 1:
        selected_property = properties.first()
        return properties.filter(id=selected_property.id), selected_property, False

    return properties, selected_property, True


def rent_roll_rows_for_properties(user, selected_month, properties):
    residents = (
        staff_managed_applications(user)
        .select_related("property")
        .filter(
            Q(user__isnull=False)
            | Q(payments__status="completed", payments__service_month=selected_month),
            property__in=properties,
        )
        .distinct()
        .order_by("property__name", "space_label", "full_name")
    )

    rows_by_room = OrderedDict()

    room_settings = (
        PropertyRoomRent.objects
        .select_related("property")
        .filter(property__in=properties, is_active=True)
        .order_by("property__name", "room_unit_label")
    )
    for setting in room_settings:
        key = rent_roll_room_key(setting.property, setting.room_unit_label)
        row = rows_by_room.setdefault(key, rent_roll_base_row(setting.property, setting.room_unit_label))
        apply_room_setting_to_rent_roll_row(row, setting)

    roster_entries = (
        CurrentResidentRosterEntry.objects
        .select_related("property")
        .filter(property__in=properties, is_active=True)
        .exclude(room_unit_label="")
        .order_by("property__name", "room_unit_label", "last_name", "first_name")
    )
    for entry in roster_entries:
        key = rent_roll_room_key(entry.property, entry.room_unit_label)
        row = rows_by_room.setdefault(key, rent_roll_base_row(entry.property, entry.room_unit_label))
        if not row["has_profile"] and row["resident"] == "No profile yet":
            row["resident"] = entry.full_name()

    for resident in residents:
        room_label = canonical_room_label(resident.space_label or "")
        if room_label:
            key = rent_roll_room_key(resident.property, room_label)
            row = rows_by_room.setdefault(key, rent_roll_base_row(resident.property, room_label))
        else:
            key = ("profile", resident.id)
            row = rows_by_room.setdefault(key, rent_roll_base_row(resident.property, ""))
        apply_resident_to_rent_roll_row(row, resident, selected_month)

    rows = sorted(rows_by_room.values(), key=lambda row: (row["property"].lower(), row["room_sort"], row["resident"].lower()))
    for row in rows:
        row["rent_balance"] = row["rent_due_for_month"]
        row["utility_balance"] = row["utility_due_for_month"]
        row["deposit_balance"] = row["deposit_due"]
    return rows


def rent_roll_totals(rows):
    total_fields = [
        "monthly_rent",
        "rent_paid",
        "rent_balance",
        "utility_monthly",
        "utility_paid",
        "utility_balance",
        "deposit_required",
        "deposit_paid",
        "deposit_balance",
    ]
    return {
        field: sum((row.get(field, Decimal("0.00")) or Decimal("0.00") for row in rows), Decimal("0.00"))
        for field in total_fields
    }


UTILITY_REPORT_KEYWORDS = [
    "power",
    "electric",
    "gas",
    "water",
    "sewer",
    "trash",
    "garbage",
    "internet",
    "utility",
    "utilities",
    "city of medford",
]


def custom_report_financial_entries(query_properties, entry_types=None, start_date=None, end_date=None):
    property_names = list(query_properties.values_list("name", flat=True))
    entries = FinancialEntry.objects.filter(property_name__in=property_names).select_related("upload")
    if entry_types:
        entries = entries.filter(entry_type__in=entry_types)
    if start_date:
        entries = entries.filter(entry_date__gte=start_date)
    if end_date:
        entries = entries.filter(entry_date__lte=end_date)
    return entries.order_by("property_name", "year", "month", "category", "description")


def custom_report_period_year(start_date):
    return start_date.year if start_date else timezone.localdate().year


def decimal_percent(numerator, denominator):
    denominator = Decimal(denominator or "0.00")
    if denominator <= 0:
        return Decimal("0.00")
    return (Decimal(numerator or "0.00") / denominator * Decimal("100.00")).quantize(Decimal("0.01"))


def summary_baseline_upload_ids_for_month(financial_entries, month_filter, entry_types=None):
    queryset = financial_entries.filter(month_filter, source_receipt__isnull=True)
    if entry_types:
        queryset = queryset.filter(entry_type__in=entry_types)
    return set(queryset.values_list("upload_id", flat=True).distinct())


def latest_summary_baseline_time(financial_entries, month_filter):
    summary_upload_ids = summary_baseline_upload_ids_for_month(financial_entries, month_filter, ["income", "operating_expense", "debt_service", "capital_expense"])
    if not summary_upload_ids:
        return None
    return (
        FinancialUpload.objects
        .filter(id__in=summary_upload_ids)
        .aggregate(latest=Max("uploaded_at"))["latest"]
    )


def entries_total(queryset):
    return queryset.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")


def recalculate_application_balances(application):
    completed_payments = application.payments.filter(status="completed")
    rent_paid = completed_payments.filter(payment_type="rent").aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    utility_paid = completed_payments.filter(payment_type="utility").aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    deposit_paid = completed_payments.filter(payment_type="deposit").aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    application.balance = max(Decimal("0.00"), recalculated_rent_due(application) - rent_paid)
    application.utility_balance = max(Decimal("0.00"), recalculated_utility_due(application) - utility_paid)
    application.deposit_paid = min(application.deposit_required, deposit_paid)
    application.save(update_fields=["balance", "utility_balance", "deposit_paid"])


def prorated_monthly_charge(monthly_amount, start_date):
    monthly_amount = Decimal(monthly_amount or "0.00")
    if not start_date or monthly_amount <= 0:
        return monthly_amount.quantize(Decimal("0.01"))

    days_in_month = calendar.monthrange(start_date.year, start_date.month)[1]
    charge_days = days_in_month - start_date.day + 1
    if charge_days >= days_in_month:
        return monthly_amount.quantize(Decimal("0.01"))

    prorated = monthly_amount * Decimal(charge_days) / Decimal(days_in_month)
    return prorated.quantize(Decimal("0.01"))


def home(request):
    if getattr(settings, "RENTAL_LEDGER_SITE", False):
        return rental_ledger_pro_home(request)

    properties = Property.objects.all()
    posts = BlogPost.objects.filter(property__isnull=True).order_by("-created_at")[:5]
    painted_lady_profile_property = (
        properties
        .filter(name__icontains="painted lady")
        .order_by("name")
        .first()
    )

    return render(request, "home.html", {
        "properties": properties,
        "posts": posts,
        "painted_lady_profile_property": painted_lady_profile_property,
        "painted_lady_profile_open": bool(
            painted_lady_profile_property
            and property_existing_resident_intake_open(painted_lady_profile_property)
        ),
    })


RENTAL_LEDGER_PRODUCT_PAGES = {
    "financial-command": {
        "eyebrow": "Financial Command",
        "title": "Reports that make property performance clear.",
        "lead": "Rental Ledger Pro is built around the reports owners, lenders, buyers, and operators ask for: T-12, rent roll, NOI, cash flow after debt, payment ledger, valuation estimate, and receipt-backed expenses.",
        "accent": "cyan",
        "dashboard_title": "Owner / Manager Portfolio View",
        "dashboard_summary": "For owners and property managers with more than one property, the first screen should compare the portfolio, then let them drill into one property without mixing records.",
        "dashboard_metrics": [
            ("Portfolio Income", "$788.3k"),
            ("NOI", "$424.0k"),
            ("Cash Flow", "$196.6k"),
            ("Occupancy", "98%"),
        ],
        "dashboard_columns": ["Property", "Units", "Occupancy", "NOI", "Watch"],
        "dashboard_rows": [
            ["Cedar Ridge", "48", "96%", "$89.2k", "1 rent open"],
            ["Hillview Commons", "32", "97%", "$55.3k", "2 deposits"],
            ["Evergreen Flats", "54", "98%", "$132.6k", "Capital review"],
            ["Riverstone Court", "40", "100%", "$146.9k", "Water trend"],
        ],
        "report_slugs": ["t12", "rent-roll", "property-performance", "valuation", "vendor-expense", "utility-trend"],
        "features": [
            ("T-12 and NOI", "Monthly income, operating expenses, debt service, cash flow, and year-to-date totals."),
            ("Rent roll and payment ledger", "Resident/unit rent status, service month tracking, payment methods, balances, and printable history."),
            ("Receipt-backed expenses", "Phone or desktop receipt upload with category review and proof attached to the property ledger."),
            ("Valuation support", "NOI-based valuation estimates using cap-rate scenarios and property performance trends."),
        ],
    },
    "operations-hub": {
        "eyebrow": "Operations Hub",
        "title": "One work queue for the daily landlord job.",
        "lead": "The operations hub keeps resident files, applications, onboarding documents, messages, maintenance requests, rent setup, and monthly collection watch lists in one property-aware workspace.",
        "accent": "amber",
        "dashboard_title": "Landlord Dashboard",
        "dashboard_summary": "The landlord view should focus on what needs action today: unpaid balances, new messages, maintenance, documents, setup requests, and property-specific resident files.",
        "dashboard_metrics": [
            ("Needs Attention", "14"),
            ("Open Rent", "$1,935"),
            ("Open Utilities", "$200"),
            ("New Messages", "6"),
        ],
        "dashboard_columns": ["Unit", "Resident", "Issue", "Amount", "Action"],
        "dashboard_rows": [
            ["103", "Diane Cole", "Utilities", "$65", "Send reminder"],
            ["105", "Lena Morales", "Partial rent", "$390", "Record split pay"],
            ["F", "Marcus Doyle", "Rent + utilities", "$1,180", "Call resident"],
            ["204", "Iris Morgan", "Utilities", "$80", "Send reminder"],
            ["2A", "Jamal Pierce", "Partial rent", "$420", "ACH follow-up"],
        ],
        "report_slugs": ["delinquency", "payment-log", "rent-roll", "vendor-expense", "capital-log", "insurance-compliance"],
        "features": [
            ("Resident files", "Applications, leases, emergency contacts, documents, payments, profile photos, and notes stay attached to the resident."),
            ("Needs-attention workflow", "New messages, documents, setup requests, applications, and unpaid balances surface without digging through pages."),
            ("Unit-level rent setup", "Rent, utilities, deposit rules, due dates, and room/unit labels stay consistent across reports and dashboards."),
            ("Property separation", "Owners and landlords see only the properties, residents, reports, and messages they are allowed to access."),
        ],
    },
    "property-app": {
        "eyebrow": "Branded Property App",
        "title": "One platform that can feel local to every property.",
        "lead": "Residents should not feel like they are using generic software. Rental Ledger Pro can present the right property name, visuals, notices, documents, payments, utility setup steps, and communication tools for each property.",
        "accent": "violet",
        "dashboard_title": "Tenant Dashboard",
        "dashboard_summary": "The resident view should be simple: what they owe, what documents need attention, how to message the property, and what property-specific setup steps remain.",
        "dashboard_metrics": [
            ("Current Balance", "$0"),
            ("Inbox", "2"),
            ("Setup Steps", "3 of 4"),
            ("Maintenance", "1 open"),
        ],
        "dashboard_columns": ["Panel", "What Resident Sees", "Status", "Next Step"],
        "dashboard_rows": [
            ["Balance", "Rent, utilities, deposit detail", "Current", "Print receipt"],
            ["Inbox", "Lease update and community notice", "Needs signature", "Open document"],
            ["Messages", "Private thread with manager", "New reply", "Respond"],
            ["Utilities", "Power and water setup links", "In progress", "Finish checklist"],
            ["Insurance", "Suggested renters insurance link", "Optional", "Open provider"],
        ],
        "report_slugs": ["payment-log", "insurance-compliance", "utility-trend"],
        "features": [
            ("Property identity", "Each property can carry its own name, photos, colors, notices, and resident instructions."),
            ("Resident payments", "Rent, utilities, deposits, balance details, payment history, and receipts stay easy to find."),
            ("Onboarding checklist", "Lease, emergency contact, property rules, utility vendor setup, and insurance links can guide new residents."),
            ("Powered by Rental Ledger Pro", "The product brand can stay present without taking over the property experience."),
        ],
    },
}


RENTAL_LEDGER_DEMO_REPORTS = {
    "t12": {
        "title": "T-12 Performance Report",
        "eyebrow": "Financial Command",
        "summary": "A lender-ready operating view showing monthly income, operating expenses, NOI, debt service, cash flow, occupancy, and collection performance.",
        "columns": ["Line", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD"],
        "rows": [
            ["Potential Rent", "$38,640", "$38,640", "$39,280", "$39,280", "$39,930", "$41,080", "$41,080", "$41,080", "$41,080", "$41,080", "$41,730", "$41,730", "$483,630"],
            ["Collected Rent", "$37,940", "$37,420", "$38,915", "$38,480", "$39,210", "$40,380", "$40,960", "$40,525", "$39,870", "$40,415", "$41,120", "$41,730", "$476,965"],
            ["Utility Recovery", "$3,420", "$3,395", "$3,510", "$3,485", "$3,560", "$3,640", "$3,705", "$3,705", "$3,650", "$3,705", "$3,760", "$3,760", "$43,295"],
            ["Other Income", "$1,050", "$920", "$1,215", "$1,060", "$1,140", "$1,335", "$1,180", "$1,025", "$1,210", "$1,145", "$1,275", "$1,160", "$13,715"],
            ["Gross Income", "$42,410", "$41,735", "$43,640", "$43,025", "$43,910", "$45,355", "$45,845", "$45,255", "$44,730", "$45,265", "$46,155", "$46,650", "$533,975"],
            ["Payroll / Labor", "$6,820", "$6,740", "$6,880", "$6,910", "$7,050", "$7,120", "$7,120", "$7,210", "$7,180", "$7,240", "$7,250", "$7,280", "$84,800"],
            ["Repairs / Maintenance", "$3,115", "$2,845", "$4,920", "$3,360", "$3,780", "$4,410", "$3,255", "$3,020", "$5,480", "$3,210", "$3,760", "$3,150", "$44,305"],
            ["Utilities", "$7,215", "$7,460", "$7,040", "$6,580", "$6,455", "$6,720", "$7,010", "$7,175", "$6,980", "$6,740", "$6,610", "$6,900", "$82,885"],
            ["Insurance / Taxes", "$4,330", "$4,330", "$4,330", "$4,330", "$4,480", "$4,480", "$4,480", "$4,480", "$4,480", "$4,630", "$4,630", "$4,630", "$53,610"],
            ["Admin / Software", "$1,180", "$1,140", "$1,235", "$1,205", "$1,260", "$1,320", "$1,310", "$1,295", "$1,345", "$1,330", "$1,350", "$1,365", "$15,335"],
            ["Total Operating Expenses", "$22,660", "$22,515", "$24,405", "$22,385", "$23,025", "$24,050", "$23,175", "$23,180", "$25,465", "$23,150", "$23,600", "$23,325", "$280,935"],
            ["NOI", "$19,750", "$19,220", "$19,235", "$20,640", "$20,885", "$21,305", "$22,670", "$22,075", "$19,265", "$22,115", "$22,555", "$23,325", "$253,040"],
            ["Debt Service", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$11,200", "$134,400"],
            ["Cash Flow After Debt", "$8,550", "$8,020", "$8,035", "$9,440", "$9,685", "$10,105", "$11,470", "$10,875", "$8,065", "$10,915", "$11,355", "$12,125", "$118,640"],
            ["Occupancy", "94%", "94%", "96%", "96%", "96%", "98%", "98%", "98%", "96%", "98%", "98%", "100%", "97%"],
            ["Collection Rate", "98%", "97%", "99%", "98%", "98%", "98%", "100%", "99%", "97%", "98%", "99%", "100%", "98%"],
        ],
    },
    "rent-roll": {
        "title": "Rent Roll",
        "eyebrow": "Operations Hub",
        "summary": "A property and unit-sorted resident view showing scheduled rent, utilities, deposits, balances, lease status, and payment status.",
        "columns": ["Property", "Unit", "Resident", "Rent", "Utilities", "Deposit Held", "Balance", "Lease Status", "Status"],
        "rows": [
            ["Cedar Ridge", "101", "Marisol Vega", "$1,175", "$65", "$1,175", "$0", "Active", "Clear"],
            ["Cedar Ridge", "102", "Anthony Reed", "$1,225", "$65", "$1,225", "$0", "Active", "Clear"],
            ["Cedar Ridge", "103", "Diane Cole", "$1,195", "$65", "$1,195", "$65", "Active", "Utilities Open"],
            ["Cedar Ridge", "104", "Omar Nash", "$1,275", "$65", "$1,275", "$0", "Active", "Clear"],
            ["Cedar Ridge", "105", "Lena Morales", "$1,310", "$65", "$1,310", "$390", "Renewal Due", "Partial Rent"],
            ["Cedar Ridge", "106", "Vacant", "$1,325", "$0", "$0", "$0", "Available", "Turnover"],
            ["Hillview Commons", "A", "Calvin Price", "$975", "$55", "$975", "$0", "Active", "Clear"],
            ["Hillview Commons", "B", "Grace Nolan", "$1,025", "$55", "$1,025", "$0", "Active", "Clear"],
            ["Hillview Commons", "C", "Priya Shah", "$1,050", "$55", "$1,050", "$0", "Active", "Clear"],
            ["Hillview Commons", "D", "Victor Lane", "$1,050", "$55", "$775", "$275", "Active", "Deposit Balance"],
            ["Hillview Commons", "E", "Samira Ellis", "$1,100", "$55", "$1,100", "$0", "Active", "Clear"],
            ["Hillview Commons", "F", "Marcus Doyle", "$1,125", "$55", "$1,125", "$1,180", "Active", "Rent Open"],
            ["Evergreen Flats", "201", "Rosa Bennett", "$1,450", "$80", "$1,450", "$0", "Active", "Clear"],
            ["Evergreen Flats", "202", "Noah Kim", "$1,475", "$80", "$1,475", "$0", "Active", "Clear"],
            ["Evergreen Flats", "203", "Hannah Wells", "$1,500", "$80", "$1,500", "$0", "Active", "Clear"],
            ["Evergreen Flats", "204", "Iris Morgan", "$1,525", "$80", "$1,525", "$80", "Active", "Utilities Open"],
            ["Evergreen Flats", "205", "Theo Brooks", "$1,550", "$80", "$1,550", "$0", "Active", "Clear"],
            ["Riverstone Court", "1A", "Elliot Fisher", "$1,625", "$0", "$1,625", "$0", "Active", "Clear"],
            ["Riverstone Court", "1B", "Natalie Soto", "$1,650", "$0", "$1,650", "$0", "Active", "Clear"],
            ["Riverstone Court", "2A", "Jamal Pierce", "$1,675", "$0", "$1,675", "$420", "Active", "Partial Rent"],
            ["Riverstone Court", "2B", "Mei Tan", "$1,700", "$0", "$1,700", "$0", "Active", "Clear"],
            ["Riverstone Court", "3A", "Vacant", "$1,725", "$0", "$0", "$0", "Available", "Market Ready"],
            ["Riverstone Court", "3B", "Olivia Grant", "$1,750", "$0", "$1,750", "$0", "New Lease", "Clear"],
        ],
    },
    "payment-log": {
        "title": "Payment Log",
        "eyebrow": "Ledger Records",
        "summary": "A month-filtered payment ledger for rent, utilities, deposits, split payments, manual payments, card payments, and bank deposits.",
        "columns": ["Date", "Property", "Unit", "Resident", "Type", "Method", "Amount", "Service Month", "Recorded By"],
        "rows": [
            ["Jun 03, 2026", "Cedar Ridge", "101", "Marisol Vega", "Rent", "Card", "$1,175", "June 2026", "Resident"],
            ["Jun 03, 2026", "Cedar Ridge", "103", "Diane Cole", "Utilities", "ACH", "$65", "June 2026", "Bank Import"],
            ["Jun 02, 2026", "Hillview Commons", "A", "Calvin Price", "Rent", "Cash", "$975", "June 2026", "Manager"],
            ["Jun 02, 2026", "Hillview Commons", "A", "Calvin Price", "Utilities", "Cash", "$55", "June 2026", "Manager"],
            ["Jun 02, 2026", "Evergreen Flats", "202", "Noah Kim", "Rent", "Check", "$1,475", "June 2026", "Manager"],
            ["Jun 01, 2026", "Riverstone Court", "1B", "Natalie Soto", "Rent", "Card", "$1,650", "June 2026", "Resident"],
            ["Jun 01, 2026", "Riverstone Court", "1B", "Natalie Soto", "Deposit", "Card", "$1,650", "-", "Resident"],
            ["May 31, 2026", "Cedar Ridge", "105", "Lena Morales", "Rent", "Split Cash", "$920", "June 2026", "Manager"],
            ["May 31, 2026", "Cedar Ridge", "105", "Lena Morales", "Rent", "Check", "$390", "June 2026", "Manager"],
            ["May 30, 2026", "Hillview Commons", "D", "Victor Lane", "Deposit", "Cash", "$250", "-", "Manager"],
            ["May 30, 2026", "Evergreen Flats", "204", "Iris Morgan", "Utilities", "Card", "$80", "June 2026", "Resident"],
            ["May 29, 2026", "Riverstone Court", "2A", "Jamal Pierce", "Rent", "ACH", "$1,255", "June 2026", "Bank Import"],
            ["May 28, 2026", "Evergreen Flats", "205", "Theo Brooks", "Rent", "ACH", "$1,550", "June 2026", "Bank Import"],
            ["May 28, 2026", "Hillview Commons", "F", "Marcus Doyle", "Late Fee", "Manual", "$45", "May 2026", "Manager"],
            ["May 27, 2026", "Cedar Ridge", "102", "Anthony Reed", "Rent", "Card", "$1,225", "June 2026", "Resident"],
            ["May 26, 2026", "Evergreen Flats", "201", "Rosa Bennett", "Deposit", "Check", "$1,450", "-", "Manager"],
            ["May 26, 2026", "Evergreen Flats", "201", "Rosa Bennett", "Rent", "Check", "$1,450", "June 2026", "Manager"],
            ["May 25, 2026", "Riverstone Court", "3B", "Olivia Grant", "Rent", "Card", "$1,750", "June 2026", "Resident"],
        ],
    },
    "valuation": {
        "title": "Valuation Estimate",
        "eyebrow": "Owner Reports",
        "summary": "An estimate view using NOI, cap-rate scenarios, occupancy, and improvement assumptions to help owners understand potential valuation ranges.",
        "columns": ["Property", "Scenario", "Cap Rate", "NOI Used", "Estimated Value", "Notes"],
        "rows": [
            ["Cedar Ridge", "Conservative", "7.75%", "$83,400", "$1,076,129", "Uses trailing NOI before rent increases"],
            ["Cedar Ridge", "Market", "6.85%", "$91,800", "$1,340,146", "Uses stabilized rent roll"],
            ["Cedar Ridge", "Upside", "6.25%", "$101,250", "$1,620,000", "After renovation and loss-to-lease capture"],
            ["Hillview Commons", "Conservative", "8.10%", "$62,900", "$776,543", "Small asset, higher cap assumption"],
            ["Hillview Commons", "Market", "7.25%", "$67,450", "$930,345", "Comparable local workforce housing"],
            ["Evergreen Flats", "Market", "6.65%", "$132,600", "$1,993,985", "Higher rent base and lower vacancy"],
            ["Riverstone Court", "Market", "6.40%", "$146,850", "$2,294,531", "Newer building, stronger resident profile"],
        ],
    },
    "vendor-expense": {
        "title": "Vendor Expense Report",
        "eyebrow": "Accounting",
        "summary": "A report that groups expenses by vendor, property, and category so owners can see what is driving operating cost.",
        "columns": ["Property", "Category", "Vendor", "YTD Total", "Receipts", "Last Receipt", "Review Status"],
        "rows": [
            ["Cedar Ridge", "Power", "Pacific Energy", "$11,824", "12", "Jun 02", "Filed"],
            ["Cedar Ridge", "Trash", "Metro Waste", "$3,984", "12", "Jun 01", "Filed"],
            ["Cedar Ridge", "Maintenance Supplies", "Lowe's", "$5,620", "31", "Jun 03", "Needs split review"],
            ["Cedar Ridge", "Cleaning Labor", "Bright Hall Services", "$8,400", "12", "May 28", "Filed"],
            ["Hillview Commons", "Plumbing", "Cascade Plumbing", "$2,875", "4", "May 19", "Filed"],
            ["Hillview Commons", "Insurance", "Harbor Mutual", "$5,280", "4", "Apr 30", "Filed"],
            ["Hillview Commons", "Landscaping", "Greenline Grounds", "$3,125", "8", "May 22", "Filed"],
            ["Evergreen Flats", "HVAC", "North Star Mechanical", "$6,940", "7", "Jun 01", "Capital review"],
            ["Evergreen Flats", "Internet", "FiberWorks", "$1,548", "6", "Jun 02", "Filed"],
            ["Evergreen Flats", "Security", "DoorCloud Access", "$2,160", "6", "May 31", "Filed"],
            ["Riverstone Court", "Water/Sewer", "City Utility", "$7,430", "12", "Jun 01", "Filed"],
            ["Riverstone Court", "Pest Control", "SafeNest Pest", "$1,200", "6", "May 17", "Filed"],
            ["Riverstone Court", "Turnover Materials", "Home Depot", "$4,775", "18", "Jun 03", "Needs receipt approval"],
            ["Portfolio", "Software", "Rental Ledger Pro", "$3,600", "6", "Jun 01", "Filed"],
        ],
    },
    "utility-trend": {
        "title": "Utility Cost Trend",
        "eyebrow": "Property Performance",
        "summary": "A utility-focused trend report for power, gas, water, trash, internet, and shared utility recovery across multiple properties.",
        "columns": ["Property", "Utility", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Trend", "Action"],
        "rows": [
            ["Cedar Ridge", "Power", "$1,444", "$1,595", "$1,246", "$781", "$736", "$802", "Down", "Track after HVAC upgrade"],
            ["Cedar Ridge", "Gas", "$146", "$157", "$153", "$134", "$142", "$126", "Stable", "No action"],
            ["Cedar Ridge", "Water", "$524", "$524", "$516", "$516", "$504", "$529", "Stable", "No action"],
            ["Cedar Ridge", "Trash", "$364", "$366", "$39", "$363", "$0", "$366", "Review", "Missing May invoice"],
            ["Hillview Commons", "Power", "$920", "$1,010", "$890", "$812", "$795", "$826", "Down", "Monitor summer"],
            ["Hillview Commons", "Water", "$410", "$405", "$426", "$451", "$488", "$522", "Up", "Check leak risk"],
            ["Evergreen Flats", "Power", "$2,140", "$2,220", "$1,980", "$1,760", "$1,690", "$1,715", "Down", "Normal seasonality"],
            ["Evergreen Flats", "Internet", "$258", "$258", "$258", "$258", "$258", "$258", "Flat", "Contract stable"],
            ["Riverstone Court", "Water", "$760", "$744", "$790", "$818", "$846", "$905", "Up", "Review irrigation"],
            ["Riverstone Court", "Trash", "$520", "$520", "$520", "$540", "$540", "$540", "Up", "Rate change noted"],
        ],
    },
    "property-performance": {
        "title": "Property Performance Summary",
        "eyebrow": "Portfolio",
        "summary": "A portfolio-level report comparing occupancy, collected income, operating expenses, NOI, debt service, and cash flow by property.",
        "columns": ["Property", "Units", "Occupied", "Occupancy", "Collected Income", "Operating Expenses", "NOI", "Debt Service", "Cash Flow", "Watch Items"],
        "rows": [
            ["Cedar Ridge", "48", "46", "96%", "$182,460", "$93,220", "$89,240", "$48,000", "$41,240", "1 rent open, 1 turnover"],
            ["Hillview Commons", "32", "31", "97%", "$119,840", "$64,515", "$55,325", "$35,400", "$19,925", "2 deposit balances"],
            ["Evergreen Flats", "54", "53", "98%", "$246,110", "$113,510", "$132,600", "$69,600", "$63,000", "HVAC capital review"],
            ["Riverstone Court", "40", "40", "100%", "$239,900", "$93,050", "$146,850", "$74,400", "$72,450", "Water trend up"],
            ["Portfolio Total", "174", "170", "98%", "$788,310", "$364,295", "$424,015", "$227,400", "$196,615", "4 open work items"],
        ],
    },
    "delinquency": {
        "title": "Delinquency and Collection Watch",
        "eyebrow": "Operations Hub",
        "summary": "A focused monthly watch list showing who still owes rent, utilities, deposits, or fees and what action is next.",
        "columns": ["Property", "Unit", "Resident", "Rent Balance", "Utility Balance", "Deposit Balance", "Days Late", "Next Action"],
        "rows": [
            ["Cedar Ridge", "103", "Diane Cole", "$0", "$65", "$0", "2", "Send utility reminder"],
            ["Cedar Ridge", "105", "Lena Morales", "$390", "$0", "$0", "2", "Record expected split payment"],
            ["Hillview Commons", "D", "Victor Lane", "$0", "$0", "$275", "0", "Deposit plan due Jun 15"],
            ["Hillview Commons", "F", "Marcus Doyle", "$1,125", "$55", "$0", "4", "Call resident"],
            ["Evergreen Flats", "204", "Iris Morgan", "$0", "$80", "$0", "2", "Send utility reminder"],
            ["Riverstone Court", "2A", "Jamal Pierce", "$420", "$0", "$0", "2", "ACH follow-up"],
        ],
    },
    "capital-log": {
        "title": "Capital Improvement Log",
        "eyebrow": "Asset Planning",
        "summary": "A project log for owner-visible improvements that should be separated from normal operating expenses.",
        "columns": ["Property", "Project", "Vendor", "Budget", "Spent To Date", "Status", "Expected Impact"],
        "rows": [
            ["Cedar Ridge", "Window replacement phase 1", "ClearView Glass", "$42,000", "$18,400", "In Progress", "Lower energy loss"],
            ["Cedar Ridge", "Unit 106 turnover renovation", "Lowe's / Local Labor", "$6,800", "$4,120", "In Progress", "Market-ready vacancy"],
            ["Hillview Commons", "Laundry room upgrade", "SpinTech Laundry", "$14,500", "$0", "Approved", "New service income"],
            ["Evergreen Flats", "HVAC compressor replacements", "North Star Mechanical", "$28,000", "$6,940", "Review", "Reduce emergency calls"],
            ["Evergreen Flats", "Access control upgrade", "DoorCloud Access", "$9,800", "$4,250", "In Progress", "Resident safety"],
            ["Riverstone Court", "Irrigation repair", "Greenline Grounds", "$5,500", "$1,125", "Scheduled", "Reduce water trend"],
        ],
    },
    "insurance-compliance": {
        "title": "Insurance and Compliance Report",
        "eyebrow": "Risk Control",
        "summary": "A property and resident compliance view for policies, renter insurance status, lease documents, and required follow-up.",
        "columns": ["Property", "Requirement", "Compliant", "Open Items", "Renewal / Due Date", "Owner Note"],
        "rows": [
            ["Cedar Ridge", "Master policy", "Yes", "0", "Oct 01, 2026", "Policy filed"],
            ["Cedar Ridge", "Resident renters insurance", "42 of 46", "4", "Monthly review", "Send reminder"],
            ["Hillview Commons", "Master policy", "Yes", "0", "Sep 15, 2026", "Filed"],
            ["Hillview Commons", "Resident renters insurance", "27 of 31", "4", "Monthly review", "Offer tenant link"],
            ["Evergreen Flats", "Fire inspection", "Yes", "0", "Jan 20, 2027", "Passed"],
            ["Evergreen Flats", "Resident renters insurance", "51 of 53", "2", "Monthly review", "New lease follow-up"],
            ["Riverstone Court", "Master policy", "Yes", "0", "Nov 12, 2026", "Filed"],
            ["Riverstone Court", "Resident renters insurance", "39 of 40", "1", "Monthly review", "One renewal missing"],
        ],
    },
}


def rental_ledger_pro_home(request):
    return render(request, "rental_ledger_pro_home.html", {
        "product_pages": RENTAL_LEDGER_PRODUCT_PAGES,
        "demo_reports": RENTAL_LEDGER_DEMO_REPORTS,
    })


def rental_ledger_demo(request):
    return render(request, "rental_ledger_demo.html", {
        "interactive_demo_enabled": getattr(settings, "DEMO_MODE", False),
        "product_pages": RENTAL_LEDGER_PRODUCT_PAGES,
        "reports": RENTAL_LEDGER_DEMO_REPORTS,
    })


def rental_ledger_product_page(request, page_slug):
    page = RENTAL_LEDGER_PRODUCT_PAGES.get(page_slug)
    if not page:
        raise Http404("Rental Ledger Pro page not found.")

    page_reports = {
        slug: RENTAL_LEDGER_DEMO_REPORTS[slug]
        for slug in page.get("report_slugs", [])
        if slug in RENTAL_LEDGER_DEMO_REPORTS
    }

    return render(request, "rental_ledger_product_page.html", {
        "page": page,
        "page_slug": page_slug,
        "reports": page_reports,
    })


def rental_ledger_demo_report(request, report_slug):
    report = RENTAL_LEDGER_DEMO_REPORTS.get(report_slug)
    if not report:
        raise Http404("Rental Ledger Pro report not found.")

    return render(request, "rental_ledger_demo_report.html", {
        "report": report,
        "report_slug": report_slug,
        "reports": RENTAL_LEDGER_DEMO_REPORTS,
    })


def demo_entry(request):
    if not getattr(settings, "DEMO_MODE", False):
        return HttpResponse(
            "Demo route is installed, but DEMO_MODE is not enabled on this Render service.",
            status=403,
            content_type="text/plain",
        )

    demo_user = User.objects.filter(username=settings.DEMO_ADMIN_USERNAME).first()
    if not demo_user:
        call_command("reset_demo_environment", "--confirm", verbosity=0)
        demo_user = User.objects.get(username=settings.DEMO_ADMIN_USERNAME)

    login(request, demo_user, backend="django.contrib.auth.backends.ModelBackend")
    request.session.set_expiry(getattr(settings, "DEMO_SESSION_SECONDS", 7200))
    messages.info(
        request,
        "You are using a temporary demo workspace. Sample data resets automatically and should not contain real information.",
    )
    return redirect("superadmin_dashboard")


def demo_status(request):
    return JsonResponse({
        "demo_route_installed": True,
        "demo_mode": getattr(settings, "DEMO_MODE", False),
        "demo_admin_username": getattr(settings, "DEMO_ADMIN_USERNAME", ""),
    })
     
def properties_list(request):
    properties = Property.objects.all().order_by("name")

    return render(request, "properties.html", {
        "properties": properties,
    })

def creed(request):
    return render(request, "creed.html")


def who_we_serve(request):
    return render(request, "who_we_serve.html")


def privacy_policy(request):
    return render(request, "privacy_policy.html")


def terms_of_service(request):
    return render(request, "terms_of_service.html")


def property_owner_intake(request):
    form = PropertyOwnerIntakeForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Your property owner questionnaire has been submitted.")
        return redirect("property_owner_intake_success")

    return render(request, "property_owner_intake.html", {"form": form})


def property_owner_intake_success(request):
    return render(request, "property_owner_intake_success.html")


def calculate_screening_score(application):
    score = 0
    factors = []

    if application.monthly_rent and application.monthly_income:
        rent_ratio = application.monthly_income / application.monthly_rent if application.monthly_rent else Decimal("0")
        if rent_ratio >= Decimal("3"):
            score += 25
            factors.append("Income is at least 3x the monthly rent.")
        elif rent_ratio >= Decimal("2"):
            score += 16
            factors.append("Income is at least 2x the monthly rent.")
        else:
            score += 6
            factors.append("Income is below 2x monthly rent and needs owner review.")
    elif application.monthly_income:
        score += 10
        factors.append("Income was provided, but no rent amount is assigned yet.")

    if application.current_address and application.current_address_length:
        score += 10
        factors.append("Current housing history is present.")

    previous_housing_fields = [
        application.previous_address_1,
        application.previous_address_2,
        application.previous_address_3,
    ]
    if any(previous_housing_fields):
        score += 10
        factors.append("Prior housing history was provided.")

    eviction_text = (application.previous_evictions or "").strip().lower()
    if eviction_text and eviction_text not in {"no", "none", "n/a", "na"}:
        factors.append("Applicant disclosed prior evictions or housing barriers.")
    else:
        score += 12
        factors.append("No prior eviction concern was disclosed.")

    if application.has_valid_odl or application.oregon_id_number or application.drivers_license_number or application.id_upload:
        score += 10
        factors.append("Applicant provided identification information.")

    if application.reference_1_name and application.reference_1_phone:
        score += 8
        factors.append("Primary reference is available.")

    if application.reference_2_name and application.reference_2_phone:
        score += 5
        factors.append("Secondary reference is available.")

    if application.background_check_status == "cleared":
        score += 20
        factors.append("Background report is marked cleared.")
    elif application.background_check_status == "needs_review":
        score += 6
        factors.append("Background report needs owner review.")
    elif application.background_check_status == "declined":
        factors.append("Background report is marked declined.")
    elif application.background_check_required:
        factors.append("Background report is required but not cleared yet.")
    else:
        score += 8
        factors.append("Background check is not required for this property.")

    score = min(score, 100)
    if score >= 85:
        rating = "strong"
    elif score >= 70:
        rating = "qualified"
    elif score >= 50:
        rating = "review"
    elif score >= 30:
        rating = "high_risk"
    else:
        rating = "declined"

    return score, rating, factors


def apply(request):
    property_id = request.GET.get("property") or request.POST.get("property")
    property_obj = None

    if property_id:
        property_obj = get_object_or_404(Property, pk=property_id)

    if request.method == "POST":
        form = HousingApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            if property_obj:
                application.property = property_obj
                if property_obj.charges_application_fee:
                    application.application_fee_amount = property_obj.application_fee_amount
                if property_obj.requires_background_check:
                    if not application.screening_consent:
                        form.add_error("screening_consent", "Consent is required before this property can process applicant screening.")
                        return render(request, "apply.html", {
                            "form": form,
                            "property": property_obj,
                        })
                    application.background_check_required = True
                    application.background_check_fee_amount = property_obj.background_check_fee_amount
                    application.background_check_status = "pending"
                    application.screening_provider_name = property_obj.screening_provider_name
            if application.sms_opted_in:
                application.sms_opted_in_at = timezone.now()
                application.communication_preference = "sms"
            if application.screening_consent:
                application.screening_consent_at = timezone.now()
            application.save()
            request.session["submitted_application_id"] = application.id
            return redirect("apply_success")
    else:
        form = HousingApplicationForm()
    return render(request, "apply.html", {
        "form": form,
        "property": property_obj,
    })


def apply_success(request):
    application = None
    application_id = request.session.get("submitted_application_id")
    if application_id:
        application = HousingApplication.objects.filter(id=application_id).first()

    return render(request, "apply_success.html", {
        "application": application,
    })


def build_listing_copy(request, listing):
    apply_url = request.build_absolute_uri(f"{reverse('apply')}?property={listing.property_id}")
    public_url = request.build_absolute_uri(reverse("public_rental_listing", args=[listing.id]))
    lines = [
        listing.headline,
        "",
        f"Property: {listing.property.name}",
    ]

    if listing.unit_label:
        lines.append(f"Unit: {listing.unit_label}")
    if listing.rent_amount:
        lines.append(f"Rent: ${listing.rent_amount}/month")
    if listing.deposit_amount:
        lines.append(f"Deposit: ${listing.deposit_amount}")
    if listing.available_date:
        lines.append(
            f"Available: {calendar.month_name[listing.available_date.month]} {listing.available_date.day}, {listing.available_date.year}"
        )
    if listing.lease_terms:
        lines.append(f"Terms: {listing.lease_terms}")
    if listing.utilities_description:
        lines.append(f"Utilities: {listing.utilities_description}")

    lines.extend(["", listing.listing_body or listing.property.description])

    if listing.property_benefits:
        lines.extend(["", "Property benefits:", listing.property_benefits])
    if listing.amenities:
        lines.extend(["", "Amenities:", listing.amenities])
    if listing.screening_summary:
        lines.extend(["", "Application / screening:", listing.screening_summary])

    lines.extend([
        "",
        f"Apply here: {apply_url}",
        f"Listing page: {public_url}",
    ])
    return "\n".join(filter(lambda value: value is not None, lines))


def public_rental_listing(request, listing_id):
    listing = get_object_or_404(
        RentalListing.objects.select_related("property").prefetch_related("photos"),
        id=listing_id,
        status="published",
    )
    return render(request, "rental_listing_public.html", {
        "listing": listing,
        "apply_url": f"{reverse('apply')}?property={listing.property_id}",
    })


def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect("login")


def signup(request):
    pending_user_id = request.session.get("pending_portal_user_id") or request.session.get("pending_resident_user_id")
    pending_profile_id = request.session.get("pending_resident_profile_id")

    if not pending_user_id:
        messages.error(request, "Please enter your invite code before creating an account.")
        return redirect("enter_invite_code")

    pending_user = get_object_or_404(User, id=pending_user_id)
    profile = None
    owner_intake = None
    landlord_intake_obj = None

    if pending_profile_id:
        profile = get_object_or_404(HousingApplication, id=pending_profile_id)
    elif pending_user.role == "property_owner":
        owner_intake = get_object_or_404(PropertyOwnerIntake, user=pending_user)
    elif pending_user.role == "landlord":
        landlord_intake_obj = get_object_or_404(LandlordIntake, user=pending_user)
    else:
        messages.error(request, "No portal intake is connected to this code yet.")
        return redirect("enter_invite_code")

    if not pending_user.invite_code_is_valid():
        request.session.pop("pending_portal_user_id", None)
        request.session.pop("pending_resident_user_id", None)
        request.session.pop("pending_resident_profile_id", None)
        messages.error(request, "That invite code expired. Request a new code to continue.")
        return redirect("request_invite_code")

    if request.method == "POST":
        form_class = LandlordSignUpForm if pending_user.role == "landlord" else SignUpForm
        form = form_class(request.POST)

        if form.is_valid():
            user = form.save(commit=False)
            user.role = pending_user.role
            user.email = form.cleaned_data.get("email") or pending_user.email
            user.is_staff = user.role in ["landlord", "assistant"]
            user.is_superuser = False
            user.save()

            if profile:
                profile.user = user
                profile.save()

            if owner_intake:
                owner_intake.user = user
                owner_intake.status = "registered"
                owner_intake.save(update_fields=["user", "status"])

            if landlord_intake_obj:
                landlord_intake_obj.full_name = form.cleaned_data.get("full_name", "")
                landlord_intake_obj.phone = form.cleaned_data.get("phone", "")
                landlord_intake_obj.address = form.cleaned_data.get("address", "")
                landlord_intake_obj.user = user
                landlord_intake_obj.status = "registered"
                landlord_intake_obj.save(update_fields=["full_name", "phone", "address", "user", "status"])

            if not pending_user.has_usable_password() and pending_user.id != user.id:
                pending_user.delete()
            else:
                pending_user.mark_invite_code_used()

            request.session.pop("pending_portal_user_id", None)
            request.session.pop("pending_resident_user_id", None)
            request.session.pop("pending_resident_profile_id", None)

            login(request, user)
            messages.success(request, "Your portal account is ready.")
            from .auth_views import dashboard_for_user
            return redirect(dashboard_for_user(user))
    else:
        form_class = LandlordSignUpForm if pending_user.role == "landlord" else SignUpForm
        initial = {"email": pending_user.email}

        if landlord_intake_obj:
            initial.update({
                "full_name": landlord_intake_obj.full_name,
                "phone": landlord_intake_obj.phone,
                "address": landlord_intake_obj.address,
            })

        form = form_class(initial=initial)

    return render(request, "signup.html", {
        "form": form,
        "application": profile,
        "pending_role": pending_user.get_role_display(),
    })


def enter_invite_code(request):
    form = InviteCodeForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        code = form.cleaned_data["invite_code"].upper()
        user_with_code = User.objects.filter(invite_code=code).first()

        if not user_with_code:
            messages.error(request, "Invalid access code.")
            return redirect("enter_invite_code")

        if not user_with_code.invite_code_is_valid():
            messages.error(request, "That invite code expired. Request a new code to continue.")
            return redirect("request_invite_code")

        profile = HousingApplication.objects.filter(user=user_with_code).first()
        if not profile and user_with_code.role == "tenant":
            profile = HousingApplication.objects.filter(email=user_with_code.email).first()

        owner_intake = PropertyOwnerIntake.objects.filter(user=user_with_code).first()
        landlord_intake_obj = LandlordIntake.objects.filter(user=user_with_code).first()

        if not profile and not owner_intake and not landlord_intake_obj:
            messages.error(request, "No approved portal intake is connected to this code yet.")
            return redirect("enter_invite_code")

        request.session["pending_portal_user_id"] = user_with_code.id
        request.session["pending_resident_user_id"] = user_with_code.id
        if profile:
            request.session["pending_resident_profile_id"] = profile.id
        else:
            request.session.pop("pending_resident_profile_id", None)

        messages.success(request, "Invite code accepted. Create your login to finish setup.")
        return redirect("signup")

    return render(request, "enter_invite_code.html", {"form": form})


def request_invite_code(request):
    form = ReplacementInviteCodeForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        email = form.cleaned_data["email"].strip()
        profile = (
            HousingApplication.objects
            .select_related("user")
            .filter(email__iexact=email, user__isnull=False)
            .first()
        )

        if profile and profile.user and not profile.user.has_usable_password():
            try:
                profile.user.refresh_invite_code()
                from .landlord_views import send_resident_invite_email
                send_resident_invite_email(profile)
            except Exception:
                pass
        else:
            portal_intake = (
                PropertyOwnerIntake.objects.select_related("user")
                .filter(email__iexact=email, user__isnull=False)
                .first()
            )
            role_label = "Property Owner"

            if not portal_intake:
                portal_intake = (
                    LandlordIntake.objects.select_related("user")
                    .filter(email__iexact=email, user__isnull=False)
                    .first()
                )
                role_label = "Landlord"

            if portal_intake and portal_intake.user and not portal_intake.user.has_usable_password():
                try:
                    portal_intake.user.refresh_invite_code()
                    from .invite_utils import send_portal_access_invite_email
                    send_portal_access_invite_email(portal_intake.user, portal_intake.full_name, role_label)
                except Exception:
                    pass

        messages.success(
            request,
            "If an approved unregistered portal intake matches that email, a new invite code has been sent.",
        )
        return redirect("enter_invite_code")

    return render(request, "request_invite_code.html", {"form": form})

def staff_managed_properties(user):
    if user.is_superuser or getattr(user, "role", "") in ["admin", "assistant"]:
        return Property.objects.all()

    if getattr(user, "role", "") == "property_owner":
        return Property.objects.filter(owner_email__iexact=user.email)

    if getattr(user, "role", "") == "landlord":
        return Property.objects.filter(landlord_email__iexact=user.email)

    return Property.objects.none()


def staff_managed_applications(user):
    return HousingApplication.objects.filter(property__in=staff_managed_properties(user))


def custom_report_accessible_properties(user):
    return staff_managed_properties(user).order_by("name")


def current_month_bounds():
    today = timezone.localdate()
    month_start = today.replace(day=1)

    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)

    return month_start, next_month


def monthly_collection_watch_rows(applications):
    month_start, next_month = current_month_bounds()
    rows = []

    for application in applications:
        monthly_payments = list(application.payments.filter(status="completed"))

        rent_paid = payment_amount_for_month(monthly_payments, month_start.year, month_start.month, ["rent"])
        utility_paid = payment_amount_for_month(monthly_payments, month_start.year, month_start.month, ["utility"])
        combined_paid = sum(
            (
                amount
                for payment in monthly_payments
                if payment.payment_type == "other" and "combined" in payment.description.lower()
                for accounting_month, amount in payment_month_allocations(payment)
                if accounting_month.year == month_start.year and accounting_month.month == month_start.month
            ),
            Decimal("0.00"),
        )

        rent_expected = expected_rent_for_month(application, month_start)
        utility_expected = expected_utility_for_month(application, month_start)

        if combined_paid > 0:
            rent_shortfall = max(rent_expected - rent_paid, Decimal("0.00"))
            combined_to_rent = min(combined_paid, rent_shortfall)
            rent_paid += combined_to_rent
            combined_paid -= combined_to_rent

            utility_shortfall = max(utility_expected - utility_paid, Decimal("0.00"))
            combined_to_utility = min(combined_paid, utility_shortfall)
            utility_paid += combined_to_utility

        missing_items = []
        rent_due = max(rent_expected - rent_paid, Decimal("0.00"))
        utility_due = max(utility_expected - utility_paid, Decimal("0.00"))

        if rent_expected > 0 and rent_due > 0:
            missing_items.append("Rent")

        if utility_expected > 0 and utility_due > 0:
            missing_items.append("Utilities")

        if missing_items:
            unit_label = canonical_room_label(application.space_label)
            rows.append({
                "application": application,
                "property": application.property.name if application.property else "No Property",
                "unit": unit_label,
                "missing": " + ".join(missing_items),
                "rent_paid": rent_paid,
                "rent_expected": rent_expected,
                "rent_due": rent_due,
                "utility_paid": utility_paid,
                "utility_expected": utility_expected,
                "utility_due": utility_due,
            })

    return sorted(
        rows,
        key=lambda row: resident_sort_key(row["application"]),
    )


def get_landlord_workspace_context(user):
    properties = staff_managed_properties(user).order_by("name")

    resident_files = (
        HousingApplication.objects
        .select_related("property", "user")
        .filter(property__in=properties, user__isnull=False)
        .order_by("property__name", "space_label", "full_name")
    )

    payments = Payment.objects.filter(application__property__in=properties).order_by("-created_at")[:25]

    resident_messages = (
        ResidentMessage.objects
        .select_related("application", "application__property")
        .filter(application__property__in=properties)
        .order_by("application__property__name", "-created_at")
    )

    new_applications_queryset = (
        HousingApplication.objects
        .select_related("property", "user")
        .filter(property__in=properties, user__isnull=True, landlord_reviewed_at__isnull=True)
        .order_by("-created_at")
    )
    new_applications = dedupe_attention_applications(new_applications_queryset)

    new_messages = (
        ResidentMessage.objects
        .select_related("application", "application__property")
        .filter(application__property__in=properties, status="submitted")
        .order_by("-created_at")
    )

    new_documents = (
        ApplicantDocument.objects
        .select_related("application", "application__property")
        .filter(application__property__in=properties, status="uploaded", landlord_notified=False)
        .order_by("-created_at")
    )

    existing_resident_intakes = list(
        ExistingResidentIntake.objects
        .select_related("property")
        .filter(property__in=properties, landlord_reviewed_at__isnull=True)
        .order_by("-created_at")
    )
    existing_resident_rows = []
    existing_resident_seen = {}

    for intake in existing_resident_intakes:
        intake_key = attention_identity_key(
            intake.property_id,
            intake.email,
            intake.full_name(),
            intake.room_unit_label,
        )
        if intake_key in existing_resident_seen:
            existing_resident_seen[intake_key]["duplicate_count"] += 1
            continue

        application = (
            HousingApplication.objects
            .select_related("user")
            .filter(property=intake.property, email__iexact=intake.email)
            .first()
        )
        pending_user = application.user if application and application.user else None

        if pending_user and pending_user.has_usable_password():
            setup_status = "completed"
        elif pending_user:
            setup_status = "invite_sent"
        else:
            setup_status = "ready"

        row = {
            "intake": intake,
            "setup_status": setup_status,
            "duplicate_count": 1,
        }
        existing_resident_seen[intake_key] = row
        existing_resident_rows.append(row)

    landlord_inbox = OrderedDict()

    for resident_message in resident_messages:
        application = resident_message.application
        property_name = "No Property"

        if application and application.property:
            property_name = application.property.name

        landlord_inbox.setdefault(property_name, [])
        landlord_inbox[property_name].append(resident_message)

    new_message_count = new_messages.count()
    collection_watch_rows = monthly_collection_watch_rows(resident_files)
    month_start, _next_month = current_month_bounds()

    return {
        "applications": sorted_resident_list(resident_files),
        "properties": properties,
        "payments": payments,
        "landlord_inbox": landlord_inbox,
        "new_message_count": new_message_count,
        "new_applications": new_applications,
        "new_application_count": len(new_applications),
        "new_messages": new_messages,
        "new_documents": new_documents,
        "new_document_count": new_documents.count(),
        "existing_resident_intakes": existing_resident_rows,
        "existing_resident_intake_count": len(existing_resident_rows),
        "collection_watch_month": month_start,
        "collection_watch_rows": collection_watch_rows,
        "collection_watch_count": len(collection_watch_rows),
        "attention_count": (
            len(new_applications)
            + new_message_count
            + new_documents.count()
            + len(existing_resident_rows)
        ),
    }


@login_required
@user_passes_test(staff_required)
def landlord_dashboard(request):
    return render(request, "landlord_dashboard.html", get_landlord_workspace_context(request.user))


@login_required
@user_passes_test(staff_required)
def landlord_attention(request):
    return render(request, "landlord_attention.html", get_landlord_workspace_context(request.user))


@login_required
@user_passes_test(staff_required)
def landlord_resident_files(request):
    return render(request, "landlord_resident_files.html", get_landlord_workspace_context(request.user))


@login_required
@user_passes_test(staff_required)
def listing_center(request):
    properties = staff_managed_properties(request.user).order_by("name")
    listings = (
        RentalListing.objects
        .filter(property__in=properties)
        .select_related("property", "created_by")
        .prefetch_related("channels", "photos")
        .order_by("-updated_at")
    )
    return render(request, "listing_center.html", {
        "listings": listings,
        "properties": properties,
    })


def ensure_listing_channels(listing):
    for channel, _label in RentalListingChannel.CHANNEL_CHOICES:
        RentalListingChannel.objects.get_or_create(listing=listing, channel=channel)


@login_required
@user_passes_test(staff_required)
def rental_listing_create(request):
    properties = staff_managed_properties(request.user).order_by("name")
    initial = {}
    property_id = request.GET.get("property")
    if property_id:
        property_obj = properties.filter(id=property_id).first()
        if property_obj:
            initial = {
                "property": property_obj,
                "headline": f"{property_obj.name} vacancy",
                "rent_amount": property_obj.rent_amount or Decimal("0.00"),
                "deposit_amount": property_obj.deposit_amount or Decimal("0.00"),
                "utilities_description": property_obj.utilities_cost,
                "lease_terms": property_obj.get_lease_type_display(),
                "available_date": property_obj.available_date,
                "property_benefits": property_obj.description,
                "screening_summary": property_obj.screening_fee_disclosure or property_obj.background_check_instructions,
            }

    form = RentalListingForm(request.POST or None, request.FILES or None, properties=properties, initial=initial)

    if request.method == "POST" and form.is_valid():
        listing = form.save(commit=False)
        listing.created_by = request.user
        if listing.status == "published" and not listing.published_at:
            listing.published_at = timezone.now()
        listing.save()
        for index, image in enumerate(form.cleaned_data.get("photos", []), start=1):
            RentalListingPhoto.objects.create(listing=listing, image=image, sort_order=index)
        ensure_listing_channels(listing)
        messages.success(request, "Rental listing saved.")
        return redirect("rental_listing_detail", listing_id=listing.id)

    return render(request, "rental_listing_form.html", {
        "form": form,
        "title": "Create Rental Listing",
    })


@login_required
@user_passes_test(staff_required)
def rental_listing_detail(request, listing_id):
    listing = get_object_or_404(
        RentalListing.objects.select_related("property", "created_by").prefetch_related("photos", "channels"),
        id=listing_id,
        property__in=staff_managed_properties(request.user),
    )
    ensure_listing_channels(listing)
    channel_forms = [
        (channel, RentalListingChannelForm(prefix=f"channel_{channel.id}", instance=channel))
        for channel in listing.channels.all()
    ]
    return render(request, "rental_listing_detail.html", {
        "listing": listing,
        "channel_forms": channel_forms,
        "posting_copy": build_listing_copy(request, listing),
        "public_url": request.build_absolute_uri(reverse("public_rental_listing", args=[listing.id])),
        "apply_url": request.build_absolute_uri(f"{reverse('apply')}?property={listing.property_id}"),
    })


@login_required
@user_passes_test(staff_required)
def rental_listing_edit(request, listing_id):
    listing = get_object_or_404(
        RentalListing.objects.select_related("property"),
        id=listing_id,
        property__in=staff_managed_properties(request.user),
    )
    old_status = listing.status
    form = RentalListingForm(
        request.POST or None,
        request.FILES or None,
        properties=staff_managed_properties(request.user).order_by("name"),
        instance=listing,
    )

    if request.method == "POST" and form.is_valid():
        listing = form.save(commit=False)
        if listing.status == "published" and old_status != "published":
            listing.published_at = timezone.now()
        if listing.status == "filled" and old_status != "filled":
            listing.filled_at = timezone.now()
        listing.save()
        starting_order = listing.photos.count() + 1
        for offset, image in enumerate(form.cleaned_data.get("photos", []), start=starting_order):
            RentalListingPhoto.objects.create(listing=listing, image=image, sort_order=offset)
        ensure_listing_channels(listing)
        messages.success(request, "Rental listing updated.")
        return redirect("rental_listing_detail", listing_id=listing.id)

    return render(request, "rental_listing_form.html", {
        "form": form,
        "listing": listing,
        "title": "Edit Rental Listing",
    })


@login_required
@user_passes_test(staff_required)
def rental_listing_update_channels(request, listing_id):
    listing = get_object_or_404(
        RentalListing,
        id=listing_id,
        property__in=staff_managed_properties(request.user),
    )
    ensure_listing_channels(listing)
    all_valid = True
    forms = []
    for channel in listing.channels.all():
        form = RentalListingChannelForm(request.POST, prefix=f"channel_{channel.id}", instance=channel)
        forms.append(form)
        all_valid = all_valid and form.is_valid()

    if all_valid:
        for form in forms:
            channel = form.save(commit=False)
            if channel.status == "posted" and not channel.posted_at:
                channel.posted_at = timezone.now()
            channel.save()
        messages.success(request, "Listing channel status updated.")
    else:
        messages.error(request, "One or more channel updates could not be saved.")

    return redirect("rental_listing_detail", listing_id=listing.id)


@login_required
@user_passes_test(staff_required)
def transfer_resident_room(request, application_id):
    application = get_object_or_404(
        HousingApplication.objects.select_related("property"),
        id=application_id,
        property__in=staff_managed_properties(request.user),
    )

    if request.method == "POST":
        form = ResidentRoomTransferForm(request.POST)
        if form.is_valid():
            old_room = f"{application.space_type} {application.space_label}".strip() or "Unassigned"
            new_space_type = form.cleaned_data["space_type"].strip()
            new_space_label = form.cleaned_data["space_label"].strip()
            application.space_type = new_space_type
            application.space_label = new_space_label

            updated_fields = ["space_type", "space_label"]
            room_setting = None
            if form.cleaned_data["apply_room_rent"]:
                room_setting = find_room_rent_setting(application.property, new_space_label)

            if room_setting:
                if application.monthly_rent != room_setting.monthly_rent:
                    RentHistory.objects.create(
                        application=application,
                        rent_amount=room_setting.monthly_rent,
                        effective_date=timezone.localdate(),
                    )
                application.monthly_rent = room_setting.monthly_rent
                application.balance = room_setting.monthly_rent
                application.rent_due_day = room_setting.rent_due_day
                application.utility_monthly = room_setting.utility_monthly
                application.utility_balance = room_setting.utility_monthly
                application.deposit_required = room_setting.deposit_required
                application.deposit_paid = min(application.deposit_paid, room_setting.deposit_required)
                updated_fields.extend([
                    "monthly_rent",
                    "balance",
                    "rent_due_day",
                    "utility_monthly",
                    "utility_balance",
                    "deposit_required",
                    "deposit_paid",
                ])

            notes = form.cleaned_data.get("notes", "").strip()
            if notes:
                application.additional_notes = (
                    f"{application.additional_notes}\n\n"
                    f"{timezone.localdate()}: Room transfer from {old_room} to {new_space_type} {new_space_label}. {notes}"
                ).strip()
                updated_fields.append("additional_notes")

            application.save(update_fields=updated_fields)
            messages.success(request, f"{application.full_name} moved from {old_room} to {new_space_type} {new_space_label}.")
            return redirect("application_detail", pk=application.id)
    else:
        form = ResidentRoomTransferForm(initial={
            "space_type": application.space_type or "Room",
            "space_label": application.space_label,
            "apply_room_rent": True,
        })

    room_setting = find_room_rent_setting(application.property, application.space_label)
    return render(request, "transfer_resident_room.html", {
        "application": application,
        "form": form,
        "current_room_setting": room_setting,
    })


def room_rent_setup_rows(user, property_obj=None):
    properties = staff_managed_properties(user).order_by("name")
    if property_obj:
        properties = properties.filter(id=property_obj.id)
    room_map = OrderedDict()

    roster_entries = (
        CurrentResidentRosterEntry.objects
        .select_related("property")
        .filter(property__in=properties, is_active=True)
        .order_by("property__name", "room_unit_label", "last_name", "first_name")
    )
    for entry in roster_entries:
        label = (entry.room_unit_label or "").strip()
        if not label:
            continue
        key = (entry.property_id, normalized_room_label(label))
        room_map.setdefault(key, {
            "property": entry.property,
            "room_unit_label": canonical_room_label(label),
            "residents": [],
            "setting": None,
        })
        room_map[key]["residents"].append(entry.full_name())

    existing_files = (
        staff_managed_applications(user)
        .select_related("property", "existing_resident_intake")
        .exclude(space_label="")
        .order_by("property__name", "space_label", "full_name")
    )
    for application in sorted_resident_list(existing_files):
        if is_orphan_existing_resident_setup_file(application):
            continue
        if not application.property_id:
            continue
        label = (application.space_label or "").strip()
        if not label:
            continue
        key = (application.property_id, normalized_room_label(label))
        room_map.setdefault(key, {
            "property": application.property,
            "room_unit_label": canonical_room_label(label),
            "residents": [],
            "setting": None,
        })
        if application.full_name not in room_map[key]["residents"]:
            room_map[key]["residents"].append(application.full_name)

    settings = PropertyRoomRent.objects.filter(property__in=properties, is_active=True).order_by("property__name", "room_unit_label", "id")
    for setting in settings:
        key = (setting.property_id, normalized_room_label(setting.room_unit_label))
        room_map.setdefault(key, {
            "property": setting.property,
            "room_unit_label": canonical_room_label(setting.room_unit_label),
            "residents": [],
            "setting": None,
        })
        room_map[key]["room_unit_label"] = canonical_room_label(setting.room_unit_label)
        room_map[key]["setting"] = setting

    return sorted(
        room_map.values(),
        key=lambda row: (
            row["property"].name.lower() if row["property"] else "",
            rent_roll_room_sort_key(row["room_unit_label"]),
            ", ".join(row["residents"]).lower(),
        ),
    )


def is_existing_resident_setup_file(application):
    return (
        application.income_source == "Existing resident intake"
        or application.housing_need == "Existing resident profile setup."
    )


def is_orphan_existing_resident_setup_file(application):
    if not is_existing_resident_setup_file(application):
        return False

    if application.user:
        return False

    try:
        application.existing_resident_intake
    except ExistingResidentIntake.DoesNotExist:
        return True

    return False


def normalized_room_label(room_unit_label):
    label = clean_match_value(room_unit_label)
    for prefix in ["room", "unit", "space", "apt", "apartment"]:
        if label.startswith(prefix) and len(label) > len(prefix):
            return label[len(prefix):]
    return label


def canonical_room_label(room_unit_label):
    label = str(room_unit_label or "").strip()
    clean_label = normalized_room_label(label)
    if not clean_label:
        return label
    if len(clean_label) == 1:
        return clean_label.upper()
    return clean_label.upper() if clean_label.isalpha() else clean_label


def find_room_rent_setting(property_obj, room_unit_label):
    if not property_obj or not room_unit_label:
        return None

    target_label = normalized_room_label(room_unit_label)
    for setting in PropertyRoomRent.objects.filter(property=property_obj, is_active=True):
        if normalized_room_label(setting.room_unit_label) == target_label:
            return setting
    return None


def attention_identity_key(property_id, email, full_name, room_unit_label=""):
    email_key = clean_match_value(email)
    if email_key:
        return ("email", property_id, email_key)

    return (
        "name-room",
        property_id,
        clean_match_value(full_name),
        normalized_room_label(room_unit_label),
    )


def dedupe_attention_applications(applications):
    deduped = []
    seen = {}

    for application in applications:
        key = attention_identity_key(
            application.property_id,
            application.email,
            application.full_name,
            application.space_label,
        )

        if key in seen:
            seen[key].attention_duplicate_count += 1
            continue

        application.attention_duplicate_count = 1
        seen[key] = application
        deduped.append(application)

    return sorted_resident_list(deduped)


def matching_room_rent_settings(property_id, room_unit_label):
    target_label = normalized_room_label(room_unit_label)
    return [
        setting for setting in PropertyRoomRent.objects.filter(property_id=property_id)
        if normalized_room_label(setting.room_unit_label) == target_label
    ]


def save_room_rent_setting(property_id, room_unit_label, defaults):
    clean_room_unit_label = canonical_room_label(room_unit_label)
    matching_settings = matching_room_rent_settings(property_id, room_unit_label)

    if matching_settings:
        primary_setting = next(
            (
                setting for setting in matching_settings
                if str(setting.room_unit_label or "").strip().lower() == clean_room_unit_label.lower()
            ),
            matching_settings[0],
        )
        primary_setting.room_unit_label = clean_room_unit_label
        for field_name, value in defaults.items():
            setattr(primary_setting, field_name, value)
        primary_setting.save(update_fields=list(dict.fromkeys(["room_unit_label", *list(defaults.keys())])))

        for room_setting in matching_settings:
            if room_setting.id == primary_setting.id:
                continue
            for field_name, value in defaults.items():
                setattr(room_setting, field_name, value)
            room_setting.is_active = False
            room_setting.save(update_fields=list(dict.fromkeys(["room_unit_label", *list(defaults.keys()), "is_active"])))
        primary_setting.refresh_from_db()
        return primary_setting

    return PropertyRoomRent.objects.create(
        property_id=property_id,
        room_unit_label=clean_room_unit_label,
        **defaults,
    )


def adjusted_balance_after_amount_change(old_amount, old_balance, new_amount):
    credited_amount = max(Decimal(old_amount or "0.00") - Decimal(old_balance or "0.00"), Decimal("0.00"))
    return max(Decimal(new_amount or "0.00") - credited_amount, Decimal("0.00"))


def apply_room_rent_setting_to_residents(room_setting, residents, effective_date, applied_resident_ids):
    updated_count = 0
    rent_history_count = 0
    room_applied_count = 0
    target_label = normalized_room_label(room_setting.room_unit_label)

    matching_residents = [
        resident for resident in residents
        if resident.id not in applied_resident_ids
        and resident.property_id == room_setting.property_id
        and normalized_room_label(resident.space_label) == target_label
    ]

    for resident in matching_residents:
        changed_fields = []
        rent_balance = adjusted_balance_after_amount_change(
            resident.monthly_rent,
            resident.balance,
            room_setting.monthly_rent,
        )
        utility_balance = adjusted_balance_after_amount_change(
            resident.utility_monthly,
            resident.utility_balance,
            room_setting.utility_monthly,
        )

        if resident.monthly_rent != room_setting.monthly_rent:
            resident.monthly_rent = room_setting.monthly_rent
            changed_fields.append("monthly_rent")
            RentHistory.objects.create(
                application=resident,
                rent_amount=room_setting.monthly_rent,
                effective_date=effective_date,
            )
            rent_history_count += 1

        if resident.balance != rent_balance:
            resident.balance = rent_balance
            changed_fields.append("balance")

        if resident.rent_due_day != room_setting.rent_due_day:
            resident.rent_due_day = room_setting.rent_due_day
            changed_fields.append("rent_due_day")

        if resident.utility_monthly != room_setting.utility_monthly:
            resident.utility_monthly = room_setting.utility_monthly
            changed_fields.append("utility_monthly")

        if resident.utility_balance != utility_balance:
            resident.utility_balance = utility_balance
            changed_fields.append("utility_balance")

        if resident.deposit_required != room_setting.deposit_required:
            resident.deposit_required = room_setting.deposit_required
            changed_fields.append("deposit_required")

        clean_deposit_paid = min(room_setting.deposit_paid, room_setting.deposit_required)
        if resident.deposit_paid != clean_deposit_paid:
            resident.deposit_paid = clean_deposit_paid
            changed_fields.append("deposit_paid")

        if changed_fields:
            resident.save(update_fields=changed_fields)
            updated_count += 1
            room_applied_count += 1
            applied_resident_ids.add(resident.id)

    return updated_count, rent_history_count, room_applied_count


@login_required
@user_passes_test(staff_required)
def landlord_rent_setup(request, property_id=None):
    accessible_properties = staff_managed_properties(request.user).order_by("name")
    selected_property = None
    selected_from_route = bool(property_id)

    if property_id:
        selected_property = get_object_or_404(accessible_properties, id=property_id)
    elif accessible_properties.count() == 1:
        selected_property = accessible_properties.first()

    room_rows = room_rent_setup_rows(request.user, selected_property)
    residents = (
        staff_managed_applications(request.user)
        .select_related("property", "user")
        .order_by("property__name", "space_label", "full_name")
    )
    if selected_property:
        residents = residents.filter(property=selected_property)

    if request.method == "POST":
        updated_count = 0
        rent_history_count = 0
        effective_date = timezone.localdate().replace(day=1)
        apply_room_rents = request.POST.get("apply_room_rents") == "on"
        accessible_property_ids = set(accessible_properties.values_list("id", flat=True))

        try:
            room_count = int(request.POST.get("room_count") or 0)
        except (TypeError, ValueError):
            room_count = 0

        selected_room_update_index = request.POST.get("room_update_index")
        if selected_room_update_index not in [None, ""]:
            try:
                selected_room_update_index = int(selected_room_update_index)
            except (TypeError, ValueError):
                selected_room_update_index = None

        room_setting_count = 0
        room_applied_count = 0
        room_applied_resident_ids = set()
        added_room_keys = set()

        add_room_property_id = request.POST.get("add_room_property_id")
        if selected_property and not add_room_property_id:
            add_room_property_id = str(selected_property.id)
        add_room_unit_label = (request.POST.get("add_room_unit_label") or "").strip()
        save_added_room = request.POST.get("save_added_room") == "1"
        added_room_saved = False

        if add_room_property_id and add_room_unit_label:
            try:
                property_id = int(add_room_property_id)
            except (TypeError, ValueError):
                property_id = None

            if property_id in accessible_property_ids:
                try:
                    add_room_rent_due_day = int(request.POST.get("add_room_rent_due_day") or 1)
                except (TypeError, ValueError):
                    add_room_rent_due_day = 1
                add_room_rent_due_day = min(max(add_room_rent_due_day, 1), 31)

                room_setting = save_room_rent_setting(
                    property_id,
                    add_room_unit_label,
                    {
                        "monthly_rent": money(request.POST.get("add_room_monthly_rent")),
                        "rent_due_day": add_room_rent_due_day,
                        "utility_monthly": money(request.POST.get("add_room_utility_monthly")),
                        "deposit_required": money(request.POST.get("add_room_deposit_required")),
                        "deposit_paid": money(request.POST.get("add_room_deposit_paid")),
                        "is_active": True,
                    },
                )
                added_room_keys.add((property_id, normalized_room_label(add_room_unit_label)))
                room_setting_count += 1

                if apply_room_rents:
                    applied_updates, applied_history, applied_rooms = apply_room_rent_setting_to_residents(
                        room_setting,
                        residents,
                        effective_date,
                        room_applied_resident_ids,
                    )
                    updated_count += applied_updates
                    rent_history_count += applied_history
                    room_applied_count += applied_rooms
                added_room_saved = True

        if save_added_room:
            if added_room_saved:
                messages.success(
                    request,
                    f"Room {canonical_room_label(add_room_unit_label)} rent saved and matching resident files updated.",
                )
            else:
                messages.error(request, "Choose a space/unit label before saving rent.")
            if selected_property and selected_from_route:
                return redirect("landlord_rent_setup_property", property_id=selected_property.id)
            return redirect("landlord_rent_setup")

        for index in range(room_count):
            if selected_room_update_index is not None and index != selected_room_update_index:
                continue

            prefix = f"room_{index}_"

            try:
                property_id = int(request.POST.get(prefix + "property_id") or 0)
            except (TypeError, ValueError):
                continue

            if property_id not in accessible_property_ids:
                continue

            room_unit_label = (request.POST.get(prefix + "room_unit_label") or "").strip()
            if not room_unit_label:
                continue
            if (property_id, normalized_room_label(room_unit_label)) in added_room_keys:
                continue

            monthly_rent = money(request.POST.get(prefix + "monthly_rent"))
            utility_monthly = money(request.POST.get(prefix + "utility_monthly"))
            deposit_required = money(request.POST.get(prefix + "deposit_required"))
            deposit_paid = money(request.POST.get(prefix + "deposit_paid"))

            try:
                rent_due_day = int(request.POST.get(prefix + "rent_due_day") or 1)
            except (TypeError, ValueError):
                rent_due_day = 1
            rent_due_day = min(max(rent_due_day, 1), 31)

            room_setting = save_room_rent_setting(
                property_id,
                room_unit_label,
                {
                    "monthly_rent": monthly_rent,
                    "rent_due_day": rent_due_day,
                    "utility_monthly": utility_monthly,
                    "deposit_required": deposit_required,
                    "deposit_paid": deposit_paid,
                    "is_active": True,
                },
            )
            room_setting_count += 1

            if apply_room_rents:
                applied_updates, applied_history, applied_rooms = apply_room_rent_setting_to_residents(
                    room_setting,
                    residents,
                    effective_date,
                    room_applied_resident_ids,
                )
                updated_count += applied_updates
                rent_history_count += applied_history
                room_applied_count += applied_rooms

        for resident in residents:
            if resident.id in room_applied_resident_ids:
                continue

            prefix = f"resident_{resident.id}_"
            if prefix + "monthly_rent" not in request.POST:
                continue

            monthly_rent = money(request.POST.get(prefix + "monthly_rent"))
            rent_balance = money(request.POST.get(prefix + "balance"))
            utility_monthly = money(request.POST.get(prefix + "utility_monthly"))
            utility_balance = money(request.POST.get(prefix + "utility_balance"))
            deposit_required = money(request.POST.get(prefix + "deposit_required"))
            deposit_paid = money(request.POST.get(prefix + "deposit_paid"))

            try:
                rent_due_day = int(request.POST.get(prefix + "rent_due_day") or resident.rent_due_day or 1)
            except (TypeError, ValueError):
                rent_due_day = resident.rent_due_day or 1
            rent_due_day = min(max(rent_due_day, 1), 31)

            changed_fields = []
            if resident.monthly_rent != monthly_rent:
                resident.monthly_rent = monthly_rent
                changed_fields.append("monthly_rent")
                RentHistory.objects.create(
                    application=resident,
                    rent_amount=monthly_rent,
                    effective_date=effective_date,
                )
                rent_history_count += 1

            if resident.balance != rent_balance:
                resident.balance = rent_balance
                changed_fields.append("balance")

            if resident.rent_due_day != rent_due_day:
                resident.rent_due_day = rent_due_day
                changed_fields.append("rent_due_day")

            if resident.utility_monthly != utility_monthly:
                resident.utility_monthly = utility_monthly
                changed_fields.append("utility_monthly")

            if resident.utility_balance != utility_balance:
                resident.utility_balance = utility_balance
                changed_fields.append("utility_balance")

            if resident.deposit_required != deposit_required:
                resident.deposit_required = deposit_required
                changed_fields.append("deposit_required")

            clean_deposit_paid = min(deposit_paid, deposit_required)
            if resident.deposit_paid != clean_deposit_paid:
                resident.deposit_paid = clean_deposit_paid
                changed_fields.append("deposit_paid")

            if changed_fields:
                resident.save(update_fields=changed_fields)
                updated_count += 1

        messages.success(
            request,
            f"Rent setup saved for {updated_count} resident file(s). {room_setting_count} room rent setting(s) saved. {room_applied_count} resident file(s) updated from room rent. {rent_history_count} rent history record(s) added.",
        )
        if selected_property and selected_from_route:
            return redirect("landlord_rent_setup_property", property_id=selected_property.id)
        return redirect("landlord_rent_setup")

    return render(request, "landlord_rent_setup.html", {
        "properties": accessible_properties,
        "selected_property": selected_property,
        "room_rows": room_rows,
        "residents": sorted_resident_list(residents),
    })


def get_superadmin_workspace_context():
    properties = Property.objects.all().order_by("name")
    users = User.objects.all().order_by("username")
    applications = (
        HousingApplication.objects
        .select_related("property", "user")
        .filter(user__isnull=False)
        .order_by("property__name", "space_label", "full_name")
    )
    completed_payments = Payment.objects.filter(status="completed")
    site_payment_total = completed_payments.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    owner_buckets = OrderedDict()

    for property_obj in properties:
        owner_email = (property_obj.owner_email or "").strip()
        owner_label = owner_email or "Unassigned Owner"

        owner_buckets.setdefault(owner_label, [])
        owner_buckets[owner_label].append(property_obj)

    owner_groups = [
        {
            "email": owner_label,
            "property_count": len(owner_properties),
            "properties": owner_properties,
        }
        for owner_label, owner_properties in owner_buckets.items()
    ]
    
    recent_messages = (
        ResidentMessage.objects
        .select_related("application", "application__property")
        .all()
        .order_by("-created_at")[:10]
)
    
    context = {
        "properties": properties,
        "users": users,
        "applications": sorted_resident_list(applications),
        "recent_messages": recent_messages,
        "owner_groups": owner_groups,
        "site_payment_total": site_payment_total,
    }
    context.update(company_mailbox_context())

    return context


GRAPH_SCOPE = "offline_access Mail.ReadWrite Mail.Send User.Read"
GRAPH_API_BASE = "https://graph.microsoft.com/v1.0"


def microsoft_graph_configured():
    return bool(
        settings.MICROSOFT_GRAPH_CLIENT_ID
        and settings.MICROSOFT_GRAPH_CLIENT_SECRET
        and settings.MICROSOFT_GRAPH_REDIRECT_URI
        and settings.MICROSOFT_GRAPH_MAILBOX_USER
    )


def microsoft_graph_authority_url(path):
    tenant = settings.MICROSOFT_GRAPH_TENANT_ID or "common"
    return f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/{path}"


def microsoft_token_request(data):
    request = Request(
        microsoft_graph_authority_url("token"),
        data=urlencode(data).encode("utf-8"),
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urlopen(request, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def graph_request(connection, method, path, data=None):
    token = get_company_mailbox_access_token(connection)
    headers = {"Authorization": f"Bearer {token}"}
    body = None

    if data is not None:
        headers["Content-Type"] = "application/json"
        body = json.dumps(data).encode("utf-8")

    request = Request(
        f"{GRAPH_API_BASE}{path}",
        data=body,
        headers=headers,
        method=method,
    )

    try:
        with urlopen(request, timeout=20) as response:
            raw = response.read()
            if not raw:
                return {}
            return json.loads(raw.decode("utf-8"))
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Microsoft Graph error {exc.code}: {detail}") from exc


def get_company_mailbox_connection():
    mailbox_email = settings.MICROSOFT_GRAPH_MAILBOX_USER or settings.EMAIL_HOST_USER or ""
    if not mailbox_email:
        return None

    connection, _ = CompanyMailboxConnection.objects.get_or_create(mailbox_email=mailbox_email)
    return connection


def get_company_mailbox_access_token(connection):
    if connection.access_token and connection.token_expires_at and connection.token_expires_at > timezone.now() + timedelta(minutes=5):
        return connection.access_token

    if not connection.refresh_token:
        raise RuntimeError("Company mailbox is not connected yet.")

    token_data = microsoft_token_request({
        "client_id": settings.MICROSOFT_GRAPH_CLIENT_ID,
        "client_secret": settings.MICROSOFT_GRAPH_CLIENT_SECRET,
        "grant_type": "refresh_token",
        "refresh_token": connection.refresh_token,
        "redirect_uri": settings.MICROSOFT_GRAPH_REDIRECT_URI,
        "scope": GRAPH_SCOPE,
    })

    expires_in = int(token_data.get("expires_in") or 3600)
    connection.access_token = token_data.get("access_token", "")
    connection.refresh_token = token_data.get("refresh_token") or connection.refresh_token
    connection.token_expires_at = timezone.now() + timedelta(seconds=expires_in)
    connection.save(update_fields=["access_token", "refresh_token", "token_expires_at", "updated_at"])

    return connection.access_token


def company_mailbox_context():
    connection = get_company_mailbox_connection()
    return {
        "mailbox_connection": connection,
        "mailbox_configured": microsoft_graph_configured(),
        "mailbox_email": settings.MICROSOFT_GRAPH_MAILBOX_USER or settings.EMAIL_HOST_USER,
    }


class EmailBodyTextParser(HTMLParser):
    block_tags = {"br", "div", "p", "li", "tr", "table", "section", "article", "h1", "h2", "h3", "h4"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []

    def handle_starttag(self, tag, attrs):
        if tag in self.block_tags:
            self.parts.append("\n")

    def handle_endtag(self, tag):
        if tag in self.block_tags:
            self.parts.append("\n")

    def handle_data(self, data):
        if data:
            self.parts.append(data)

    def text(self):
        return "".join(self.parts)


def clean_email_body(content, content_type="html"):
    text = html.unescape(content or "")

    if (content_type or "").lower() == "html":
        text = re.sub(r"<!--.*?-->", " ", text, flags=re.DOTALL)
        text = re.sub(r"<(style|script|head|meta|title)[^>]*>.*?</\1>", " ", text, flags=re.DOTALL | re.IGNORECASE)
        parser = EmailBodyTextParser()
        parser.feed(text)
        text = parser.text()
    else:
        text = strip_tags(text)

    text = html.unescape(text)
    text = "".join(
        ""
        if unicodedata.category(character) in {"Cf", "Cc"} and character not in "\n\r\t"
        else character
        for character in text
    )
    text = text.replace("\xa0", " ")
    text = re.sub(r"\[\[https?://[^\]]+\]\]", "", text)
    text = re.sub(r"https?://\S*utm_\S+", "", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    cleaned_lines = []
    seen_lines = set()
    for line in text.splitlines():
        clean_line = line.strip()
        if not clean_line:
            cleaned_lines.append("")
            continue
        if len(clean_line) <= 2 and not clean_line.isalnum():
            continue
        if "{" in clean_line or "}" in clean_line:
            continue
        if clean_line.startswith(("@media", ".", "#")):
            continue
        if clean_line.lower().startswith(("padding-", "line-height-", "background-", "mj-", "mso-")):
            continue
        if re.fullmatch(r"\[.*?(icon|stripe).*?\]", clean_line, flags=re.IGNORECASE):
            continue
        line_key = clean_line.lower()
        if line_key in seen_lines:
            continue
        seen_lines.add(line_key)
        cleaned_lines.append(clean_line)

    normalized_lines = []
    previous_blank = False
    for line in cleaned_lines:
        is_blank = not line.strip()
        if is_blank and previous_blank:
            continue
        normalized_lines.append(line)
        previous_blank = is_blank

    practical_text = "\n".join(normalized_lines).strip()
    if len(practical_text) > 5000:
        practical_text = practical_text[:5000].rsplit("\n", 1)[0].strip()
        practical_text += "\n\n[Message shortened. Open in Outlook to view the full email.]"

    return practical_text


def parse_graph_message(message):
    sender = message.get("from", {}).get("emailAddress", {})
    body = message.get("body", {})
    body_content = body.get("content", "")
    body_type = body.get("contentType", "html")

    return {
        "id": message.get("id"),
        "subject": message.get("subject") or "(No subject)",
        "sender_name": sender.get("name") or sender.get("address") or "Unknown sender",
        "sender_email": sender.get("address", ""),
        "received": message.get("receivedDateTime", ""),
        "preview": message.get("bodyPreview", ""),
        "body_text": clean_email_body(body_content, body_type),
        "is_read": message.get("isRead"),
        "web_link": message.get("webLink", ""),
    }


@login_required
@user_passes_test(staff_required)
def superadmin_dashboard(request):

    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    return render(
        request,
        "superadmin_dashboard.html",
        get_superadmin_workspace_context()
    )


@login_required
@user_passes_test(staff_required)
def company_mailbox_connect(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    if not microsoft_graph_configured():
        messages.error(request, "Microsoft mailbox integration is not configured in Render yet.")
        return redirect("company_mailbox")

    state = secrets.token_urlsafe(24)
    request.session["microsoft_graph_oauth_state"] = state

    auth_params = {
        "client_id": settings.MICROSOFT_GRAPH_CLIENT_ID,
        "response_type": "code",
        "redirect_uri": settings.MICROSOFT_GRAPH_REDIRECT_URI,
        "response_mode": "query",
        "scope": GRAPH_SCOPE,
        "state": state,
        "prompt": "select_account",
    }

    return redirect(f"{microsoft_graph_authority_url('authorize')}?{urlencode(auth_params)}")


@login_required
@user_passes_test(staff_required)
def company_mailbox_callback(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    if request.GET.get("state") != request.session.pop("microsoft_graph_oauth_state", None):
        messages.error(request, "Microsoft mailbox connection failed because the security state did not match.")
        return redirect("company_mailbox")

    code = request.GET.get("code")
    if not code:
        messages.error(request, request.GET.get("error_description") or "Microsoft did not return an authorization code.")
        return redirect("company_mailbox")

    try:
        token_data = microsoft_token_request({
            "client_id": settings.MICROSOFT_GRAPH_CLIENT_ID,
            "client_secret": settings.MICROSOFT_GRAPH_CLIENT_SECRET,
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": settings.MICROSOFT_GRAPH_REDIRECT_URI,
            "scope": GRAPH_SCOPE,
        })
    except Exception as exc:
        messages.error(request, f"Microsoft mailbox connection failed: {exc}")
        return redirect("company_mailbox")

    expires_in = int(token_data.get("expires_in") or 3600)
    connection = get_company_mailbox_connection()
    connection.access_token = token_data.get("access_token", "")
    connection.refresh_token = token_data.get("refresh_token", "")
    connection.token_expires_at = timezone.now() + timedelta(seconds=expires_in)
    connection.connected_by = request.user
    connection.save(update_fields=["access_token", "refresh_token", "token_expires_at", "connected_by", "updated_at"])

    messages.success(request, "Company mailbox connected.")
    return redirect("company_mailbox")


@login_required
@user_passes_test(staff_required)
def company_mailbox(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    context = company_mailbox_context()
    messages_list = []
    error_message = ""

    connection = context["mailbox_connection"]
    if context["mailbox_configured"] and connection and connection.is_connected:
        query = urlencode({
            "$top": "25",
            "$orderby": "receivedDateTime desc",
            "$select": "id,subject,from,receivedDateTime,isRead,bodyPreview,webLink",
        })
        try:
            response = graph_request(connection, "GET", f"/me/messages?{query}")
            messages_list = [parse_graph_message(message) for message in response.get("value", [])]
        except Exception as exc:
            error_message = str(exc)

    context.update({
        "messages_list": messages_list,
        "error_message": error_message,
    })
    return render(request, "company_mailbox.html", context)


@login_required
@user_passes_test(staff_required)
def company_mailbox_message(request, message_id):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    connection = get_company_mailbox_connection()
    if not connection or not connection.is_connected:
        messages.error(request, "Connect the company mailbox first.")
        return redirect("company_mailbox")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "delete":
            try:
                graph_request(connection, "DELETE", f"/me/messages/{message_id}")
            except Exception as exc:
                messages.error(request, f"Email delete failed: {exc}")
                return redirect("company_mailbox_message", message_id=message_id)
            messages.success(request, "Email deleted from company mailbox.")
            return redirect("company_mailbox")

        form = CompanyEmailReplyForm(request.POST)
        if form.is_valid():
            try:
                graph_request(connection, "POST", f"/me/messages/{message_id}/reply", {
                    "comment": form.cleaned_data["body"],
                })
            except Exception as exc:
                messages.error(request, f"Reply failed: {exc}")
            else:
                messages.success(request, "Reply sent from company mailbox.")
                return redirect("company_mailbox_message", message_id=message_id)
    else:
        form = CompanyEmailReplyForm()

    query = urlencode({
        "$select": "id,subject,from,toRecipients,receivedDateTime,isRead,body,bodyPreview,webLink",
    })
    try:
        raw_message = graph_request(connection, "GET", f"/me/messages/{message_id}?{query}")
        message = parse_graph_message(raw_message)
        if raw_message.get("isRead") is False:
            graph_request(connection, "PATCH", f"/me/messages/{message_id}", {"isRead": True})
    except Exception as exc:
        messages.error(request, f"Could not load email: {exc}")
        return redirect("company_mailbox")

    return render(request, "company_mailbox_message.html", {
        **company_mailbox_context(),
        "message": message,
        "form": form,
    })


@login_required
@user_passes_test(staff_required)
def company_mailbox_compose(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    connection = get_company_mailbox_connection()
    if not connection or not connection.is_connected:
        messages.error(request, "Connect the company mailbox first.")
        return redirect("company_mailbox")

    form = CompanyEmailComposeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            graph_request(connection, "POST", "/me/sendMail", {
                "message": {
                    "subject": form.cleaned_data["subject"],
                    "body": {
                        "contentType": "Text",
                        "content": form.cleaned_data["body"],
                    },
                    "toRecipients": [
                        {"emailAddress": {"address": form.cleaned_data["to_email"]}}
                    ],
                },
                "saveToSentItems": True,
            })
        except Exception as exc:
            messages.error(request, f"Email send failed: {exc}")
        else:
            messages.success(request, "Email sent from company mailbox.")
            return redirect("company_mailbox")

    return render(request, "company_mailbox_compose.html", {
        **company_mailbox_context(),
        "form": form,
    })


@login_required
@user_passes_test(staff_required)
def superadmin_owners(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    return render(request, "superadmin_owners.html", get_superadmin_workspace_context())


@login_required
@user_passes_test(staff_required)
def superadmin_owner_intakes(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    owner_intakes = PropertyOwnerIntake.objects.select_related("user").all()
    return render(request, "superadmin_owner_intakes.html", {
        "submitted_owner_intakes": owner_intakes.filter(status="submitted").order_by("follow_up_date", "-created_at"),
        "active_leads": owner_intakes.exclude(lead_stage__in=["closed_won", "closed_lost"]).exclude(status="submitted").order_by("follow_up_date", "-created_at"),
        "closed_leads": owner_intakes.filter(lead_stage__in=["closed_won", "closed_lost"]).order_by("-created_at"),
    })


@login_required
@user_passes_test(staff_required)
def superadmin_owner_intake_detail(request, intake_id):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    intake = get_object_or_404(PropertyOwnerIntake.objects.select_related("user"), id=intake_id)

    if request.method == "POST" and request.POST.get("action") == "update_lead_pipeline":
        pipeline_form = PropertyOwnerLeadPipelineForm(request.POST, instance=intake)
        if pipeline_form.is_valid():
            pipeline_form.save()
            messages.success(request, "Owner lead pipeline updated.")
            return redirect("superadmin_owner_intake_detail", intake_id=intake.id)
    else:
        pipeline_form = PropertyOwnerLeadPipelineForm(instance=intake)

    owner_properties = Property.objects.filter(owner_email__iexact=intake.email).order_by("name")
    return render(request, "superadmin_owner_intake_detail.html", {
        "intake": intake,
        "pipeline_form": pipeline_form,
        "owner_properties": owner_properties,
    })


@login_required
@user_passes_test(staff_required)
def superadmin_send_owner_invite(request, intake_id):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    if request.method != "POST":
        return redirect("superadmin_owner_intake_detail", intake_id=intake_id)

    intake = get_object_or_404(PropertyOwnerIntake.objects.select_related("user"), id=intake_id)
    user = intake.user

    if user and user.has_usable_password():
        intake.status = "registered"
        intake.save(update_fields=["status"])
        messages.info(request, "This property owner already has a registered portal login.")
        return redirect("superadmin_owner_intake_detail", intake_id=intake.id)

    if not user:
        user = create_pending_portal_user(intake.full_name, intake.email, "property_owner", intake.id)
        intake.user = user

    user.refresh_invite_code()
    intake.status = "invited"
    intake.invite_sent_at = timezone.now()
    intake.save(update_fields=["user", "status", "invite_sent_at"])

    try:
        sent = send_portal_access_invite_email(user, intake.full_name, "Property Owner")
    except Exception as exc:
        messages.warning(request, f"Owner setup code created, but email failed: {exc}")
    else:
        if sent:
            messages.success(request, "Owner setup invite email sent.")
        else:
            messages.warning(request, "Owner setup code created, but this intake has no email to send.")

    messages.info(request, f"Backup owner setup code: {user.invite_code}")
    return redirect("superadmin_owner_intake_detail", intake_id=intake.id)


@login_required
@user_passes_test(staff_required)
def superadmin_residents(request):
    if not request.user.is_superuser and request.user.role != "admin":
        return redirect("tenant_dashboard")

    return render(request, "superadmin_residents.html", get_superadmin_workspace_context())
@login_required
@user_passes_test(staff_required)
def landlord_message_detail(request, message_id):
    resident_message = get_object_or_404(
        ResidentMessage.objects.select_related("application", "application__property").prefetch_related("replies", "replies__sender"),
        id=message_id,
        application__in=staff_managed_applications(request.user),
    )

    if request.method == "GET" and resident_message.status == "submitted":
        resident_message.status = "reviewed"
        resident_message.save(update_fields=["status"])

    if request.method == "POST":
        reply_body = request.POST.get("reply_body", "").strip()
        if reply_body:
            ResidentMessageReply.objects.create(
                message=resident_message,
                sender=request.user,
                body=reply_body,
                visible_to_resident=True,
            )
            if resident_message.status == "submitted":
                resident_message.status = "reviewed"
                resident_message.save(update_fields=["status"])
            try:
                email_sent = notify_resident_of_portal_reply(request, resident_message)
            except Exception as exc:
                messages.warning(
                    request,
                    f"Reply saved to resident portal, but email notification failed: {exc}",
                )
            else:
                if email_sent:
                    messages.success(request, "Reply sent to resident portal and email notification sent.")
                else:
                    messages.warning(request, "Reply saved to resident portal. No email is on file for this resident.")
            sms_log = notify_resident_of_portal_reply_sms(request, resident_message)
            if sms_log.status == "sent":
                messages.success(request, "Text notification sent.")
            elif sms_log.status in ["skipped_no_consent", "not_configured"]:
                messages.info(request, f"Text notification not sent: {sms_log.get_status_display()}.")
            else:
                messages.warning(request, f"Text notification failed: {sms_log.error_message}")
            return redirect("landlord_message_detail", message_id=resident_message.id)

        new_status = request.POST.get("status")

        if new_status in ["submitted", "reviewed", "closed"]:
            resident_message.status = new_status
            resident_message.save()
            messages.success(request, "Message status updated.")

        return redirect("landlord_message_detail", message_id=resident_message.id)

    return render(request, "landlord_message_detail.html", {
        "resident_message": resident_message,
        "application": resident_message.application,
    })


@login_required
@user_passes_test(reporting_required)
def group_resident_message(request):
    properties = staff_managed_properties(request.user).order_by("name")
    form = GroupResidentMessageForm(request.POST or None, properties=properties)
    preview_count = staff_managed_applications(request.user).filter(user__isnull=False).count()

    if request.method == "POST" and form.is_valid():
        selected_property_id = form.cleaned_data["property_id"]
        target_properties = properties

        if selected_property_id != "all":
            target_properties = properties.filter(id=selected_property_id)

        recipients = (
            HousingApplication.objects
            .select_related("property", "user")
            .filter(property__in=target_properties, user__isnull=False)
            .order_by("property__name", "space_label", "full_name")
        )
        created_count = 0
        sms_attempt_count = 0

        for application in recipients:
            resident_message = ResidentMessage.objects.create(
                application=application,
                message_type="general",
                subject=form.cleaned_data["subject"],
                message=form.cleaned_data["message"],
                status="reviewed",
                locked=True,
            )
            created_count += 1
            send_resident_portal_notification_email(
                request,
                application,
                f"New secure portal message: {form.cleaned_data['subject']}",
                "You have a new secure message in your Bowling Legacy resident portal.",
                "resident_requests",
            )

            if form.cleaned_data["delivery_method"] == "portal_sms":
                send_sms_message(
                    application,
                    sms_body(form.cleaned_data["subject"], form.cleaned_data["message"]),
                    request.user,
                    resident_message=resident_message,
                )
                sms_attempt_count += 1

        if sms_attempt_count:
            messages.success(request, f"Secure portal message sent to {created_count} resident file(s). SMS attempted for {sms_attempt_count} resident file(s).")
        else:
            messages.success(request, f"Secure portal message sent to {created_count} resident file(s).")
        return redirect("group_resident_message")

    return render(request, "group_resident_message.html", {
        "form": form,
        "preview_count": preview_count,
    })


@csrf_exempt
def twilio_sms_webhook(request):
    if request.method != "POST":
        return HttpResponse("", content_type="text/xml")

    from_phone = request.POST.get("From", "")
    body = request.POST.get("Body", "").strip().lower()

    if body in {"stop", "stopall", "unsubscribe", "cancel", "end", "quit"}:
        phone_digits = normalize_phone_digits(from_phone)
        applications = HousingApplication.objects.filter(sms_opted_in=True)
        for application in applications:
            if normalize_phone_digits(application.phone) == phone_digits:
                application.sms_opted_in = False
                application.sms_opted_out_at = timezone.now()
                application.save(update_fields=["sms_opted_in", "sms_opted_out_at"])

    return HttpResponse("<Response></Response>", content_type="text/xml")


@login_required
@user_passes_test(staff_required)
def mark_document_reviewed(request, document_id):
    document = get_object_or_404(
        ApplicantDocument.objects.select_related("application"),
        id=document_id,
        application__in=staff_managed_applications(request.user),
    )

    document.landlord_notified = True
    document.save(update_fields=["landlord_notified"])

    messages.success(request, f"{document.name} marked reviewed.")
    return redirect("landlord_dashboard")


@login_required
@user_passes_test(staff_required)
def open_applicant_document(request, document_id):
    document = get_object_or_404(
        ApplicantDocument.objects.select_related("application"),
        id=document_id,
        application__in=staff_managed_applications(request.user),
    )

    if not document.landlord_notified:
        document.landlord_notified = True
        document.save(update_fields=["landlord_notified"])

    if document.file:
        return redirect(document.file.url)

    messages.warning(request, "This document file is not available.")
    return redirect("landlord_attention")


@login_required
def tenant_dashboard(request):
    request.session.set_expiry(0)

    application, is_superadmin_inspecting = get_resident_portal_application(request)

    payments = []
    resident_messages = []
    profile_photo_form = None
    property_blog_posts = []
    total_due = Decimal("0.00")
    rent_due = Decimal("0.00")
    deposit_due = Decimal("0.00")
    utility_due = Decimal("0.00")
    show_utilities = False
    utility_setup_items = []
    utility_setup_complete = False

    if application:
        payments = application.payments.all().order_by("-created_at")
        resident_messages = application.resident_messages.all().order_by("-created_at")
        if not is_superadmin_inspecting:
            profile_photo_form = ResidentProfilePhotoForm(instance=application)

        rent_due = application.balance if application.balance > 0 else Decimal("0.00")
        deposit_due = max(application.deposit_required - application.deposit_paid, Decimal("0.00"))
        utility_due = application.utility_balance if application.utility_balance > 0 else Decimal("0.00")
        show_utilities = resident_has_portal_utility_charge(application)
        utility_setup_items = resident_utility_setup_items(application)
        utility_setup_complete = bool(utility_setup_items) and all(item.completed_at for item in utility_setup_items)

        total_due = rent_due + deposit_due + utility_due
        if application.property:
            property_blog_posts = (
                application.property.blog_posts
                .prefetch_related("comments")
                .select_related("author")
                .order_by("-created_at")[:3]
            )

    return render(request, "tenant_dashboard.html", {
        "application": application,
        "payments": payments,
        "resident_messages": resident_messages,
        "property_blog_posts": property_blog_posts,
        "profile_photo_form": profile_photo_form,
        "is_superadmin_inspecting": is_superadmin_inspecting,
        "total_due": total_due,
        "rent_due": rent_due,
        "deposit_due": deposit_due,
        "utility_due": utility_due,
        "show_utilities": show_utilities,
        "utility_setup_items": utility_setup_items,
        "utility_setup_complete": utility_setup_complete,
        "stripe_public_key": settings.STRIPE_PUBLIC_KEY,
        "resident_balance_url": resident_portal_url("resident_balance_detail", is_superadmin_inspecting, application),
        "resident_payment_history_url": resident_portal_url("resident_payment_history", is_superadmin_inspecting, application),
        "resident_requests_url": resident_portal_url("resident_requests", is_superadmin_inspecting, application),
    })


def get_resident_portal_application(request):
    application = getattr(request.user, "resident_profile", None)
    is_superadmin_inspecting = False

    if (request.user.is_superuser or getattr(request.user, "role", "") == "admin") and request.GET.get("resident"):
        application = get_object_or_404(
            HousingApplication.objects.select_related("property", "user"),
            id=request.GET.get("resident"),
        )
        is_superadmin_inspecting = True

    return application, is_superadmin_inspecting


def resident_portal_url(view_name, is_superadmin_inspecting, application):
    url = reverse(view_name)
    if is_superadmin_inspecting and application:
        return f"{url}?resident={application.id}"
    return url


def resident_has_portal_utility_charge(application):
    if not application:
        return False
    return application.utility_balance > 0 or application.utility_monthly > 0


def resident_utility_setup_items(application):
    if not application or not application.property_id:
        return []

    vendors = list(
        PropertyUtilityVendor.objects
        .filter(property=application.property, is_active=True)
        .order_by("sort_order", "service_type", "provider_name")
    )

    for vendor in vendors:
        ResidentUtilitySetup.objects.get_or_create(application=application, vendor=vendor)

    return list(
        ResidentUtilitySetup.objects
        .select_related("vendor")
        .filter(application=application, vendor__in=vendors)
        .order_by("vendor__sort_order", "vendor__service_type", "vendor__provider_name")
    )


def notify_landlord_utility_setup_complete(application):
    if not application:
        return

    subject = "Utility setup completed"
    if ResidentMessage.objects.filter(application=application, subject=subject, message_type="general").exists():
        return

    ResidentMessage.objects.create(
        application=application,
        message_type="general",
        subject=subject,
        message=(
            f"{application.full_name} has opened every required tenant utility setup item "
            f"for {application.property.name if application.property else 'their property'}."
        ),
        status="submitted",
        locked=True,
    )


@login_required
def resident_utility_setup_open(request, setup_id):
    application, is_superadmin_inspecting = get_resident_portal_application(request)

    setup_queryset = ResidentUtilitySetup.objects.select_related("vendor", "application", "application__property")

    if application:
        setup = get_object_or_404(setup_queryset, id=setup_id, application=application)
    elif staff_required(request.user):
        setup = get_object_or_404(
            setup_queryset,
            id=setup_id,
            application__property__in=staff_managed_properties(request.user),
        )
        application = setup.application
        is_superadmin_inspecting = True
    else:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    now = timezone.now()
    update_fields = []

    if not setup.opened_at:
        setup.opened_at = now
        update_fields.append("opened_at")

    if not setup.completed_at:
        setup.completed_at = now
        update_fields.append("completed_at")

    if update_fields:
        setup.save(update_fields=update_fields)

    remaining = ResidentUtilitySetup.objects.filter(
        application=application,
        vendor__property=application.property,
        vendor__is_active=True,
        completed_at__isnull=True,
    ).exists()

    if not remaining:
        notify_landlord_utility_setup_complete(application)

    if setup.vendor.setup_url:
        return redirect(setup.vendor.setup_url)

    messages.success(request, f"{setup.vendor.service_type} setup marked complete.")
    return redirect(resident_portal_url("tenant_dashboard", is_superadmin_inspecting, application))


@login_required
def resident_balance_detail(request):
    application, is_superadmin_inspecting = get_resident_portal_application(request)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    rent_due = application.balance if application.balance > 0 else Decimal("0.00")
    deposit_due = max(application.deposit_required - application.deposit_paid, Decimal("0.00"))
    utility_due = application.utility_balance if application.utility_balance > 0 else Decimal("0.00")
    show_utilities = resident_has_portal_utility_charge(application)

    return render(request, "resident_balance_detail.html", {
        "application": application,
        "rent_due": rent_due,
        "deposit_due": deposit_due,
        "utility_due": utility_due,
        "show_utilities": show_utilities,
        "total_due": rent_due + deposit_due + utility_due,
        "dashboard_url": resident_portal_url("tenant_dashboard", is_superadmin_inspecting, application),
        "is_superadmin_inspecting": is_superadmin_inspecting,
    })


@login_required
def resident_payment_history(request):
    application, is_superadmin_inspecting = get_resident_portal_application(request)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    return render(request, "resident_payment_history.html", {
        "application": application,
        "payments": application.payments.all().order_by("-created_at"),
        "dashboard_url": resident_portal_url("tenant_dashboard", is_superadmin_inspecting, application),
    })


@login_required
def resident_requests(request):
    application, is_superadmin_inspecting = get_resident_portal_application(request)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    if request.method == "POST":
        message_id = request.POST.get("message_id")
        reply_body = request.POST.get("reply_body", "").strip()
        resident_message = get_object_or_404(
            ResidentMessage,
            id=message_id,
            application=application,
        )

        if not reply_body:
            messages.error(request, "Reply cannot be blank.")
            return redirect("resident_requests")

        ResidentMessageReply.objects.create(
            message=resident_message,
            sender=request.user,
            body=reply_body,
            visible_to_resident=True,
        )
        resident_message.status = "submitted"
        resident_message.save(update_fields=["status"])
        messages.success(request, "Reply sent.")
        return redirect("resident_requests")

    return render(request, "resident_requests.html", {
        "application": application,
        "requests": application.resident_messages.prefetch_related("replies", "replies__sender").all().order_by("-created_at"),
        "dashboard_url": resident_portal_url("tenant_dashboard", is_superadmin_inspecting, application),
    })


@login_required
def update_resident_profile_photo(request):
    if request.method != "POST":
        return redirect("tenant_dashboard")

    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    form = ResidentProfilePhotoForm(request.POST, request.FILES, instance=application)

    if form.is_valid():
        old_photo_name = application.profile_photo.name if application.profile_photo else ""
        new_photo = request.FILES.get("profile_photo")
        form.save()

        if old_photo_name and new_photo and old_photo_name != application.profile_photo.name:
            application.profile_photo.storage.delete(old_photo_name)

        messages.success(request, "Profile photo updated.")
    else:
        messages.error(request, "Please choose a valid image file.")

    return redirect("tenant_dashboard")


@login_required
def upload_resident_document(request):
    if request.method != "POST":
        return redirect("tenant_dashboard")

    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    document_type = request.POST.get("document_type", "other")
    resident_upload_types = {"id", "income", "bank", "other"}
    name = request.POST.get("name", "").strip()
    uploaded_file = request.FILES.get("file")

    if document_type not in resident_upload_types:
        messages.error(request, "Choose a valid resident upload document type.")
        return redirect("tenant_dashboard")

    if not name or not uploaded_file:
        messages.error(request, "Document name and file are required.")
        return redirect("tenant_dashboard")

    ApplicantDocument.objects.create(
        application=application,
        document_type=document_type,
        name=name,
        file=uploaded_file,
        status="uploaded",
        locked=False,
    )

    owner_email = "BowlingLegacyLLC@outlook.com"
    if application.property and application.property.owner_email:
        owner_email = application.property.owner_email

    notification = EmailMessage(
        subject=f"New Resident Document Uploaded: {name}",
        body=f"""
A resident uploaded a new document.

Property: {application.property.name if application.property else "No Property"}
Resident: {application.full_name}
Resident Email: {application.email}
Room/Space: {application.space_type} {application.space_label}

Document:
{name}

Type:
{document_type}

Replying to this email will go to the resident's email address. Use the portal file tools for official document review.
""",
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=[owner_email],
        reply_to=[application.email] if application.email else None,
    )
    notification.send(fail_silently=True)

    messages.success(request, "Your document has been uploaded and filed.")
    return redirect("tenant_dashboard")


@login_required
def submit_resident_message(request):
    if request.method != "POST":
        return redirect("tenant_dashboard")

    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    message_type = request.POST.get("message_type", "general")
    subject = request.POST.get("subject", "").strip()
    message = request.POST.get("message", "").strip()

    if not subject or not message:
        messages.error(request, "Subject and message are required.")
        return redirect("tenant_dashboard")

    resident_message = ResidentMessage.objects.create(
        application=application,
        message_type=message_type,
        subject=subject,
        message=message,
        status="submitted",
        locked=True,
    )

    owner_email = "BowlingLegacyLLC@outlook.com"
    if application.property and application.property.owner_email:
        owner_email = application.property.owner_email

    message_url = request.build_absolute_uri(reverse("landlord_message_detail", args=[resident_message.id]))

    notification = EmailMessage(
        subject=f"New Resident Request Filed: {subject}",
        body=f"""
A new resident request/message has been filed.

Property: {application.property.name if application.property else "No Property"}
Resident: {application.full_name}
Resident Email: {application.email}
Unit/Space: {application.space_type} {application.space_label}

Type: {message_type}

Subject:
{subject}

Message:
{message}

Open secure portal thread:
{message_url}

Replying to this email will go to the resident's email address. To keep the conversation in the portal file, open the secure portal thread above.
""",
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=[owner_email],
        reply_to=[application.email] if application.email else None,
    )
    notification.send(fail_silently=True)

    messages.success(request, "Your message/request has been submitted and filed.")
    return redirect("tenant_dashboard")


@login_required
@user_passes_test(staff_required)
def payment_log(request):
    raw_month = (request.GET.get("month") or "").strip()
    month_filter_active = bool(raw_month)
    selected_month = selected_report_month(request) if month_filter_active else None
    previous_month = add_months(selected_month, -1) if selected_month else None
    next_month = add_months(selected_month, 1) if selected_month else None
    completed_payments = (
        Payment.objects
        .filter(application__in=staff_managed_applications(request.user), status="completed")
        .select_related("application", "application__property")
        .order_by("application__property__name", "-created_at", "application__space_label", "application__full_name")
    )

    grouped = OrderedDict()

    for payment in completed_payments:
        application = payment.application
        payment.display_unit_label = canonical_room_label(application.space_label or application.space_type)
        payment.display_paid_at = payment.received_at or payment.created_at
        property_name = application.property.name if application.property else "No Property"
        accounting_month = payment_service_month(payment)
        if selected_month and accounting_month != selected_month:
            continue
        month_label = accounting_month.strftime("%B %Y")

        grouped.setdefault(property_name, OrderedDict())
        grouped[property_name].setdefault(month_label, {
            "month_date": accounting_month,
            "payments": [],
        })
        grouped[property_name][month_label]["payments"].append(payment)

    payment_log_data = []

    for property_name, months in grouped.items():
        month_data = []

        for month_label, month_group in sorted(months.items(), key=lambda item: item[1]["month_date"]):
            payments = month_group["payments"]
            payments_sorted = sorted(
                payments,
                key=lambda p: resident_sort_key(p.application),
            )

            month_data.append({
                "month_label": month_label,
                "payments": payments_sorted,
            })

        payment_log_data.append({
            "property_name": property_name,
            "months": month_data,
        })

    return render(request, "payment_log.html", {
        "payment_log": payment_log_data,
        "month_filter_active": month_filter_active,
        "selected_month": selected_month,
        "previous_month": previous_month,
        "next_month": next_month,
    })


@login_required
@user_passes_test(staff_required)
def record_manual_payment(request, property_id=None):
    accessible_properties = staff_managed_properties(request.user).order_by("name")
    selected_property = None
    application_id = request.GET.get("application")

    if application_id:
        selected_application = get_object_or_404(
            staff_managed_applications(request.user).select_related("property"),
            id=application_id,
        )
        selected_property = selected_application.property
    elif property_id:
        selected_property = get_object_or_404(accessible_properties, id=property_id)
    elif accessible_properties.count() == 1:
        selected_property = accessible_properties.first()

    application_queryset = (
        staff_managed_applications(request.user)
        .select_related("property")
        .order_by("space_label", "full_name")
    )
    if selected_property:
        application_queryset = application_queryset.filter(property=selected_property)

    if request.method == "POST":
        form = ManualPaymentForm(request.POST)
        form.fields["application"].queryset = application_queryset

        if form.is_valid():
            payment = form.save(commit=False)
            payment.status = "completed"
            payment.recorded_by = request.user

            if not payment.received_at:
                payment.received_at = timezone.now()

            if not payment.description:
                payment.description = f"Manual {payment.get_payment_method_display()} payment"

            payment.save()
            apply_completed_payment_to_balance(payment)

            messages.success(request, "Manual payment recorded and resident balance updated.")
            return redirect("payment_receipt", payment_id=payment.id)
    else:
        initial = {}

        if application_id:
            initial["application"] = application_id

        form = ManualPaymentForm(initial=initial)
        form.fields["application"].queryset = application_queryset

    return render(request, "record_manual_payment.html", {
        "form": form,
        "properties": accessible_properties,
        "selected_property": selected_property,
        "show_property_picker": not selected_property and not application_id,
    })


@login_required
@user_passes_test(staff_required)
def edit_manual_payment(request, payment_id):
    payment = get_object_or_404(
        Payment.objects.select_related("application", "application__property"),
        id=payment_id,
        application__in=staff_managed_applications(request.user),
    )

    if request.method == "POST":
        form = ManualPaymentForm(request.POST, instance=payment)
        form.fields["application"].queryset = (
            staff_managed_applications(request.user)
            .select_related("property")
            .order_by("property__name", "space_label", "full_name")
        )

        if form.is_valid():
            payment = form.save()
            if payment.status == "completed":
                recalculate_application_balances(payment.application)
            messages.success(request, "Payment record corrected and resident balances recalculated.")
            return redirect("payment_receipt", payment_id=payment.id)
    else:
        form = ManualPaymentForm(instance=payment)
        form.fields["application"].queryset = (
            staff_managed_applications(request.user)
            .select_related("property")
            .order_by("property__name", "space_label", "full_name")
        )

    return render(request, "record_manual_payment.html", {
        "form": form,
        "payment": payment,
        "is_edit": True,
    })


@login_required
@user_passes_test(staff_required)
def edit_resident_balances(request, application_id):
    application = get_object_or_404(
        HousingApplication.objects.select_related("property", "user"),
        id=application_id,
        id__in=staff_managed_applications(request.user).filter(user__isnull=False).values_list("id", flat=True),
    )

    form = ResidentBalanceCorrectionForm(request.POST or None, instance=application)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Resident balances updated.")
        return redirect("landlord_resident_files")

    return render(request, "edit_resident_balances.html", {
        "application": application,
        "form": form,
    })


@login_required
@user_passes_test(staff_required)
def payment_receipt(request, payment_id):
    payment = get_object_or_404(
        Payment.objects.select_related("application", "application__property", "recorded_by"),
        id=payment_id,
        application__in=staff_managed_applications(request.user),
    )

    return render(request, "payment_receipt.html", {"payment": payment})


@login_required
@user_passes_test(staff_required)
def rent_roll(request):
    selected_month = selected_report_month(request)
    previous_month = add_months(selected_month, -1)
    next_month = add_months(selected_month, 1)
    properties, selected_property, show_property_picker = selected_property_scope(request)
    rows = [] if show_property_picker else rent_roll_rows_for_properties(request.user, selected_month, properties)
    totals = rent_roll_totals(rows)

    return render(request, "rent_roll.html", {
        "rows": rows,
        "totals": totals,
        "properties": properties,
        "selected_property": selected_property,
        "show_property_picker": show_property_picker,
        "selected_month": selected_month,
        "previous_month": previous_month,
        "next_month": next_month,
    })


@login_required
@user_passes_test(reporting_required)
def custom_reports(request):
    properties = custom_report_accessible_properties(request.user)
    form = CustomReportForm(
        request.GET or None,
        properties=properties,
        initial={
            "report_type": "resident_phone_list",
            "financial_entry_types": ["operating_expense"],
        },
    )

    selected_property = None
    report_title = ""
    report_columns = []
    report_rows = []
    totals = {}
    generated = bool(request.GET)

    if form.is_valid() and generated:
        report_type = form.cleaned_data["report_type"]
        selected_property_id = form.cleaned_data.get("property_id")
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")

        filtered_properties = properties
        if selected_property_id:
            selected_property = get_object_or_404(properties, id=selected_property_id)
            filtered_properties = properties.filter(id=selected_property.id)

        residents = (
            HousingApplication.objects
            .select_related("property")
            .filter(property__in=filtered_properties)
            .order_by("property__name", "space_label", "full_name")
        )
        sorted_residents = sorted_resident_list(residents)

        if report_type == "resident_phone_list":
            report_title = "Resident Phone List"
            report_columns = ["Property", "Room / Unit", "Resident", "Phone", "Email"]
            report_rows = [
                [
                    resident.property.name if resident.property else "No Property",
                    resident.space_label or resident.space_type or "",
                    resident.full_name,
                    phone_format(resident.phone),
                    resident.email,
                ]
                for resident in sorted_residents
            ]

        elif report_type == "resident_roster":
            report_title = "Resident Roster"
            report_columns = ["Property", "Room / Unit", "Resident", "Phone", "Monthly Rent", "Balance", "Deposit Due"]
            for resident in sorted_residents:
                deposit_due = max(resident.deposit_required - resident.deposit_paid, Decimal("0.00"))
                report_rows.append([
                    resident.property.name if resident.property else "No Property",
                    resident.space_label or resident.space_type or "",
                    resident.full_name,
                    phone_format(resident.phone),
                    resident.monthly_rent,
                    resident.balance,
                    deposit_due,
                ])
            totals = {
                "Scheduled Rent": sum((resident.monthly_rent for resident in residents), Decimal("0.00")),
                "Open Balances": sum((resident.balance for resident in residents), Decimal("0.00")),
            }

        elif report_type == "payment_summary":
            report_title = "Payment Summary"
            report_columns = ["Date Received", "Applies To", "Months Covered", "Property", "Resident", "Type", "Method", "Amount", "Status"]
            payments = (
                Payment.objects
                .select_related("application", "application__property")
                .filter(application__in=residents)
                .order_by("-created_at")
            )
            if start_date:
                payments = payments.filter(created_at__date__gte=start_date)
            if end_date:
                payments = payments.filter(created_at__date__lte=end_date)

            for payment in payments:
                report_rows.append([
                    timezone.localtime(payment.created_at).strftime("%Y-%m-%d"),
                    payment.accounting_month_label,
                    payment.months_covered,
                    payment.application.property.name if payment.application.property else "No Property",
                    payment.application.full_name,
                    payment.get_payment_type_display(),
                    payment.get_payment_method_display(),
                    payment.amount,
                    payment.get_status_display(),
                ])
            totals = {
                "Completed Total": payments.filter(status="completed").aggregate(total=Sum("amount"))["total"] or Decimal("0.00"),
                "All Matching Payments": payments.aggregate(total=Sum("amount"))["total"] or Decimal("0.00"),
            }

        elif report_type == "delinquency_report":
            report_title = "Delinquency Report"
            report_columns = ["Property", "Room / Unit", "Resident", "Rent Balance", "Utility Balance", "Deposit Due", "Total Due"]
            total_due_sum = Decimal("0.00")
            for resident in sorted_residents:
                deposit_due = max(resident.deposit_required - resident.deposit_paid, Decimal("0.00"))
                total_due = resident.balance + resident.utility_balance + deposit_due
                if total_due <= 0:
                    continue
                total_due_sum += total_due
                report_rows.append([
                    resident.property.name if resident.property else "No Property",
                    resident.space_label or resident.space_type or "",
                    resident.full_name,
                    resident.balance,
                    resident.utility_balance,
                    deposit_due,
                    total_due,
                ])
            totals = {"Total Due": total_due_sum}

        elif report_type == "deposit_liability":
            report_title = "Deposit Liability Report"
            report_columns = ["Property", "Room / Unit", "Resident", "Deposit Required", "Deposit Held", "Deposit Balance"]
            deposit_required_total = Decimal("0.00")
            deposit_held_total = Decimal("0.00")
            deposit_balance_total = Decimal("0.00")
            for resident in sorted_residents:
                deposit_balance = max(resident.deposit_required - resident.deposit_paid, Decimal("0.00"))
                deposit_required_total += resident.deposit_required
                deposit_held_total += resident.deposit_paid
                deposit_balance_total += deposit_balance
                report_rows.append([
                    resident.property.name if resident.property else "No Property",
                    resident.space_label or resident.space_type or "",
                    resident.full_name,
                    resident.deposit_required,
                    resident.deposit_paid,
                    deposit_balance,
                ])
            totals = {
                "Deposit Required": deposit_required_total,
                "Deposit Held": deposit_held_total,
                "Deposit Balance": deposit_balance_total,
            }

        elif report_type == "property_performance_summary":
            report_title = "Property Performance Summary"
            report_columns = ["Property", "Year", "Units", "Occupied", "Occupancy %", "Income", "Operating Expenses", "NOI", "Debt Service", "Cash Flow"]
            report_year = custom_report_period_year(start_date)
            for property_obj in filtered_properties:
                property_qs = Property.objects.filter(id=property_obj.id)
                months, property_totals = t12_report_rows(request.user, report_year, property_qs)
                units = PropertyRoomRent.objects.filter(property=property_obj, is_active=True).count()
                occupied = (
                    HousingApplication.objects
                    .filter(property=property_obj)
                    .filter(Q(user__isnull=False) | Q(payments__status="completed"))
                    .exclude(space_label="")
                    .values("space_label")
                    .distinct()
                    .count()
                )
                report_rows.append([
                    property_obj.name,
                    report_year,
                    units,
                    occupied,
                    f"{decimal_percent(occupied, units)}%",
                    property_totals["total_income"],
                    property_totals["operating_expenses"],
                    property_totals["net_operating_income"],
                    property_totals["debt_service"],
                    property_totals["cash_flow_after_debt"],
                ])

        elif report_type == "valuation_estimate":
            report_title = "Valuation Estimate Report"
            report_columns = ["Property", "NOI Used", "Cap Rate", "Estimated Value", "Method"]
            report_year = custom_report_period_year(start_date)
            cap_rates = [Decimal("0.060"), Decimal("0.065"), Decimal("0.070"), Decimal("0.075"), Decimal("0.080"), Decimal("0.085")]
            for property_obj in filtered_properties:
                property_qs = Property.objects.filter(id=property_obj.id)
                months, property_totals = t12_report_rows(request.user, report_year, property_qs)
                active_months = sum(1 for month in months if month["total_income"] > 0 or month["operating_expenses"] > 0)
                noi_used = property_totals["net_operating_income"]
                method = f"{report_year} actual"
                if active_months and active_months < 12:
                    noi_used = (noi_used / Decimal(active_months) * Decimal("12.00")).quantize(Decimal("0.01"))
                    method = f"{active_months} months annualized"
                for cap_rate in cap_rates:
                    estimated_value = (noi_used / cap_rate).quantize(Decimal("0.01")) if cap_rate else Decimal("0.00")
                    report_rows.append([
                        property_obj.name,
                        noi_used,
                        f"{(cap_rate * Decimal('100')).quantize(Decimal('0.1'))}%",
                        estimated_value,
                        method,
                    ])

        elif report_type == "income_statement":
            report_title = "Income Statement / P&L"
            report_columns = ["Property", "Month", "Type", "Category", "Amount"]
            entries = custom_report_financial_entries(filtered_properties, ["income", "operating_expense", "debt_service", "capital_expense"], start_date, end_date)
            for entry in entries:
                report_rows.append([
                    entry.property_name,
                    entry.entry_date.strftime("%B %Y") if entry.entry_date else f"{entry.year or ''}-{entry.month or ''}",
                    entry.get_entry_type_display(),
                    entry.category,
                    entry.amount,
                ])
            totals = {"Report Total": entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")}

        elif report_type == "expense_by_category":
            report_title = "Expense Detail by Category"
            report_columns = ["Property", "Type", "Category", "Amount"]
            entries = custom_report_financial_entries(filtered_properties, ["operating_expense", "capital_expense", "debt_service"], start_date, end_date)
            grouped_entries = (
                entries.values("property_name", "entry_type", "category")
                .annotate(total=Sum("amount"))
                .order_by("property_name", "entry_type", "category")
            )
            for entry in grouped_entries:
                report_rows.append([
                    entry["property_name"],
                    dict(FinancialEntry.ENTRY_TYPE_CHOICES).get(entry["entry_type"], entry["entry_type"]),
                    entry["category"],
                    entry["total"],
                ])
            totals = {"Expense Total": entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")}

        elif report_type == "vendor_expense":
            report_title = "Vendor Expense Report"
            report_columns = ["Property", "Vendor", "Category", "Amount", "Status"]
            receipts = AccountingReceipt.objects.filter(property__in=filtered_properties)
            if start_date:
                receipts = receipts.filter(receipt_date__gte=start_date)
            if end_date:
                receipts = receipts.filter(receipt_date__lte=end_date)
            grouped_receipts = (
                receipts.values("property__name", "vendor", "category__name", "status")
                .annotate(total=Sum("amount"))
                .order_by("property__name", "vendor", "category__name")
            )
            for receipt in grouped_receipts:
                report_rows.append([
                    receipt["property__name"],
                    receipt["vendor"] or "Unassigned Vendor",
                    receipt["category__name"] or "Unassigned Category",
                    receipt["total"],
                    receipt["status"],
                ])
            totals = {"Vendor Expense Total": receipts.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")}

        elif report_type == "occupancy_vacancy":
            report_title = "Occupancy / Vacancy Report"
            report_columns = ["Property", "Units", "Occupied", "Vacant", "Occupancy %", "Vacant Units"]
            total_units = 0
            total_occupied = 0
            for property_obj in filtered_properties:
                room_settings = PropertyRoomRent.objects.filter(property=property_obj, is_active=True)
                unit_labels = {normalized_room_label(room.room_unit_label) for room in room_settings}
                occupied_labels = {
                    normalized_room_label(application.space_label)
                    for application in HousingApplication.objects.filter(property=property_obj)
                    .filter(Q(user__isnull=False) | Q(payments__status="completed"))
                    .exclude(space_label="")
                }
                vacant_labels = sorted(label.upper() for label in unit_labels if label and label not in occupied_labels)
                units = len(unit_labels)
                occupied = len([label for label in unit_labels if label in occupied_labels])
                total_units += units
                total_occupied += occupied
                report_rows.append([
                    property_obj.name,
                    units,
                    occupied,
                    max(units - occupied, 0),
                    f"{decimal_percent(occupied, units)}%",
                    ", ".join(vacant_labels) or "-",
                ])
            totals = {
                "Units": Decimal(total_units),
                "Occupied": Decimal(total_occupied),
                "Vacant": Decimal(max(total_units - total_occupied, 0)),
            }

        elif report_type == "capital_improvement_log":
            report_title = "Capital Improvement Log"
            report_columns = ["Date", "Property", "Category", "Description", "Amount", "Source"]
            entries = custom_report_financial_entries(filtered_properties, ["capital_expense"], start_date, end_date)
            for entry in entries:
                report_rows.append([
                    entry.entry_date or f"{entry.year or ''}-{entry.month or ''}",
                    entry.property_name,
                    entry.category,
                    entry.description,
                    entry.amount,
                    entry.upload.name,
                ])
            totals = {"Capital Improvements": entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")}

        elif report_type == "utility_cost_trend":
            report_title = "Utility Usage / Cost Trend"
            report_columns = ["Property", "Month", "Category", "Amount"]
            entries = custom_report_financial_entries(filtered_properties, ["operating_expense"], start_date, end_date)
            entries = entries.filter(
                Q(category__icontains="power") |
                Q(category__icontains="electric") |
                Q(category__icontains="gas") |
                Q(category__icontains="water") |
                Q(category__icontains="trash") |
                Q(category__icontains="internet") |
                Q(category__icontains="utility") |
                Q(category__icontains="utilities") |
                Q(category__icontains="City of Medford")
            )
            for entry in entries:
                report_rows.append([
                    entry.property_name,
                    entry.entry_date.strftime("%B %Y") if entry.entry_date else f"{entry.year or ''}-{entry.month or ''}",
                    entry.category,
                    entry.amount,
                ])
            totals = {"Utility Cost Total": entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")}

        elif report_type == "insurance_compliance":
            report_title = "Insurance / Compliance Report"
            report_columns = ["Property", "Renters Insurance Provider", "Renters Insurance Link", "Insurance Expense Total", "Notes"]
            for property_obj in filtered_properties:
                insurance_entries = custom_report_financial_entries(Property.objects.filter(id=property_obj.id), ["operating_expense"], start_date, end_date).filter(category__icontains="insurance")
                report_rows.append([
                    property_obj.name,
                    property_obj.renters_insurance_provider_name or "Not set",
                    property_obj.renters_insurance_url or "Not set",
                    insurance_entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00"),
                    property_obj.renters_insurance_notes or "",
                ])

        elif report_type == "financial_entries":
            selected_entry_types = form.cleaned_data.get("financial_entry_types") or ["operating_expense"]
            report_title = "Financial Entries / Expenses"
            report_columns = ["Date", "Ledger", "Property", "Type", "Category", "Description", "Amount", "Source"]
            entries = custom_report_financial_entries(filtered_properties, selected_entry_types, start_date, end_date)

            for entry in entries:
                report_rows.append([
                    entry.entry_date or f"{entry.year or ''}-{entry.month or ''}",
                    entry.get_ledger_scope_display(),
                    entry.property_name,
                    entry.get_entry_type_display(),
                    entry.category,
                    entry.description,
                    entry.amount,
                    entry.upload.name,
                ])
            totals = {
                "Report Total": entries.aggregate(total=Sum("amount"))["total"] or Decimal("0.00"),
            }

    return render(request, "custom_reports.html", {
        "form": form,
        "generated": generated,
        "selected_property": selected_property,
        "report_title": report_title,
        "report_columns": report_columns,
        "report_rows": report_rows,
        "totals": totals,
        "row_count": len(report_rows),
        "generated_at": timezone.localtime(),
    })


@login_required
@user_passes_test(staff_required)
def export_payment_log_csv(request):
    raw_month = (request.GET.get("month") or "").strip()
    selected_month = selected_report_month(request) if raw_month else None
    response = HttpResponse(content_type="text/csv")
    month_suffix = f"_{selected_month.strftime('%Y_%m')}" if selected_month else ""
    response["Content-Disposition"] = f'attachment; filename="payment_log{month_suffix}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Resident", "Property", "Payment Type", "Amount", "Status", "Date Received", "Applies To", "Months Covered"])

    payments = (
        Payment.objects
        .filter(application__in=staff_managed_applications(request.user), status="completed")
        .select_related("application", "application__property")
        .order_by("application__property__name", "application__space_label", "application__full_name", "-created_at")
    )

    sorted_payments = sorted(
        payments,
        key=lambda payment: (
            resident_sort_key(payment.application),
            -(payment.received_at or payment.created_at).timestamp(),
            payment.id,
        ),
    )

    for payment in sorted_payments:
        accounting_month = payment_service_month(payment)
        if selected_month and accounting_month != selected_month:
            continue
        display_paid_at = payment.received_at or payment.created_at
        writer.writerow([
            payment.application.full_name,
            payment.application.property.name if payment.application.property else "",
            payment.get_payment_type_display(),
            payment.amount,
            payment.status,
            timezone.localtime(display_paid_at).strftime("%Y-%m-%d %H:%M"),
            payment.accounting_month_label,
            payment.months_covered,
        ])

    return response


@login_required
@user_passes_test(staff_required)
def export_rent_roll_csv(request):
    selected_month = selected_report_month(request)
    properties, selected_property, _show_property_picker = selected_property_scope(request)
    response = HttpResponse(content_type="text/csv")
    property_suffix = f'_{selected_property.name.replace(" ", "_")}' if selected_property else ""
    response["Content-Disposition"] = f'attachment; filename="rent_roll{property_suffix}_{selected_month.strftime("%Y_%m")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Month",
        "Resident",
        "Unit",
        "Rent",
        "Rent Paid",
        "Rent Balance",
        "Utilities",
        "Utilities Paid",
        "Utilities Balance",
        "Deposit",
        "Deposit Paid",
        "Deposit Balance",
    ])

    rows = rent_roll_rows_for_properties(request.user, selected_month, properties)
    totals = rent_roll_totals(rows)
    for row in rows:
        writer.writerow([
            selected_month.strftime("%B %Y"),
            row["resident"],
            row["room"],
            row["monthly_rent"],
            row["rent_paid"],
            row["rent_balance"],
            row["utility_monthly"],
            row["utility_paid"],
            row["utility_balance"],
            row["deposit_required"],
            row["deposit_paid"],
            row["deposit_balance"],
        ])
    writer.writerow([
        selected_month.strftime("%B %Y"),
        "TOTAL",
        "",
        totals["monthly_rent"],
        totals["rent_paid"],
        totals["rent_balance"],
        totals["utility_monthly"],
        totals["utility_paid"],
        totals["utility_balance"],
        totals["deposit_required"],
        totals["deposit_paid"],
        totals["deposit_balance"],
    ])

    return response


@login_required
@user_passes_test(reporting_required)
def export_t12_csv(request):
    year = selected_report_year(request)
    report_properties, selected_property = selected_report_properties(request)
    months, totals = t12_report_rows(request.user, year, report_properties)
    property_suffix = f'_{selected_property.name.replace(" ", "_")}' if selected_property else ""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="t12_report{property_suffix}_{year}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Month",
        "Income Source",
        "Online Income",
        "Spreadsheet Income",
        "Total Income",
        "Operating Expenses",
        "Debt Service",
        "Capital Expenses",
        "NOI",
        "Cash Flow After Debt",
    ])

    for month in months:
        writer.writerow([
            month["month_name"],
            month["income_source"],
            csv_money(month["online_income"]),
            csv_money(month["spreadsheet_income"]),
            csv_money(month["total_income"]),
            csv_money(month["operating_expenses"]),
            csv_money(month["debt_service"]),
            csv_money(month["capital_expenses"]),
            csv_money(month["net_operating_income"]),
            csv_money(month["cash_flow_after_debt"]),
        ])

    writer.writerow([
        "Total",
        "",
        csv_money(totals["online_income"]),
        csv_money(totals["spreadsheet_income"]),
        csv_money(totals["total_income"]),
        csv_money(totals["operating_expenses"]),
        csv_money(totals["debt_service"]),
        csv_money(totals["capital_expenses"]),
        csv_money(totals["net_operating_income"]),
        csv_money(totals["cash_flow_after_debt"]),
    ])

    return response


@login_required
@user_passes_test(reporting_required)
def t12_report(request):
    year = selected_report_year(request)
    report_properties, selected_property = selected_report_properties(request)
    months, totals = t12_report_rows(request.user, year, report_properties)

    return render(request, "t12_report.html", {
        "year": year,
        "months": months,
        "totals": totals,
        "properties": staff_managed_properties(request.user).order_by("name"),
        "selected_property": selected_property,
    })


def selected_report_year(request):
    raw_year = (request.GET.get("year") or "").strip()
    if raw_year:
        try:
            year = int(raw_year)
            if 2000 <= year <= 2100:
                return year
        except ValueError:
            pass
    return timezone.localdate().year


def selected_report_properties(request):
    accessible_properties = staff_managed_properties(request.user).order_by("name")
    selected_property = None
    property_id = (request.GET.get("property_id") or "").strip()

    if property_id:
        selected_property = get_object_or_404(accessible_properties, id=property_id)
        return accessible_properties.filter(id=selected_property.id), selected_property

    if accessible_properties.count() == 1:
        selected_property = accessible_properties.first()
        return accessible_properties.filter(id=selected_property.id), selected_property

    return accessible_properties, selected_property


def t12_report_rows(user, year, report_properties=None):
    accessible_properties = report_properties or staff_managed_properties(user)
    accessible_property_names = list(accessible_properties.values_list("name", flat=True))
    financial_entries = FinancialEntry.objects.filter(
        Q(upload__property__in=accessible_properties) | Q(property_name__in=accessible_property_names),
    ).filter(
        Q(year=year) | Q(entry_date__year=year),
    )
    completed_payments = list(
        Payment.objects
        .filter(application__in=staff_managed_applications(user), status="completed")
        .filter(application__property__in=accessible_properties)
    )
    months = []
    totals = {
        "online_income": Decimal("0.00"),
        "spreadsheet_income": Decimal("0.00"),
        "total_income": Decimal("0.00"),
        "operating_expenses": Decimal("0.00"),
        "debt_service": Decimal("0.00"),
        "capital_expenses": Decimal("0.00"),
        "net_operating_income": Decimal("0.00"),
        "cash_flow_after_debt": Decimal("0.00"),
    }

    for month_number in range(1, 13):
        portal_income = payment_amount_for_month(completed_payments, year, month_number, T12_INCOME_PAYMENT_TYPES)
        month_filter = Q(month=month_number) | Q(entry_date__month=month_number)
        month_entries = financial_entries.filter(month_filter)
        summary_entries = month_entries.filter(source_receipt__isnull=True)
        receipt_entries = month_entries.filter(source_receipt__isnull=False)
        baseline_time = latest_summary_baseline_time(financial_entries, month_filter)
        if baseline_time:
            receipt_entries = receipt_entries.filter(created_at__gt=baseline_time)

        spreadsheet_income = entries_total(
            summary_entries
            .filter(entry_type="income")
            .exclude(category__icontains="deposit")
        )
        receipt_income = entries_total(
            receipt_entries
            .filter(entry_type="income")
            .exclude(category__icontains="deposit")
        )
        operating_expenses = (
            entries_total(summary_entries.filter(entry_type="operating_expense"))
            + entries_total(receipt_entries.filter(entry_type="operating_expense"))
        )
        debt_service = (
            entries_total(summary_entries.filter(entry_type="debt_service"))
            + entries_total(receipt_entries.filter(entry_type="debt_service"))
        )
        capital_expenses = (
            entries_total(summary_entries.filter(entry_type="capital_expense"))
            + entries_total(receipt_entries.filter(entry_type="capital_expense"))
        )

        if spreadsheet_income > 0:
            online_income = Decimal("0.00")
            total_income = spreadsheet_income + receipt_income
            income_source = "Spreadsheet + Receipts" if receipt_income > 0 else "Spreadsheet"
        else:
            online_income = portal_income
            total_income = portal_income + receipt_income
            income_source = "Portal + Receipts" if receipt_income > 0 else "Portal"

        net_operating_income = total_income - operating_expenses
        cash_flow_after_debt = net_operating_income - debt_service

        row = {
            "month_name": date(year, month_number, 1).strftime("%B"),
            "income_source": income_source if total_income > 0 else "",
            "online_income": online_income,
            "spreadsheet_income": spreadsheet_income,
            "total_income": total_income,
            "operating_expenses": operating_expenses,
            "debt_service": debt_service,
            "capital_expenses": capital_expenses,
            "net_operating_income": net_operating_income,
            "cash_flow_after_debt": cash_flow_after_debt,
        }
        months.append(row)

        for key in totals:
            totals[key] += row[key]

    return months, totals


@login_required
@user_passes_test(reporting_required)
def financial_upload(request):
    properties = staff_managed_properties(request.user).order_by("name")

    if request.method == "POST":
        form = FinancialUploadForm(request.POST, request.FILES, properties=properties)

        if form.is_valid():
            upload = form.save()
            return redirect("parse_financial_upload", upload_id=upload.id)
    else:
        form = FinancialUploadForm(properties=properties)

    uploads = (
        FinancialUpload.objects
        .filter(property__in=properties)
        .select_related("property")
        .order_by("-uploaded_at")
    )
    return render(request, "financial_upload.html", {"form": form, "uploads": uploads})


@login_required
@user_passes_test(reporting_required)
def accounting_receipts(request):
    properties = staff_managed_properties(request.user).order_by("name")

    if request.method == "POST":
        form = AccountingReceiptForm(
            request.POST,
            request.FILES,
            properties=properties,
            user=request.user,
        )

        if form.is_valid():
            receipt = form.save()
            process_receipt_ocr(receipt)
            if receipt.ocr_status == "extracted":
                messages.success(request, "Receipt uploaded, text was extracted, and available fields were prefilled for review.")
            elif receipt.ocr_status == "needs_ocr_provider":
                messages.warning(request, "Receipt uploaded. This looks like a scanned image or image-only PDF, so it is stored for review until OCR is connected.")
            else:
                messages.success(request, "Receipt uploaded and saved for accounting review.")
            return redirect("accounting_receipts")
    else:
        form = AccountingReceiptForm(properties=properties, user=request.user)

    receipts = (
        AccountingReceipt.objects
        .select_related("property", "category", "uploaded_by", "financial_entry")
        .filter(property__in=properties)
        .order_by("status", "-uploaded_at")
    )

    return render(request, "accounting_receipts.html", {
        "form": form,
        "receipts": receipts,
    })


@login_required
@user_passes_test(reporting_required)
def approve_accounting_receipt(request, receipt_id):
    if request.method != "POST":
        return redirect("accounting_receipts")

    receipt = get_object_or_404(
        AccountingReceipt.objects.select_related("property", "category", "financial_entry"),
        id=receipt_id,
        property__in=staff_managed_properties(request.user),
    )

    if not receipt.category:
        messages.error(request, "Choose or create a category before approving this receipt.")
        return redirect("accounting_receipts")

    if receipt.amount <= 0:
        messages.error(request, "Enter a valid amount before approving this receipt.")
        return redirect("accounting_receipts")

    if not receipt.financial_upload:
        receipt.financial_upload = FinancialUpload.objects.create(
            property=receipt.property,
            file=receipt.receipt_file,
            name=f"Receipt - {receipt.vendor or receipt.property.name}",
            notes=receipt.notes,
            parsed_at=timezone.now(),
        )

    if not receipt.financial_entry and duplicate_receipt_financial_entry_exists(receipt):
        receipt.status = "ignored"
        receipt.reviewed_by = request.user
        receipt.reviewed_at = timezone.now()
        receipt.save(update_fields=["status", "reviewed_by", "reviewed_at"])
        messages.warning(request, "Receipt matched an existing ledger entry and was marked as duplicate instead of being counted again.")
        return redirect("accounting_receipts")

    if not receipt.financial_entry:
        receipt.financial_entry = FinancialEntry.objects.create(
            upload=receipt.financial_upload,
            ledger_scope=receipt.financial_upload.ledger_scope,
            property_name=receipt.property.name,
            sheet_name="Receipt Upload",
            row_number=receipt.id,
            entry_date=receipt.receipt_date,
            month=receipt.receipt_date.month if receipt.receipt_date else None,
            year=receipt.receipt_date.year if receipt.receipt_date else None,
            entry_type=receipt.entry_type,
            category=receipt.category.name,
            description=receipt.description or receipt.vendor,
            amount=receipt.amount,
        )

    receipt.status = "approved"
    receipt.reviewed_by = request.user
    receipt.reviewed_at = timezone.now()
    receipt.save()

    messages.success(request, "Receipt approved and added to the financial ledger.")
    return redirect("accounting_receipts")


@login_required
@user_passes_test(reporting_required)
def process_accounting_receipt_ocr(request, receipt_id):
    if request.method != "POST":
        return redirect("accounting_receipts")

    receipt = get_object_or_404(
        AccountingReceipt.objects.select_related("property"),
        id=receipt_id,
        property__in=staff_managed_properties(request.user),
    )
    process_receipt_ocr(receipt)

    if receipt.ocr_status == "extracted":
        messages.success(request, "OCR text was extracted and receipt suggestions were updated.")
    elif receipt.ocr_status == "needs_ocr_provider":
        messages.warning(request, "This file is stored, but scanned images and image-only PDFs need an OCR provider before text can be read automatically.")
    else:
        messages.error(request, "OCR could not process this receipt.")

    return redirect("accounting_receipts")


@login_required
@user_passes_test(reporting_required)
def parse_financial_upload(request, upload_id):
    properties = staff_managed_properties(request.user).order_by("name")
    upload = get_object_or_404(
        FinancialUpload.objects.select_related("property"),
        id=upload_id,
        property__in=properties,
    )

    sheet_names = financial_upload_sheet_names(upload)
    selected_sheet_name = request.POST.get("sheet_name") or request.GET.get("sheet_name") or (sheet_names[0] if sheet_names else None)

    try:
        sheet_name, headers, rows = read_financial_upload_rows(upload, selected_sheet_name=selected_sheet_name)
    except Exception as exc:
        messages.error(request, f"Rental Ledger Pro could not read that file yet: {exc}")
        return redirect("financial_upload")

    guesses = guess_financial_columns(headers)
    entry_type_choices = FinancialEntry.ENTRY_TYPE_CHOICES
    created = 0
    skipped = 0

    if request.method == "POST":
        import_mode = request.POST.get("import_mode", "detail")
        date_column = request.POST.get("date_column", "")
        description_column = request.POST.get("description_column", "")
        amount_column = request.POST.get("amount_column", "")
        utility_amount_column = request.POST.get("utility_amount_column", "")
        deposit_amount_column = request.POST.get("deposit_amount_column", "")
        other_income_amount_column = request.POST.get("other_income_amount_column", "")
        category_column = request.POST.get("category_column", "")
        entry_type_column = request.POST.get("entry_type_column", "")
        property_column = request.POST.get("property_column", "")
        default_entry_type = request.POST.get("default_entry_type", "operating_expense")
        default_category = request.POST.get("default_category", "").strip()
        summary_category_column = request.POST.get("summary_category_column", "")
        summary_year_raw = request.POST.get("summary_year", "")
        summary_entry_type = request.POST.get("summary_entry_type", "operating_expense")
        summary_month_columns = request.POST.getlist("summary_month_columns")

        accessible_properties = {property_obj.name.lower(): property_obj for property_obj in properties}
        upload.entries.all().delete()

        if import_mode == "summary_grid":
            try:
                summary_year = int(summary_year_raw)
            except (TypeError, ValueError):
                summary_year = timezone.localdate().year

            summary_categories = [
                str(row["data"].get(summary_category_column, "") or "").strip()
                for row in rows
            ]

            for row in rows:
                row_data = row["data"]
                category = str(row_data.get(summary_category_column, "") or "").strip()
                if should_skip_summary_category(category, summary_categories):
                    skipped += 1
                    continue

                created_for_row = 0
                for column_name in summary_month_columns:
                    month_number = parse_month_header(column_name)
                    if not month_number:
                        continue

                    amount = money(row_data.get(column_name))
                    if amount == Decimal("0.00"):
                        continue

                    summary_row_entry_type = summary_category_entry_type(category, summary_entry_type)
                    if not summary_row_entry_type:
                        continue

                    entry = create_financial_entry_from_import(
                        upload,
                        upload.property,
                        sheet_name,
                        row,
                        date(summary_year, month_number, 1),
                        summary_row_entry_type,
                        category,
                        f"{category} - {column_name} summary",
                        amount,
                    )
                    if entry:
                        created += 1
                        created_for_row += 1

                if created_for_row == 0:
                    skipped += 1
        else:
            for row in rows:
                row_data = row["data"]

                raw_property_name = str(row_data.get(property_column, "") or "").strip()
                property_obj = upload.property
                if raw_property_name:
                    property_obj = accessible_properties.get(raw_property_name.lower())
                    if not property_obj:
                        skipped += 1
                        continue

                description = str(row_data.get(description_column, "") or "").strip()
                category = str(row_data.get(category_column, "") or "").strip() or default_category or "Uncategorized"
                primary_amount = money(row_data.get(amount_column)) if amount_column else Decimal("0.00")
                entry_type = normalize_entry_type(
                    row_data.get(entry_type_column, ""),
                    category,
                    description,
                    primary_amount,
                    default_entry_type,
                )
                entry_date = parse_import_date(row_data.get(date_column))

                created_for_row = 0
                amount_columns = [
                    (amount_column, entry_type, category, description),
                    (utility_amount_column, "income", "Utility Payment", f"{description} - Utility payment".strip(" -")),
                    (deposit_amount_column, "income", "Deposit Payment", f"{description} - Deposit payment".strip(" -")),
                    (other_income_amount_column, "income", "Other Income", f"{description} - Other income".strip(" -")),
                ]

                for column_name, column_entry_type, column_category, column_description in amount_columns:
                    if not column_name:
                        continue

                    amount = money(row_data.get(column_name))
                    if amount == Decimal("0.00"):
                        continue

                    if column_entry_type != "income" and column_category:
                        ExpenseCategory.objects.get_or_create(
                            name=column_category,
                            defaults={
                                "entry_type": column_entry_type if column_entry_type in dict(ExpenseCategory.ENTRY_TYPE_CHOICES) else "other",
                                "created_by": request.user,
                            },
                        )

                    entry = create_financial_entry_from_import(
                        upload,
                        property_obj,
                        sheet_name,
                        row,
                        entry_date,
                        column_entry_type,
                        column_category,
                        column_description,
                        amount,
                    )
                    if entry:
                        created += 1
                        created_for_row += 1

                if created_for_row == 0:
                    skipped += 1

        upload.parsed_at = timezone.now()
        upload.save(update_fields=["parsed_at"])
        messages.success(request, f"Imported {created} ledger entries. Skipped {skipped} rows that were blank, zero, or outside your properties.")
        return redirect("financial_upload")

    preview_rows = rows[:10]

    return render(request, "financial_upload_parsed.html", {
        "upload": upload,
        "sheet_names": sheet_names,
        "selected_sheet_name": sheet_name,
        "headers": headers,
        "preview_rows": preview_rows,
        "guesses": guesses,
        "entry_type_choices": entry_type_choices,
        "summary_month_headers": summary_month_header_options(headers),
        "created": created,
        "skipped": skipped,
        "existing_entries": upload.entries.count(),
    })


@login_required
@user_passes_test(staff_required)
def property_financials(request, property_name):
    property_obj = get_object_or_404(Property, name=property_name)
    residents = HousingApplication.objects.filter(property=property_obj)

    monthly_rent = sum([r.monthly_rent for r in residents], Decimal("0.00"))
    balances_due = sum([r.balance for r in residents], Decimal("0.00"))
    utilities_due = sum([r.utility_balance for r in residents], Decimal("0.00"))
    deposits_held = sum([r.deposit_paid for r in residents], Decimal("0.00"))

    completed_payments = Payment.objects.filter(application__property=property_obj, status="completed")
    total_collected = completed_payments.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    return render(request, "property_financials.html", {
        "property": property_obj,
        "residents": sorted_resident_list(residents),
        "monthly_rent": monthly_rent,
        "balances_due": balances_due,
        "utilities_due": utilities_due,
        "deposits_held": deposits_held,
        "total_collected": total_collected,
    })


def property_detail(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    gallery_images = property_obj.images.all()
    active_listings = property_obj.rental_listings.filter(status="published").prefetch_related("photos")
    can_manage_property_blog = user_can_manage_property_blog(request.user, property_obj)

    return render(request, "property_detail.html", {
        "property": property_obj,
        "gallery_images": gallery_images,
        "active_listings": active_listings,
        "can_view_property_blog": False,
        "can_manage_property_blog": can_manage_property_blog,
        "existing_resident_intake_open": property_existing_resident_intake_open(property_obj),
    })


def blog_detail(request, pk):
    post = get_object_or_404(BlogPost, pk=pk)
    return render(request, "blog_detail.html", {"post": post})


def user_can_view_property_blog(user, property_obj):
    if not user.is_authenticated:
        return False

    if user.is_superuser or getattr(user, "role", "") in ["admin", "assistant"]:
        return True

    if property_obj.owner_email and user.email and property_obj.owner_email.lower() == user.email.lower():
        return True

    if (
        getattr(user, "role", "") == "landlord"
        and property_obj.landlord_email
        and user.email
        and property_obj.landlord_email.lower() == user.email.lower()
    ):
        return True

    application = getattr(user, "resident_profile", None)
    return bool(application and application.property_id == property_obj.id)


def user_can_manage_property_blog(user, property_obj):
    if not user.is_authenticated:
        return False

    if user.is_superuser or getattr(user, "role", "") in ["admin", "assistant"]:
        return True

    if property_obj.owner_email and user.email and property_obj.owner_email.lower() == user.email.lower():
        return True

    return bool(
        getattr(user, "role", "") == "landlord"
        and property_obj.landlord_email
        and user.email
        and property_obj.landlord_email.lower() == user.email.lower()
    )


def property_existing_resident_intake_open(property_obj):
    return timezone.now() <= property_obj.created_at + timedelta(days=30)


def clean_match_value(value):
    return "".join(character.lower() for character in str(value or "") if character.isalnum())


def normalize_phone_digits(value):
    return "".join(character for character in str(value or "") if character.isdigit())


def find_current_roster_match(intake):
    entries = CurrentResidentRosterEntry.objects.filter(
        property=intake.property,
        is_active=True,
    )
    intake_email = str(intake.email or "").strip().lower()
    intake_phone = normalize_phone_digits(intake.phone)
    intake_unit = clean_match_value(intake.room_unit_label)
    intake_first = clean_match_value(intake.first_name)
    intake_last = clean_match_value(intake.last_name)

    for entry in entries:
        email_matches = bool(intake_email and entry.email and intake_email == entry.email.strip().lower())
        phone_matches = bool(intake_phone and normalize_phone_digits(entry.phone) == intake_phone)
        unit_matches = bool(intake_unit and clean_match_value(entry.room_unit_label) == intake_unit)
        name_only_roster_entry = not entry.email and not entry.phone and not entry.room_unit_label
        name_matches = (
            clean_match_value(entry.first_name) == intake_first
            and clean_match_value(entry.last_name) == intake_last
        )

        if email_matches or (name_matches and (unit_matches or phone_matches or name_only_roster_entry)):
            return entry

    return None


def current_roster_match_status(intake):
    if find_current_roster_match(intake):
        return "matched"

    if CurrentResidentRosterEntry.objects.filter(property=intake.property, is_active=True).exists():
        return "not_matched"

    return "no_roster"


def ensure_existing_resident_portal_application(intake):
    room_rent_setting = find_room_rent_setting(intake.property, intake.room_unit_label)
    monthly_rent = (
        room_rent_setting.monthly_rent
        if room_rent_setting
        else intake.property.rent_amount or Decimal("0.00")
    )
    rent_due_day = room_rent_setting.rent_due_day if room_rent_setting else 1
    utility_monthly = room_rent_setting.utility_monthly if room_rent_setting else Decimal("0.00")
    deposit_required = room_rent_setting.deposit_required if room_rent_setting else Decimal("0.00")
    deposit_paid = min(room_rent_setting.deposit_paid, deposit_required) if room_rent_setting else Decimal("0.00")

    application = (
        HousingApplication.objects
        .select_related("user")
        .filter(id=intake.application_id)
        .first()
    )

    if not application:
        application = HousingApplication.objects.create(
            property=intake.property,
            full_name=intake.full_name(),
            phone=intake.phone,
            email=intake.email,
            age=0,
            profile_photo=intake.profile_photo,
            space_type="Room",
            space_label=intake.room_unit_label,
            monthly_rent=monthly_rent,
            balance=monthly_rent,
            rent_due_day=rent_due_day,
            deposit_required=deposit_required,
            deposit_paid=deposit_paid,
            utility_monthly=utility_monthly,
            utility_balance=utility_monthly,
            communication_preference="sms" if intake.sms_opted_in else "portal",
            sms_opted_in=intake.sms_opted_in,
            sms_opted_in_at=intake.sms_opted_in_at,
            income_source="Existing resident intake",
            monthly_income=Decimal("0.00"),
            housing_need="Existing resident profile setup.",
            additional_notes=intake.additional_notes,
        )
        intake.application = application
        intake.save(update_fields=["application"])
    else:
        update_fields = []

        if intake.room_unit_label and application.space_label != intake.room_unit_label:
            application.space_type = application.space_type or "Room"
            application.space_label = intake.room_unit_label
            update_fields.extend(["space_type", "space_label"])

        if intake.sms_opted_in and not application.sms_opted_in:
            application.sms_opted_in = True
            application.sms_opted_in_at = intake.sms_opted_in_at or timezone.now()
            application.communication_preference = "sms"
            update_fields.extend(["sms_opted_in", "sms_opted_in_at", "communication_preference"])

        if update_fields:
            application.save(update_fields=update_fields)

    if not application.user:
        application.user = create_pending_portal_user(
            intake.full_name(),
            intake.email,
            "tenant",
            application.id,
        )
        application.save(update_fields=["user"])

    ensure_existing_resident_onboarding_documents(application)

    return application


def ensure_existing_resident_onboarding_documents(application):
    document_specs = [
        ("lease", "Resident Lease Agreement"),
        ("emergency_contact", "Emergency Contact Sheet"),
        ("painted_lady_acknowledgment", "Who We Are / Painted Lady Acknowledgment"),
    ]

    for document_type, title in document_specs:
        document, _ = SignedDocument.objects.get_or_create(
            application=application,
            document_type=document_type,
            defaults={
                "title": title,
                "lease_sent_date": timezone.localdate(),
                "landlord_name": "Michael Bowling",
                "landlord_signature": "Michael Bowling",
            },
        )

        if not document.locked:
            document.title = title
            document.lease_sent_date = document.lease_sent_date or timezone.localdate()
            document.landlord_name = document.landlord_name or "Michael Bowling"
            document.landlord_signature = document.landlord_signature or "Michael Bowling"
            document.save()


def send_existing_resident_portal_invite(request, intake, allow_roster_override=False):
    if current_roster_match_status(intake) != "matched" and not allow_roster_override:
        messages.warning(
            request,
            "No setup invite was sent and no resident file was created because this profile setup does not match the approved current resident roster.",
        )
        return None

    if current_roster_match_status(intake) != "matched" and allow_roster_override:
        messages.warning(
            request,
            "Roster match was manually overridden by staff. Confirm this resident belongs to the property before sharing the setup code.",
        )

    application = ensure_existing_resident_portal_application(intake)

    if application.user.has_usable_password():
        messages.info(request, "This resident already has a registered portal login.")
        return application

    application.user.refresh_invite_code()

    try:
        from .landlord_views import send_resident_invite_email
        send_resident_invite_email(application)
    except Exception as exc:
        messages.warning(
            request,
            f"Resident profile was saved and a setup code was created, but email failed: {exc}",
        )
    else:
        messages.success(request, "Resident profile saved and portal setup email sent.")

    return application


def roster_value(row, *keys):
    for key in keys:
        value = row.get(key)
        if value:
            return str(value).strip()
    return ""


ROSTER_HEADER_ALIASES = {
    "first name",
    "firstname",
    "first",
    "last name",
    "lastname",
    "last",
    "name",
    "resident",
    "resident name",
    "tenant",
    "email",
    "email address",
    "phone",
    "phone number",
    "mobile",
    "cell",
    "room unit label",
    "room",
    "unit",
    "room/unit",
    "room number",
    "unit number",
}


def has_roster_headers(headers):
    return any(header in ROSTER_HEADER_ALIASES for header in headers)


def import_headerless_roster_rows(property_obj, rows, user):
    row_iterable = []
    for row in rows:
        values = [str(value or "").strip() for value in row if str(value or "").strip()]
        if not values:
            continue

        if len(values) == 1:
            row_iterable.append({"name": values[0]})
        else:
            first_value = values[0]
            second_value = values[1]
            if len(first_value) <= 10 and len(second_value.split()) >= 2:
                row_iterable.append({"room": first_value, "name": second_value})
            else:
                row_iterable.append({"name": first_value, "room": second_value})

    return import_current_resident_roster_rows(property_obj, row_iterable, user)


def import_current_resident_roster(property_obj, file_obj, user):
    filename = (getattr(file_obj, "name", "") or "").lower()
    if filename.endswith((".xlsx", ".xls")):
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return 0, 0

        headers = [normalized_header(value) for value in header_row]
        if not has_roster_headers(headers):
            return import_headerless_roster_rows(property_obj, [header_row, *rows], user)

        row_iterable = (
            {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers)
            }
            for row in rows
        )
        return import_current_resident_roster_rows(property_obj, row_iterable, user)

    decoded_file = TextIOWrapper(file_obj, encoding="utf-8-sig", newline="")
    reader = csv.reader(decoded_file)
    rows = list(reader)
    if not rows:
        return 0, 0

    headers = [normalized_header(value) for value in rows[0]]
    if not has_roster_headers(headers):
        return import_headerless_roster_rows(property_obj, rows, user)

    row_iterable = (
        {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers)
        }
        for row in rows[1:]
    )
    return import_current_resident_roster_rows(property_obj, row_iterable, user)


def import_current_resident_roster_rows(property_obj, row_iterable, user):
    created = 0
    skipped = 0
    active_entry_ids = []

    for normalized_row in row_iterable:
        first_name = roster_value(normalized_row, "first name", "firstname", "first")
        last_name = roster_value(normalized_row, "last name", "lastname", "last")
        full_name = roster_value(normalized_row, "name", "resident", "resident name", "tenant")

        if (not first_name or not last_name) and full_name:
            parts = full_name.split()
            first_name = first_name or (parts[0] if parts else "")
            last_name = last_name or (" ".join(parts[1:]) if len(parts) > 1 else "")

        email = roster_value(normalized_row, "email", "email address")
        phone = roster_value(normalized_row, "phone", "phone number", "mobile", "cell")
        room_unit_label = roster_value(normalized_row, "room unit label", "room", "unit", "room/unit", "room number", "unit number")

        if not first_name or not last_name:
            skipped += 1
            continue

        roster_entry, _created = CurrentResidentRosterEntry.objects.update_or_create(
            property=property_obj,
            first_name=first_name,
            last_name=last_name,
            email=email,
            room_unit_label=room_unit_label,
            defaults={
                "phone": phone,
                "is_active": True,
                "uploaded_by": user,
            },
        )
        active_entry_ids.append(roster_entry.id)
        created += 1

    if active_entry_ids:
        CurrentResidentRosterEntry.objects.filter(property=property_obj).exclude(id__in=active_entry_ids).update(is_active=False)

    return created, skipped


def existing_resident_intake(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)

    if not property_existing_resident_intake_open(property_obj):
        messages.info(request, "Existing resident profile intake has closed for this property.")
        return redirect("property_detail", pk=property_obj.id)

    form = ExistingResidentIntakeForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        intake = form.save(commit=False)
        intake.property = property_obj
        if intake.sms_opted_in:
            intake.sms_opted_in_at = timezone.now()
        intake.save()

        if find_current_roster_match(intake):
            send_existing_resident_portal_invite(request, intake)
        else:
            messages.warning(
                request,
                "Profile saved for landlord review. No automatic invite was sent because this submission did not match the approved current resident list.",
            )

        return redirect("existing_resident_intake_success", pk=property_obj.id)

    return render(request, "existing_resident_intake.html", {
        "form": form,
        "property": property_obj,
    })


def existing_resident_intake_success(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    return render(request, "existing_resident_intake_success.html", {"property": property_obj})


@login_required
@user_passes_test(staff_required)
def landlord_existing_resident_intake_detail(request, intake_id):
    intake = get_object_or_404(
        ExistingResidentIntake.objects.select_related("property"),
        id=intake_id,
        property__in=staff_managed_properties(request.user),
    )
    if not intake.landlord_reviewed_at:
        intake.landlord_reviewed_at = timezone.now()
        intake.save(update_fields=["landlord_reviewed_at"])

    application = (
        HousingApplication.objects
        .select_related("user", "property")
        .filter(id=intake.application_id)
        .first()
    )
    pending_user = application.user if application and application.user else None

    if pending_user and pending_user.has_usable_password():
        setup_status = "completed"
    elif pending_user:
        setup_status = "invite_sent"
    else:
        setup_status = "ready"

    return render(request, "landlord_existing_resident_intake_detail.html", {
        "intake": intake,
        "application": application,
        "pending_user": pending_user,
        "setup_status": setup_status,
        "roster_match_status": current_roster_match_status(intake),
        "roster_match": find_current_roster_match(intake),
    })


@login_required
@user_passes_test(staff_required)
def current_resident_roster_upload(request):
    properties = staff_managed_properties(request.user).order_by("name")
    form = CurrentResidentRosterUploadForm(request.POST or None, request.FILES or None, properties=properties)

    if request.method == "POST" and form.is_valid():
        property_obj = form.cleaned_data["property"]
        file_obj = form.cleaned_data["file"]
        created, skipped = import_current_resident_roster(property_obj, file_obj, request.user)
        messages.success(request, f"Current resident list imported for {property_obj.name}. {created} rows saved, {skipped} rows skipped.")
        return redirect("current_resident_roster_upload")

    roster_entries = (
        CurrentResidentRosterEntry.objects
        .filter(property__in=properties, is_active=True)
        .select_related("property")
        .order_by("property__name", "room_unit_label", "last_name")
    )

    return render(request, "current_resident_roster_upload.html", {
        "form": form,
        "roster_entries": roster_entries,
    })


@login_required
@user_passes_test(staff_required)
def landlord_send_existing_resident_invite(request, intake_id):
    if request.method != "POST":
        return redirect("landlord_attention")

    intake = get_object_or_404(
        ExistingResidentIntake.objects.select_related("property"),
        id=intake_id,
        property__in=staff_managed_properties(request.user),
    )
    allow_roster_override = request.POST.get("allow_roster_override") == "on"
    application = send_existing_resident_portal_invite(
        request,
        intake,
        allow_roster_override=allow_roster_override,
    )

    if application and application.user and not application.user.has_usable_password():
        messages.info(request, f"Backup resident setup code: {application.user.invite_code}")

    return redirect("landlord_attention")


@login_required
@user_passes_test(staff_required)
def delete_existing_resident_intake(request, intake_id):
    if request.method != "POST":
        return redirect("landlord_attention")

    intake = get_object_or_404(
        ExistingResidentIntake.objects.select_related("property"),
        id=intake_id,
        property__in=staff_managed_properties(request.user),
    )
    application = HousingApplication.objects.select_related("user").filter(id=intake.application_id).first()

    if application and application.user and application.user.has_usable_password():
        messages.error(
            request,
            "This setup attempt is connected to a completed resident login. Open the resident file before deleting anything.",
        )
        return redirect("landlord_existing_resident_intake_detail", intake_id=intake.id)

    if application and (
        application.payments.exists()
        or application.documents.exists()
        or application.signed_documents.filter(locked=True).exists()
        or application.resident_messages.exists()
    ):
        messages.error(
            request,
            "This setup attempt is connected to resident activity. Open the resident file before deleting anything.",
        )
        return redirect("landlord_existing_resident_intake_detail", intake_id=intake.id)

    resident_name = intake.full_name()
    if application:
        temp_user = application.user
        application.delete()
        if temp_user and not temp_user.has_usable_password():
            temp_user.delete()

    intake.delete()
    messages.success(request, f"Deleted invalid current resident setup attempt for {resident_name}.")
    return redirect("landlord_attention")


def add_blog_comment(request, post_id):
    post = get_object_or_404(BlogPost, id=post_id)

    if post.property and not user_can_view_property_blog(request.user, post.property):
        return redirect(f"{reverse('login')}?next={reverse('property_detail', args=[post.property.id])}")

    if request.method == "POST":
        form = BlogCommentForm(request.POST)

        if form.is_valid():
            comment = form.save(commit=False)
            comment.post = post
            comment.approved = False
            comment.save()

    if post.property:
        return redirect("property_detail", pk=post.property.id)

    return redirect("home")


def printable_application(request, pk):
    application = get_object_or_404(HousingApplication, pk=pk)
    if (
        request.user.is_authenticated
        and staff_required(request.user)
        and application.property_id in staff_managed_properties(request.user).values_list("id", flat=True)
        and not application.landlord_reviewed_at
    ):
        application.landlord_reviewed_at = timezone.now()
        application.save(update_fields=["landlord_reviewed_at"])
    return render(request, "printable_application.html", {"application": application})


@login_required
@user_passes_test(staff_required)
def application_screening_review(request, pk):
    application = get_object_or_404(
        HousingApplication.objects.select_related("property"),
        pk=pk,
        property__in=staff_managed_properties(request.user),
    )
    suggested_score, suggested_rating, score_factors = calculate_screening_score(application)
    initial = {}

    if application.screening_score is None:
        initial["screening_score"] = suggested_score
    if application.screening_rating == "unrated":
        initial["screening_rating"] = suggested_rating
    if not application.screening_review_summary:
        initial["screening_review_summary"] = "\n".join(score_factors)

    old_report_name = application.background_report.name if application.background_report else ""
    old_decision = application.owner_final_decision
    form = ScreeningReviewForm(request.POST or None, request.FILES or None, instance=application, initial=initial)

    if request.method == "POST" and form.is_valid():
        application = form.save(commit=False)

        if application.background_report and application.background_report.name != old_report_name:
            application.background_report_received_at = timezone.now()
            if application.background_check_status in ["pending", "ordered", "not_required"]:
                application.background_check_status = "needs_review"

        if application.owner_final_decision != "pending" and application.owner_final_decision != old_decision:
            application.owner_decision_at = timezone.now()

        application.save()

        if application.background_report:
            ApplicantDocument.objects.get_or_create(
                application=application,
                document_type="background_report",
                file=application.background_report.name,
                defaults={
                    "name": f"Background Screening Report - {application.full_name}",
                    "status": "locked",
                },
            )

        messages.success(request, "Screening review saved.")
        return redirect("application_screening_review", pk=application.id)

    adverse_notices = application.adverse_action_notices.all()
    return render(request, "application_screening_review.html", {
        "application": application,
        "form": form,
        "suggested_score": suggested_score,
        "suggested_rating": suggested_rating,
        "score_factors": score_factors,
        "adverse_notices": adverse_notices,
    })


@login_required
@user_passes_test(staff_required)
def create_adverse_action_notice(request, pk):
    application = get_object_or_404(
        HousingApplication.objects.select_related("property"),
        pk=pk,
        property__in=staff_managed_properties(request.user),
    )
    property_obj = application.property
    default_owner_contact = ""
    if property_obj:
        default_owner_contact = "\n".join(filter(None, [property_obj.owner_email, property_obj.landlord_email]))

    initial = {
        "screening_company_name": application.screening_provider_name or (property_obj.screening_provider_name if property_obj else ""),
        "owner_landlord_name": property_obj.name if property_obj else "",
        "owner_landlord_contact": default_owner_contact,
        "notice_body": (
            f"Dear {application.full_name},\n\n"
            "After reviewing your rental application and screening information, the property owner or landlord has taken the action selected above.\n\n"
            "If this decision was based in whole or in part on a consumer report or background screening report, you have rights under the Fair Credit Reporting Act. "
            "The screening company did not make the rental decision and cannot explain the specific reason for the decision. "
            "You may request a free copy of the report from the screening company and dispute inaccurate or incomplete information.\n\n"
            "Sincerely,\n"
            "Property Owner / Landlord"
        ),
    }
    form = AdverseActionNoticeForm(request.POST or None, initial=initial)

    if request.method == "POST" and form.is_valid():
        notice = form.save(commit=False)
        notice.application = application
        notice.created_by = request.user
        notice.save()

        application.owner_final_decision = (
            "approved_conditions" if notice.action_type == "approved_conditions" else "declined"
        )
        application.owner_decision_at = timezone.now()
        application.owner_decision_notes = notice.reasons
        application.save(update_fields=["owner_final_decision", "owner_decision_at", "owner_decision_notes"])

        messages.success(request, "Adverse action notice drafted.")
        return redirect("adverse_action_notice_detail", notice_id=notice.id)

    return render(request, "adverse_action_notice_form.html", {
        "application": application,
        "form": form,
    })


@login_required
@user_passes_test(staff_required)
def adverse_action_notice_detail(request, notice_id):
    notice = get_object_or_404(
        AdverseActionNotice.objects.select_related("application", "application__property"),
        id=notice_id,
        application__property__in=staff_managed_properties(request.user),
    )
    return render(request, "adverse_action_notice_detail.html", {"notice": notice})


def get_resident_signed_document(request, document_id):
    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return None

    return get_object_or_404(
        SignedDocument,
        id=document_id,
        application=application,
    )


@login_required
def lease_sign(request):

    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    signed_document = SignedDocument.objects.filter(
        application=application,
        document_type="lease",
    ).first()

    if not signed_document:
        signed_document = SignedDocument.objects.create(
            application=application,
            document_type="lease",
            title="Resident Lease Agreement",
            lease_sent_date=timezone.localdate(),
        )

    return render(request, "lease_sign.html", {
        "application": application,
        "signed_document": signed_document,
    })


@login_required
def onboarding_document(request, document_id):
    signed_document = get_resident_signed_document(request, document_id)

    if not signed_document:
        return redirect("tenant_dashboard")

    template_name = "lease_sign.html"

    if signed_document.document_type != "lease":
        template_name = "onboarding_document_sign.html"

    return render(request, template_name, {
        "application": signed_document.application,
        "signed_document": signed_document,
    })


@login_required
def submit_lease_signature(request, document_id=None):

    if request.method != "POST":
        return redirect("tenant_dashboard")

    application = getattr(request.user, "resident_profile", None)

    if not application:
        messages.error(request, "No resident file connected.")
        return redirect("tenant_dashboard")

    if document_id:
        signed_document = get_object_or_404(
            SignedDocument,
            id=document_id,
            application=application,
            document_type="lease",
        )
    else:
        signed_document = SignedDocument.objects.filter(
            application=application,
            document_type="lease",
        ).first()

    if not signed_document:
        messages.error(request, "Lease document not found.")
        return redirect("tenant_dashboard")

    if signed_document.locked:
        messages.info(request, "This lease has already been signed.")
        return redirect("tenant_dashboard")

    signed_document.rent_initials = request.POST.get("rent_initials", "").strip()
    signed_document.sobriety_initials = request.POST.get("sobriety_initials", "").strip()
    signed_document.testing_initials = request.POST.get("testing_initials", "").strip()
    signed_document.guest_policy_initials = request.POST.get("guest_policy_initials", "").strip()
    signed_document.cleanliness_initials = request.POST.get("cleanliness_initials", "").strip()
    signed_document.disclosure_initials = request.POST.get("disclosure_initials", "").strip()

    signed_document.resident_signature = request.POST.get("resident_signature", "").strip()

    signed_document.signature_agreement = bool(
        request.POST.get("signature_agreement")
    )

    if not signed_document.resident_signature:
        messages.error(request, "Signature is required.")
        return redirect("lease_sign")

    if not signed_document.signature_agreement:
        messages.error(request, "You must agree to electronically sign.")
        return redirect("lease_sign")

    signed_document.signed_at = timezone.now()
    signed_document.locked = True

    signed_document.save()

    messages.success(
        request,
        "Lease agreement successfully signed and filed."
    )

    return redirect("tenant_dashboard")


@login_required
def submit_onboarding_document(request, document_id):
    if request.method != "POST":
        return redirect("tenant_dashboard")

    signed_document = get_resident_signed_document(request, document_id)

    if not signed_document:
        return redirect("tenant_dashboard")

    if signed_document.locked:
        messages.info(request, "This document has already been signed and filed.")
        return redirect("tenant_dashboard")

    if signed_document.document_type == "lease":
        return submit_lease_signature(request, document_id=signed_document.id)

    signed_document.emergency_contact_name = request.POST.get("emergency_contact_name", "").strip()
    signed_document.emergency_contact_phone = request.POST.get("emergency_contact_phone", "").strip()
    signed_document.emergency_contact_relationship = request.POST.get("emergency_contact_relationship", "").strip()
    signed_document.emergency_medical_notes = request.POST.get("emergency_medical_notes", "").strip()
    signed_document.resident_signature = request.POST.get("resident_signature", "").strip()
    signed_document.signature_agreement = bool(request.POST.get("signature_agreement"))

    if signed_document.document_type == "emergency_contact":
        if not signed_document.emergency_contact_name or not signed_document.emergency_contact_phone:
            messages.error(request, "Emergency contact name and phone are required.")
            return redirect("onboarding_document", document_id=signed_document.id)

    if not signed_document.resident_signature:
        messages.error(request, "Signature is required.")
        return redirect("onboarding_document", document_id=signed_document.id)

    if not signed_document.signature_agreement:
        messages.error(request, "You must agree to electronically sign.")
        return redirect("onboarding_document", document_id=signed_document.id)

    signed_document.signed_at = timezone.now()
    signed_document.locked = True
    signed_document.save()

    messages.success(request, "Document signed and filed.")
    return redirect("tenant_dashboard")


def create_checkout_session(request, application_id, payment_type="rent"):
    application = get_object_or_404(HousingApplication, id=application_id)

    if not staff_required(request.user):
        user_application = getattr(request.user, "resident_profile", None)
        session_application_id = request.session.get("submitted_application_id")
        is_session_fee_payment = (
            payment_type in ["application_fee", "background_check_fee"]
            and session_application_id == application.id
        )

        if (not user_application or user_application.id != application.id) and not is_session_fee_payment:
            return JsonResponse({"error": "You are not authorized to pay this account."}, status=403)

    stale_before = timezone.now() - timedelta(minutes=30)
    Payment.objects.filter(
        application=application,
        payment_type=payment_type,
        status="pending",
        created_at__lt=stale_before,
    ).update(status="failed")

    existing_pending = Payment.objects.filter(
        application=application,
        payment_type=payment_type,
        status="pending",
    ).exists()

    if existing_pending:
        return JsonResponse({
            "error": "A payment is already pending. Please wait before trying again."
        })

    if payment_type == "rent" and application.balance <= 0:
        return JsonResponse({
            "error": "No rent balance due."
        })

    amount = Decimal("0.00")
    description = ""

    if payment_type == "rent":
        amount = application.balance if application.balance > 0 else application.monthly_rent
        description = "Rent Payment"

    elif payment_type == "deposit":
        amount = max(application.deposit_required - application.deposit_paid, Decimal("0.00"))
        description = "Deposit Payment"

    elif payment_type == "utility":
        amount = application.utility_balance if application.utility_balance > 0 else application.utility_monthly
        description = "Utility Payment"

    elif payment_type == "application_fee":
        amount = max(application.application_fee_amount - application.application_fee_paid, Decimal("0.00"))
        description = "Application Fee"

    elif payment_type == "background_check_fee":
        amount = max(application.background_check_fee_amount - application.background_check_fee_paid, Decimal("0.00"))
        description = "Background Check Fee"

    elif payment_type == "total":
        rent_due = application.balance if application.balance > 0 else Decimal("0.00")
        deposit_due = max(application.deposit_required - application.deposit_paid, Decimal("0.00"))
        utility_due = application.utility_balance if application.utility_balance > 0 else Decimal("0.00")

        amount = rent_due + deposit_due + utility_due
        description = "Combined Payment - Total Due"
        payment_type = "other"

    else:
        return JsonResponse({"error": "Invalid payment type"})

    if amount <= 0:
        return JsonResponse({"error": "No balance due"})

    payment = Payment.objects.create(
        application=application,
        payment_type=payment_type,
        payment_method="stripe_card",
        description=description,
        amount=amount,
        status="pending",
    )

    if getattr(settings, "DEMO_MODE", False):
        payment.payment_method = "other"
        payment.status = "completed"
        payment.description = f"Demo payment - {description}"
        payment.received_at = timezone.now()
        payment.service_month = timezone.localdate().replace(day=1)
        payment.save(update_fields=["payment_method", "status", "description", "received_at", "service_month"])
        apply_completed_payment_to_balance(payment)
        messages.success(request, "Demo payment recorded. No real card or bank transaction was processed.")
        return redirect("payment_success")

    session = stripe.checkout.Session.create(
        payment_method_types=["card", "cashapp"],
        mode="payment",
        line_items=[{
            "price_data": {
                "currency": "usd",
                "product_data": {"name": description},
                "unit_amount": int(amount * 100),
            },
            "quantity": 1,
        }],
        success_url=request.build_absolute_uri("/payment-success/"),
        cancel_url=request.build_absolute_uri("/tenant-dashboard/"),
        metadata={"payment_id": str(payment.id)},
    )

    payment.stripe_session_id = session.id
    payment.save()

    return redirect(session.url)


def payment_success(request):
    return render(request, "payment_success.html")


@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, settings.STRIPE_WEBHOOK_SECRET)
    except Exception:
        return HttpResponse(status=400)

    if event["type"] != "checkout.session.completed":
        return HttpResponse(status=200)

    session = event["data"]["object"]
    payment_id = session["metadata"]["payment_id"]

    if not payment_id:
        return HttpResponse(status=200)

    payment = Payment.objects.filter(id=payment_id).first()

    if not payment or payment.status == "completed":
        return HttpResponse(status=200)

    payment.status = "completed"
    payment.stripe_payment_intent = session["payment_intent"]
    payment.save()

    payment_method_types = session.get("payment_method_types", [])
    if "cashapp" in payment_method_types:
        payment.payment_method = "stripe_cashapp"
        payment.save()

    apply_completed_payment_to_balance(payment)

    application = payment.application
    owner_email = "BowlingLegacyLLC@outlook.com"

    if application.property and application.property.owner_email:
        owner_email = application.property.owner_email

    send_mail(
        subject="Resident Payment Received",
        message=f"""
Resident: {application.full_name}

Payment Type:
{payment.get_payment_type_display()}

Amount:
${payment.amount}
""",
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        recipient_list=[owner_email],
        fail_silently=True,
    )

    return HttpResponse(status=200)
